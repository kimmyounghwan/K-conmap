# -*- coding: utf-8 -*-
"""
tools/checkbook.py — 설계변경 통합 엑셀이 «정말로 계산되는지» 확인합니다. (2026-09-05)

수식을 눈으로 읽는 것은 확인이 아닙니다. 그래서:
  ① 시험 자료를 넣은 사본을 만들고
  ② LibreOffice 로 실제로 열어 재계산시킨 뒤
  ③ 파이썬으로 «따로» 계산한 정답과 칸 하나하나 대조합니다.

한 군데라도 어긋나면 빨갛게 찍고 1 을 돌려줍니다.

쓰는 법:  python tools/checkbook.py
"""
import os, shutil, subprocess, sys, tempfile
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, "web", "public", "forms", "chg-tonghap.xlsx")
HEAD = 4

# ── 시험 자료 ────────────────────────────────────────────────
KUM, PLN, RATE = 100_000_000, 114_000_000, 87.7193
GEN, PRF, VAT = 6.0, 15.0, 10.0
MASTER = [           # 코드, 품명, 규격, 단위, 재료, 노무, 경비
    ("토공-001", "터파기", "기계", "m3", 0, 8000, 2000),
    ("토공-002", "되메우기", "다짐", "m3", 0, 5000, 1000),
    ("포장-001", "아스콘포장", "T=5cm", "m2", 12000, 3000, 1000),
    ("신규-001", "배수관 부설", "D300", "m", 20000, 6000, 2000),
]
SUR = [              # 구분, 코드, 가로, 세로, 높이, 개소
    ("당초", "토공-001", 10, 5, 2, 1),
    ("당초", "토공-002", 10, 5, 1, 1),
    ("당초", "포장-001", 20, 10, None, None),
    ("변경", "토공-001", 10, 5, 3, 1),
    ("변경", "토공-002", 10, 4, 1, 1),
    ("변경", "포장-001", 20, 10, None, None),
    ("변경", "신규-001", 30, None, None, None),
]
DANG = ["토공-001", "토공-002", "포장-001", None]
BYEON = ["토공-001", "토공-002", "포장-001", "신규-001"]


def fill(path):
    wb = load_workbook(path)
    c = wb["설정"]
    c["C13"], c["C14"], c["C15"] = KUM, PLN, RATE
    c["C19"], c["C20"], c["C21"] = GEN, PRF, VAT
    m = wb["단가마스터"]
    for i, row in enumerate(MASTER):
        for j, v in enumerate(row):
            m.cell(HEAD + i, j + 1, v)
    s = wb["수량산출서"]
    for i, (g, code, a, b, h, n) in enumerate(SUR):
        r = HEAD + i
        s.cell(r, 1, g); s.cell(r, 2, code)
        for j, v in enumerate((a, b, h, n)):
            if v is not None:
                s.cell(r, 4 + j, v)
    for name, codes in (("당초내역", DANG), ("변경내역", BYEON)):
        w = wb[name]
        for i, code in enumerate(codes):
            if code:
                w.cell(HEAD + i, 1, i + 1)
                w.cell(HEAD + i, 2, code)
    g = wb["공종별집계"]
    for i, code in enumerate(["토공", "포장", "신규"]):
        g.cell(HEAD + i, 1, code)
    wb.save(path)


RECALC_XCU = """<?xml version="1.0" encoding="UTF-8"?>
<oor:items xmlns:oor="http://openoffice.org/2001/registry"
           xmlns:xs="http://www.w3.org/2001/XMLSchema">
 <item oor:path="/org.openoffice.Office.Calc/Formula/Load">
  <prop oor:name="OOXMLRecalcMode" oor:op="fuse"><value>0</value></prop>
 </item>
 <item oor:path="/org.openoffice.Office.Calc/Formula/Load">
  <prop oor:name="ODFRecalcMode" oor:op="fuse"><value>0</value></prop>
 </item>
</oor:items>
"""


def recalc(path, outdir):
    """LibreOffice 로 실제 열어 계산시킵니다 (수식을 «읽는» 것이 아니라 «돌립니다»).

    ⚠️ 기본 설정은 «엑셀 파일을 열 때 다시 계산하지 않음» 입니다.
       그대로 두면 값이 전부 비어 나와 «확인» 이 아니라 «아무것도 못 본 것» 이 됩니다.
       그래서 임시 프로필에 «항상 다시 계산» 을 넣고, 원본과 «다른 폴더» 로 내보냅니다."""
    prof = os.path.join(outdir, "loprof")
    d = os.path.join(prof, "user")
    os.makedirs(d, exist_ok=True)
    with open(os.path.join(d, "registrymodifications.xcu"), "w", encoding="utf-8") as f:
        f.write(RECALC_XCU)
    dst = os.path.join(outdir, "out")
    os.makedirs(dst, exist_ok=True)
    env = dict(os.environ, HOME=outdir)
    r = subprocess.run(["soffice", "-env:UserInstallation=file://" + prof,
                        "--headless", "--norestore", "--nolockcheck",
                        "--convert-to", "xlsx", "--outdir", dst, path],
                       capture_output=True, text=True, timeout=300, env=env)
    out = os.path.join(dst, os.path.basename(path))
    if not os.path.exists(out):
        print("⛔ LibreOffice 변환 실패:", r.stdout[-400:], r.stderr[-400:])
        sys.exit(2)
    return out


# ── 파이썬으로 «따로» 계산한 정답 ─────────────────────────────
def expected():
    up = {c: (mat, lab, exp) for c, _, _, _, mat, lab, exp in MASTER}
    tot = {c: sum(up[c]) for c in up}
    q = {}
    for gu, code, a, b, h, n in SUR:
        v = 1.0
        for x in (a, b, h, n):
            v *= (x if x is not None else 1)
        q[(gu, code)] = q.get((gu, code), 0) + v
    e = {}
    for name, gu, codes in (("당초내역", "당초", DANG), ("변경내역", "변경", BYEON)):
        amt = 0
        for code in codes:
            if code:
                amt += round(q.get((gu, code), 0) * tot[code])
        e[name + "합계"] = amt
    rt = RATE / 100.0
    jg, rows = 0, []
    for i, code in enumerate(BYEON):
        d = DANG[i] if i < len(DANG) else None
        q0 = q.get(("당초", d), 0) if d else 0
        q1 = q.get(("변경", code), 0)
        dq = q1 - q0
        if not d:
            kind, unit = "신규", round(tot[code] * rt)
        elif dq < 0:
            kind, unit = "감소", tot[d]
        elif dq > 0:
            kind, unit = "증가", tot[d]
        else:
            kind, unit = "-", 0
        amt = round(dq * unit)
        rows.append((code, kind, unit, amt))
        jg += amt
    e["증감합계"] = jg
    e["변경후"] = KUM + jg
    e["증감줄"] = rows
    for name, gu, codes in (("당초", "당초", DANG), ("변경", "변경", BYEON)):
        mat = lab = exp = 0
        for code in codes:
            if not code:
                continue
            v = q.get((gu, code), 0)
            mat += v * up[code][0]; lab += v * up[code][1]; exp += v * up[code][2]
        net = mat + lab + exp
        gen = round(net * GEN / 100)
        prf = round((lab + exp + gen) * PRF / 100)
        sup = net + gen + prf
        vat = round(sup * VAT / 100)
        e[name + "재료비"], e[name + "노무비"], e[name + "경비"] = mat, lab, exp
        e[name + "순공사원가"], e[name + "일반관리비"], e[name + "이윤"] = net, gen, prf
        e[name + "공급가액"], e[name + "부가세"], e[name + "총액"] = sup, vat, sup + vat
    return e


def main():
    if not os.path.exists(SRC):
        print("⛔ 먼저 python chgbook.py 를 돌리세요."); sys.exit(2)
    tmp = tempfile.mkdtemp(prefix="chgbook_")
    work = os.path.join(tmp, "시험.xlsx")
    shutil.copy(SRC, work)
    fill(work)
    print("① 시험 자료를 넣었습니다 (단가 4가지 · 수량 7줄 · 당초 3 · 변경 4)")
    out = recalc(work, tmp)
    print("② LibreOffice 로 실제 재계산했습니다")
    wb = load_workbook(out, data_only=True)
    e = expected()
    jg, wg, ck = wb["증감대비표"], wb["원가계산서"], wb["검증시트"]
    dn, bn = wb["당초내역"], wb["변경내역"]

    def num(v):
        try: return float(v)
        except Exception: return None

    bad = []
    def cmp(label, got, want, tol=1.0):
        g = num(got)
        ok = g is not None and abs(g - want) <= tol
        print("   %s %-34s 화면 %-16s 정답 %-16s"
              % ("✅" if ok else "❌", label,
                 ("%.0f" % g) if g is not None else str(got), "%.0f" % want))
        if not ok:
            bad.append(label)

    print("③ 대조 — 내역 합계")
    cmp("당초내역 합계", dn.cell(HEAD + 200, 13).value, e["당초내역합계"])
    cmp("변경내역 합계", bn.cell(HEAD + 200, 13).value, e["변경내역합계"])

    print("③ 대조 — 증감대비표 (규정 단가)")
    KIND = {}
    for i, (code, kind, unit, amt) in enumerate(e["증감줄"]):
        r = HEAD + i
        got_kind = jg.cell(r, 12).value
        okk = (got_kind == kind)
        print("   %s %-12s 구분 %-4s (정답 %-4s)" % ("✅" if okk else "❌", code, got_kind, kind))
        if not okk:
            bad.append(code + " 구분")
        cmp("  %s 적용단가" % code, jg.cell(r, 15).value, unit)
        cmp("  %s 증감금액" % code, jg.cell(r, 16).value, amt)
    cmp("증감 합계", jg.cell(HEAD + 203, 16).value, e["증감합계"])
    cmp("변경 후 계약금액", jg.cell(HEAD + 204, 16).value, e["변경후"])

    print("③ 대조 — 원가계산서")
    ROW = {"재료비": 4, "노무비": 5, "경비": 6, "순공사원가": 7,
           "일반관리비": 8, "이윤": 9, "공급가액": 10, "부가세": 11, "총액": 12}
    for k, r in ROW.items():
        cmp("당초 " + k, wg.cell(r, 4).value, e["당초" + k])
        cmp("변경 " + k, wg.cell(r, 5).value, e["변경" + k])

    print("③ 대조 — 검증시트 (전부 ✅ 여야 합니다)")
    warn = []
    for r in range(HEAD, HEAD + 12):
        what, v = ck.cell(r, 2).value, ck.cell(r, 4).value
        if v == "⚠️":
            warn.append(what)
        print("   %s %s = %s" % (v, what, ck.cell(r, 3).value))
    # 계약금액 차이 검사는 시험 자료가 계약금액과 안 맞으므로 «걸리는 것이 정상» 입니다
    warn = [w for w in warn if "계약금액 차이" not in str(w)]
    if warn:
        bad.extend(warn)

    print()
    if bad:
        print("❌ 어긋난 곳 %d 가지: %s" % (len(bad), ", ".join(str(x) for x in bad)))
        sys.exit(1)
    print("✅ 전부 일치합니다 — 수식이 실제로 돌아갑니다.")
    print("   (시험 파일: %s)" % out)


# ── 반대 시험 — 일부러 틀리게 해서 «검증시트가 잡는지» 봅니다 ──────
#    (잡지 못하는 검사는 있으나 마나입니다. CLAUDE.md 의 원칙)
def negative():
    tmp = tempfile.mkdtemp(prefix="chgbook_neg_")
    work = os.path.join(tmp, "반대시험.xlsx")
    shutil.copy(SRC, work)
    fill(work)
    wb = load_workbook(work)
    wb["단가마스터"].cell(HEAD + 4, 1, "토공-001")        # 코드 중복
    wb["변경내역"].cell(HEAD + 1, 2, "포장-001")          # 줄 어긋남
    wb["당초내역"].cell(HEAD + 2, 7, -5)                 # 수량 음수(직접)
    wb["설정"]["C15"] = 0                                # 낙찰률 비움
    wb.save(work)
    out = recalc(work, tmp)
    ck = load_workbook(out, data_only=True)["검증시트"]
    want = {1: "낙찰률", 4: "줄이 서로 어긋난", 5: "코드가 겹치는", 6: "수량이 음수"}
    bad = []
    print("\n④ 반대 시험 — 일부러 4가지를 틀리게 넣었습니다")
    for i, key in want.items():
        r = HEAD + i - 1
        v, what = ck.cell(r, 4).value, str(ck.cell(r, 2).value)
        ok = (v == "⚠️") and (key in what)
        print("   %s %-34s %s (값 %s)" % ("✅잡음" if ok else "❌못잡음", what, v, ck.cell(r, 3).value))
        if not ok:
            bad.append(what)
    if bad:
        print("❌ 검증시트가 못 잡은 것: %s" % ", ".join(bad))
        return 1
    print("✅ 넣어 둔 4가지를 검증시트가 전부 잡았습니다.")
    return 0


if __name__ == "__main__":
    if "--negative" in sys.argv:
        sys.exit(negative())
    main()
    sys.exit(negative())

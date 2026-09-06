# -*- coding: utf-8 -*-
"""
collect.py — 조달청 나라장터에서 최근 공고·개찰 결과를 받아온다.

기존 auto_collector.py 와 다른 점
  · Firebase 에 쓰지 않는다. 파일로 떨군다 → 읽기/쓰기 과금 0
  · 받은 것을 data/store/ 에 누적 보관하고,
    사이트에는 최근 것만 잘라서 web/public/data/ 로 내보낸다
  · API 키를 코드에 박지 않고 .env 에서 읽는다

실행
  python collect.py              최근 3일 (평소 운영)
  python collect.py --days 10    최근 10일
  python collect.py --backfill 90  90일치 몰아서 채우기 (최초 1회)
"""
import io
import shutil
import os
import re
import csv
import json
import time
import argparse
import urllib.parse
from datetime import datetime, timedelta, timezone

import requests
import urllib3

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

ROOT = os.path.dirname(os.path.abspath(__file__))
STORE = os.path.join(ROOT, "data", "store")
OUT = os.path.join(ROOT, "web", "public", "data")
KST = timezone(timedelta(hours=9))

KEEP_DAYS = 70      # 파일로 보관하는 기간 (10주 — 여유를 두고 보관)
SHOW_DAYS = 49      # 사이트에 싣는 기간 (7주)
MAX_ROWS = 300      # first.json / live.json 에 싣는 건수 (첫 화면·옛 사이트용)
BOARD_CHUNK = 500   # 한 달치를 나눠 담는 묶음 크기
# 순위 30곳(corps)만 따로 담는 작은 묶음. 목록에서 빼고 «펼칠 때만» 받습니다.
BOARD_RANK_CHUNK = 50
BOARD_RANK_KEYS = ("corps", "rq", "drw")

# 2026-08-31 — 용역을 빼고 «공사»만 다룹니다.
#   3년치 482,630건 중 용역이 363,783건(75%)이라, 빼면 사이트가 크게 가벼워집니다.
#   되돌리려면 여기에 "serv" 를 다시 넣기만 하면 됩니다. 원본 파일은 지우지 않았습니다.
KINDS = ("con",)

# 3년치 원본과 같은 모양으로 계속 쌓이는 누적 파일.
# build_json.py 가 data/extra_*.csv 를 자동으로 함께 읽으므로,
# 이 파일만 자라면 «3년치 분석»이 저절로 최신 상태를 유지한다.
# (store/*.json 은 45일이면 잘리므로 분석용 이력은 여기에 남긴다)
# 달마다 파일을 나눕니다.
#   자동화(GitHub)가 매일 이 파일을 저장소에 올리는데, 한 덩어리로 두면
#   14MB 파일이 매일 통째로 다시 올라가 저장소가 금방 무거워집니다.
#   달별로 쪼개면 이번 달 것(1~2MB)만 바뀝니다.
#   build_json.py 는 data/extra_*.csv 를 전부 읽으므로 그대로 동작합니다.
ARCHIVE_DIR = os.path.join(ROOT, "data")


def archive_path(ym):
    return os.path.join(ARCHIVE_DIR, f"extra_{ym}.csv")
# 사업자번호를 반드시 남깁니다.
#   같은 이름의 다른 법인이 아주 많습니다 — «대영건설» 한 이름에 법인 40곳.
#   3년치의 46%가 그런 이름이라, 번호가 없으면 업체 화면 절반이 남의 실적입니다.
#   조달청 응답에 이미 들어 있는데(업체명^사업자번호^대표^금액^투찰률) 그동안 버리고 있었습니다.
# ⚠️ A값 칸을 넣은 이유 —
#   가상 시뮬레이션의 «합격 판정»에 A값이 필요한데, 개찰결과 API 는 A값을 주지 않습니다.
#   그래서 그동안 «A값 5% 가정» 으로 돌렸습니다. 가정은 결과를 흔듭니다.
#   A값은 «공고» 쪽 오퍼레이션에만 있으므로, 받아 둘 때 누적 CSV 에도 같이 적어
#   앞으로는 실제값으로 시뮬레이션합니다.
# ⚠️ 예가범위·참가업체수 칸을 뒤늦게 넣은 이유 (2026-09-03) —
#   가상 시뮬레이션이 이 CSV 만 읽는데, 여기에 예가범위가 없어서
#   **모든 공고를 ±3% 로 가정**하고 있었습니다. 실제로는 11.1%가 ±2% 입니다.
#   ±2% 는 사정률 폭이 좁아 σ 가 0.7676 → 0.5118 로 달라지고,
#   등급도 +1 이 아니라 −2 를 받아야 합니다. 즉 화면과 시뮬레이션이 어긋나 있었습니다.
#   참가업체수는 «무작위로 넣었으면 1/N» 이라는 비교 기준을 내기 위한 것입니다.
#   옛 줄에는 이 칸이 비어 있습니다 — 오늘부터 쌓입니다.
ARCH_COLS = ["공고번호", "날짜", "발주기관", "공고명",
             "1순위업체", "사업자번호", "대표자", "투찰금액", "투찰률", "기초금액",
             "A값", "A값적용", "예가하한", "예가상한", "참가업체수"]

BASE = "http://apis.data.go.kr/1230000"
ENDPOINTS = {
    ("first", "con"):  f"{BASE}/as/ScsbidInfoService/getOpengResultListInfoCnstwk",
    ("first", "serv"): f"{BASE}/as/ScsbidInfoService/getOpengResultListInfoServc",
    ("live",  "con"):  f"{BASE}/ad/BidPublicInfoService/getBidPblancListInfoCnstwk",
    ("live",  "serv"): f"{BASE}/ad/BidPublicInfoService/getBidPblancListInfoServc",
}

# ══════════════════════════════════════════════════════════════
#  ★ 개찰 순위 (1위 ~ 꼴찌) — 2026-09-02 발견
#
#  «조달청은 1순위만 준다» 고 제가 두 번 말했습니다. **둘 다 틀렸습니다.**
#  아래 오퍼레이션이 투찰업체를 **전부** 줍니다. 다만 조건이 있습니다:
#     · 날짜로 부르면 아무것도 안 옵니다
#     · **공고번호(bidNtceNo + bidNtceOrd)로 불러야** 옵니다
#  예전 진단은 날짜로만 두드렸고, 저는 그 빈 응답을 «없다» 로 읽었습니다.
#
#  실제 응답 항목:
#     opengRank(개찰순위) · prcbdrNm(투찰업체) · prcbdrBizno · prcbdrCeoNm
#     bidprcAmt(투찰금액) · bidprcrt(투찰률) · bidprcDt(투찰일시)
#     drwtNo1 · drwtNo2 (그 업체가 뽑은 예비가격 추첨번호)
# ══════════════════════════════════════════════════════════════
OPENG_RANK = f"{BASE}/as/ScsbidInfoService/getOpengResultListInfoOpengCompt"

# 기초금액(예정가격 산정의 기준이 되는 금액). 공고 목록에는 안 들어있고 별도 오퍼레이션이다.
BSIS = {
    "con":  f"{BASE}/ad/BidPublicInfoService/getBidPblancListInfoCnstwkBsisAmount",
    "serv": f"{BASE}/ad/BidPublicInfoService/getBidPblancListInfoServcBsisAmount",
}
# 목록조회로 못 채운 건은 공고번호로 하나씩 조회합니다.
#   A값까지 채우게 되면서 대상이 늘었습니다 (마감 전 공고 1,500건쯤).
#   한 번에 250건이면 하루 세 번 돌아 이틀이면 다 찹니다.
#   조달청이 먹통이면 차단기가 8번 만에 끊으므로 오래 붙잡히지 않습니다.
BSIS_ONE_CAP = 250

# 면허·업종 제한. 입찰에서 이게 제일 먼저 걸리는 조건인데
# 공고 목록에는 안 들어 있고 별도 오퍼레이션으로 옵니다.
# 낙찰자 상세. 2026-09-02 진단으로 확인했습니다.
#   bidwinnrNm / bidwinnrBizno / bidwinnrCeoNm / bidwinnrAdrs / bidwinnrTelNo
#   sucsfbidAmt / sucsfbidRate / prtcptCnum(참가업체수) 까지 옵니다.
#   개찰결과 오퍼레이션에는 주소·전화가 없어서 이걸 따로 받습니다.
SCSBID = {
    "con":  f"{BASE}/as/ScsbidInfoService/getScsbidListSttusCnstwk",
    "serv": f"{BASE}/as/ScsbidInfoService/getScsbidListSttusServc",
}

LIC = {
    "con":  f"{BASE}/ad/BidPublicInfoService/getBidPblancListInfoLicenseLimit",
    "serv": f"{BASE}/ad/BidPublicInfoService/getBidPblancListInfoLicenseLimit",
}


def load_env():
    """.env 를 읽어 환경변수로 (python-dotenv 없이도 동작)"""
    p = os.path.join(ROOT, ".env")
    if not os.path.exists(p):
        return
    with open(p, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))


def api_key():
    k = os.environ.get("G2B_API_KEY", "").strip()
    if not k:
        raise SystemExit(
            "❌ G2B_API_KEY 가 없습니다.\n"
            "   .env 파일을 만들고 아래 한 줄을 넣어주세요:\n"
            "   G2B_API_KEY=발급받은_디코딩_키")
    return urllib.parse.unquote(k)


# 조달청이 통째로 먹통일 때를 대비한 «차단기».
#   2026-09-01 에 조달청 API 가 전부 ConnectTimeout 이 났는데,
#   한 건당 25초씩 150건을 계속 기다리느라 배치가 40분 넘게 늘어졌다.
#   연달아 실패하면 이번 회차는 통신을 포기하고 넘어간다 — 자료는 누적분이 지킨다.
FIELDS_SEEN = set()      # 오퍼레이션별로 응답 항목 이름을 한 번씩만 찍기 위한 표시
DIAG = {}                # 진단 결과 — 사이트에 파일로 남겨 나중에 읽습니다
NET_FAILS = 0            # 연속 통신 실패 횟수
NET_DOWN = False         # 차단기가 내려갔는지 (apis.data.go.kr 전용)
# ⚠️ NET_DOWN 은 «조달청 OpenAPI(apis.data.go.kr)» 의 차단기입니다.
#    나라장터 첨부파일은 www.g2b.go.kr — 다른 서버입니다.
#    한쪽 차단기로 다른 쪽을 막으면, 수집 막바지에 차단기가 내려간 회차마다
#    내역서를 «조용히» 한 개도 안 받습니다(2026-09-05 에 실제로 겪었습니다).
NO_NET = False           # --exportonly 처럼 «바깥을 아예 안 부른다» 는 뜻
NET_LIMIT = 8            # 이만큼 연달아 실패하면 포기
NET_TIMEOUT = 15         # 한 건당 기다리는 시간(초)
# ⚠️ 2026-09-06 (일요일) — «연달아 8번» 만으로는 안 잡히는 날이 있었다.
#    조달청이 «느리게» 아픈 날: 열에 여덟은 ReadTimeout, 둘은 성공. 성공 한 번이 NET_FAILS 를 0 으로
#    되돌리니 차단기가 영영 안 내려갔고, 15초짜리 타임아웃 109번 = 27분을 기다리다
#    Actions 45분 상한에 걸려 **배포 자체가 취소**됐다 (#109). 화면 수정이 사이트에 못 올라갔다.
#    → 두 번째 차단기: «이 회차에서 조달청에 쓴 시간» 이 예산을 넘으면 그 뒤 호출은 전부 건너뛴다.
#      뒤에 남은 단계(집계·빌드·굽기·배포)가 10분쯤 필요하므로 수집에는 20분만 준다.
#      건너뛴 건은 bask/rask 가 «오래된 것부터» 규칙으로 다음 회차에 다시 묻는다 — 잃는 것이 없다.
NET_BUDGET_S = int(os.environ.get("NET_BUDGET_S", "1200"))   # 조달청 호출에 쓸 수 있는 시간(초)
NET_T0 = time.time()     # 이 회차 시작 시각
NET_FAILS_ALL = 0        # 누적 실패(연속 아님) — 로그에 «몇 번 기다렸나» 를 남기기 위한 것


def fetch(url, key, day=None, extra=None, label="", why=None):
    """why 에 dict 를 넘기면 실패 이유(HTTP·resultCode·resultMsg)를 담아 줍니다.
       진단에서 «응답 없음» 과 «필수값이 달라서 안 됨» 을 구별하기 위한 것입니다."""
    """조달청 공통 호출.
    예전에 기초금액이 '계속 실패'했던 건 대부분 조용히 삼켜서 원인이 안 보였기 때문이다.
    그래서 여기서는 HTTP 코드 / resultCode / 본문 앞머리를 반드시 찍는다."""
    global NET_FAILS, NET_DOWN, NET_FAILS_ALL
    if NET_DOWN:
        return []
    if time.time() - NET_T0 > NET_BUDGET_S:
        NET_DOWN = True
        print(f"    ⛔ 조달청 호출에 {NET_BUDGET_S // 60}분을 다 썼습니다 (실패 {NET_FAILS_ALL}번 기다림). "
              f"남은 호출은 이번 회차에서 건너뜁니다 — 다음 회차가 오래된 것부터 다시 묻습니다.")
        if why is not None:
            why.update({"net": "budget"})
        return []
    params = {
        "serviceKey": key, "numOfRows": "999", "pageNo": "1",
        "inqryDiv": "1", "type": "json",
    }
    if day is not None:
        d = day.strftime("%Y%m%d")
        params["inqryBgnDt"] = d + "0000"
        params["inqryEndDt"] = d + "2359"
    if extra:
        params.update(extra)
    tag = label or url.rsplit("/", 1)[-1]
    try:
        r = requests.get(url, params=params, timeout=NET_TIMEOUT, verify=False,
                         headers={"User-Agent": "Mozilla/5.0"})
    except Exception as e:
        NET_FAILS += 1
        NET_FAILS_ALL += 1
        print(f"    ! {tag} 통신 실패 ({type(e).__name__})")
        if why is not None:
            why.update({"net": type(e).__name__})
        if NET_FAILS >= NET_LIMIT:
            NET_DOWN = True
            print(f"    ⛔ 조달청 통신이 {NET_LIMIT}번 연달아 실패했습니다. "
                  f"이번 회차는 수집을 건너뜁니다 — 사이트는 누적 자료로 그대로 올라갑니다.")
        return []
    NET_FAILS = 0
    if r.status_code != 200:
        print(f"    ! {tag} HTTP {r.status_code}")
        if why is not None:
            why.update({"http": r.status_code, "body": " ".join(r.text.split())[:200]})
        return []
    try:
        j = r.json()
    except Exception:
        head = " ".join(r.text.split())[:180]
        print(f"    ! {tag} JSON 아님 → {head}")
        if why is not None:
            why.update({"http": r.status_code, "json": False, "body": head})
        return []
    resp = j.get("response", {}) if isinstance(j, dict) else {}
    head = resp.get("header", {}) or {}
    code = str(head.get("resultCode", "")).strip()
    if code and code not in ("00", "0"):
        print(f"    ! {tag} 응답코드 {code} · {head.get('resultMsg', '')}")
        if why is not None:
            why.update({"http": r.status_code, "code": code,
                        "msg": str(head.get("resultMsg", ""))[:160]})
        return []
    items = (resp.get("body", {}) or {}).get("items", [])
    # 오퍼레이션마다 어떤 항목이 오는지 한 번씩 찍어봅니다.
    #   아이건설넷은 A값(사회보험료 등 법정경비)과 관급자재금액을 자동으로 채웁니다.
    #   그렇다면 조달청 응답 어딘가에 들어 있다는 뜻입니다. 그걸 찾으려는 것입니다.
    op = url.rsplit("/", 1)[-1]
    if op not in FIELDS_SEEN and isinstance(items, (list, dict)):
        one = items[0] if isinstance(items, list) and items else (
            items if isinstance(items, dict) else None)
        if isinstance(one, dict) and one:
            FIELDS_SEEN.add(op)
            DIAG[op] = {"fields": sorted(one.keys()),
                        "sample": {k: str(v)[:40] for k, v in list(one.items())[:60]}}
            print(f"    · [진단] {op} 항목: {', '.join(sorted(one.keys()))}")
            # A값·관급자재로 보이는 항목이 있으면 값까지 보여줍니다
            hint = {k: v for k, v in one.items()
                    if any(w in str(k).lower() for w in
                           ("insrnc", "sfty", "safe", "retire", "govsply", "gov",
                            "mtrl", "amt", "prce", "rate"))}
            if hint:
                print(f"    · [진단] {op} 금액·비율 항목: "
                      + ", ".join(f"{k}={v}" for k, v in list(hint.items())[:24]))
    if isinstance(items, dict):
        items = items.get("item", [items])
    if isinstance(items, str):
        return []
    return items or []


def pick(item, *names):
    """조달청 응답은 문서와 필드 대소문자가 다를 때가 있다 (bssamt vs bssAmt).
    이름을 소문자로 눕혀서 찾는다 — 기초금액이 늘 비어 보이던 진짜 원인."""
    if not isinstance(item, dict):
        return None
    low = {str(k).lower(): v for k, v in item.items()}
    for n in names:
        v = low.get(n.lower())
        if v not in (None, "", "-", "0"):
            return v
    return None


def to_int(v):
    try:
        return int(float(str(v).replace(",", "").strip()))
    except Exception:
        return 0


def to_rate(v):
    try:
        f = float(str(v).replace("%", "").strip())
        return round(f, 3) if 0 < f <= 200 else None
    except Exception:
        return None


def to_f(v):
    try:
        return round(float(str(v).replace("%", "").strip()), 3)
    except Exception:
        return None


def bsis_row(it):
    """기초금액 응답 한 줄 → 화면에 쓸 형태.
    base 예정가격 기준금액 / lo,hi 예가범위(%) / a,b 낙찰하한율 참고용"""
    base = to_int(pick(it, "bssamt", "bssAmt", "bsisAmt", "BssAmt", "basisAmt"))
    if not base:
        return None
    return {
        "base": base,
        "lo": to_f(pick(it, "rsrvtnPrceRngBgnRate", "rsrvtnPrceRngBgnRt")),
        "hi": to_f(pick(it, "rsrvtnPrceRngEndRate", "rsrvtnPrceRngEndRt")),
    }


# ── A값(법정경비) 항목들 ────────────────────────
#   2026-09-01 진단으로 확인했습니다. 조달청 기초금액 응답에 항목별로 다 들어 있습니다.
#   그동안 이걸 안 뒤져봐서 «API 로는 못 가져온다» 고 잘못 말했습니다.
#
#   A값은 투찰률을 곱하지 않고 그대로 더하는 금액입니다.
#       투찰금액 = (예정가격 − A값) × 투찰률 + A값
#   그래서 A값을 빼먹으면 낙찰하한을 잘못 계산해 실격합니다.
A_PARTS = [
    ("sftyMngcst", "산업안전보건관리비"),
    ("sftyChckMngcst", "안전관리비"),
    ("rtrfundNon", "퇴직공제부금비"),
    ("envCnsrvcst", "환경보전비"),
    ("scontrctPayprcePayGrntyFee", "하도급대금지급보증수수료"),
    ("mrfnHealthInsrprm", "국민건강보험료"),
    ("npnInsrprm", "국민연금보험료"),
    ("odsnLngtrmrcprInsrprm", "노인장기요양보험료"),
    ("qltyMngcst", "품질관리비"),
]


def extra_amounts(item):
    """A값을 항목별로 모으고 합계를 냅니다.

    bidPrceCalclAYn 이 «Y» 면 이 공고가 A값을 적용한다는 뜻입니다.
    품질관리비는 qltyMngcstAObjYn 이 «Y» 일 때만 A값에 넣습니다.
    """
    out = {}
    parts, total = [], 0
    qobj = str(pick(item, "qltyMngcstAObjYn") or "").upper() == "Y"
    for k, label in A_PARTS:
        v = to_int(pick(item, k) or 0)
        if v <= 0:
            continue
        if k == "qltyMngcst" and not qobj:
            continue
        parts.append([label, v])
        total += v
    if total > 0:
        out["aval"] = total
        out["aparts"] = parts
    yn = str(pick(item, "bidPrceCalclAYn") or "").upper()
    if yn in ("Y", "N"):
        out["ayn"] = yn        # 이 공고가 A값을 적용하는지
    g = to_int(pick(item, "govsplyAmt", "govcnstrtnGovsplyMtrlAmt",
                    "contrctrcnstrtnGovsplyMtrlAmt", "govsplyMtrlAmt") or 0)
    if g > 0:
        out["gmtrl"] = g
    return out


# ─────────────────────────────────────────────────────────────
#  누락 방지 두 가지 — 2026-09-04 (소장님: 「누락이 있으면 절대 안돼」)
#
#  ① 쪽 넘기기
#     목록을 «1쪽(999건)» 만 받고 있었습니다. 하루 1,000건이 넘으면 나머지가
#     **에러 없이** 사라집니다. 실측으로는 하루 최다 개찰 575건 · 공고 502건이라
#     아직 잘린 적은 없지만, 공고가 몰리는 시기에는 넘을 수 있고 넘어도 아무도 모릅니다.
#     → 999건이 꽉 차서 오면 다음 쪽을 이어 받습니다. 999건 미만이면 호출은 그대로 1번입니다.
#
#  ② 받은 날짜 적어 두기
#     매 회차가 «최근 3일» 만 훑습니다. 조달청이 3일 넘게 먹통이면 그 날짜가
#     창 밖으로 밀려나 영영 안 들어옵니다. 받은 날을 적어 두고,
#     빠진 날이 있으면 다음 회차가 그날을 다시 훑습니다.
#     ⚠️ 한 회차에 덧붙이는 «빠진 날» 은 BACK_MAX 개로 막습니다 —
#        기록이 통째로 날아갔을 때 조달청을 한꺼번에 두드리지 않기 위해서입니다.
# ─────────────────────────────────────────────────────────────

PAGE_ROWS = 999          # 조달청 한 쪽 최대
PAGE_CAP = 8             # 하루 최대 8쪽(7,992건). 여기까지 차면 진단(diag)에 남깁니다
DAYS_LOG = os.path.join(STORE, "days.json")
BACK_DAYS = 14           # 빠진 날을 얼마나 거슬러 올라가 찾을지
BACK_MAX = 4             # 한 회차에 덧붙일 «빠진 날» 최대 개수


def fetch_paged(url, key, day, extra=None, label=""):
    """999건이 꽉 차서 오면 다음 쪽을 이어 받습니다.

    돌려주는 것: (줄들, 실패이유). 실패이유가 비어 있으면 «제대로 받았다» 는 뜻입니다.
    이 값으로 «그날을 받았다» 를 기록하므로, 빈 날(주말)과 실패한 날이 구별됩니다.
    """
    if NET_DOWN:
        return [], {"net": "down"}
    out, why = [], {}
    for pg in range(1, PAGE_CAP + 1):
        ex = dict(extra or {})
        ex["pageNo"] = str(pg)
        w = {}
        got = fetch(url, key, day, ex,
                    label=(f"{label} {pg}쪽" if pg > 1 else label), why=w)
        if w:
            why = w
            break
        out.extend(got)
        if len(got) < PAGE_ROWS:
            break
        print(f"    · {label} {pg}쪽이 가득 찼습니다 — 다음 쪽을 이어 받습니다")
    else:
        msg = f"{label}: {PAGE_CAP}쪽({PAGE_CAP * PAGE_ROWS:,}건)을 다 채웠습니다 — 더 있을 수 있습니다"
        print(f"    ⚠️ {msg}")
        DIAG.setdefault("_page_cap", []).append(msg)
    return out, why


def load_days():
    try:
        with io.open(DAYS_LOG, encoding="utf-8") as f:
            v = json.load(f)
        return v if isinstance(v, dict) else {}
    except Exception:
        return {}


def save_days(v):
    try:
        if len(v) > 120:                      # 넉 달치만 둡니다
            for k in sorted(v)[:len(v) - 120]:
                v.pop(k, None)
        os.makedirs(STORE, exist_ok=True)
        with io.open(DAYS_LOG, "w", encoding="utf-8") as f:
            json.dump(v, f, ensure_ascii=False)
    except Exception as e:
        print(f"    ! 받은 날짜 기록 실패 ({type(e).__name__})")


def days_to_scan(today, days):
    """이번 회차에 훑을 날짜. 기본 창(최근 days일) + 기록에 빠져 있는 날."""
    base = [today - timedelta(days=i) for i in range(days - 1, -1, -1)]
    have = load_days()
    miss = []
    for i in range(days, BACK_DAYS + 1):
        d = today - timedelta(days=i)
        if d.weekday() >= 5:                  # 토·일은 개찰이 없습니다
            continue
        rec = have.get(d.strftime("%Y-%m-%d"))
        if not (isinstance(rec, dict) and rec.get("ok")):
            miss.append(d)
    miss.sort()                               # 오래된 날부터 — 기본 창과 같은 순서
    return miss[:BACK_MAX] + base


def scsbid_by_day(key, day, kind):
    """하루치 낙찰자 상세를 받아 {공고번호: {...}} 로.

    사용자가 «1순위 업체 주소·전화를 보고 싶다» 고 해서 붙였습니다.
    나라장터 개찰 결과에 공개되는 정보입니다."""
    out = {}
    for pg in range(1, 6):
        got = fetch(SCSBID[kind], key, day, {"pageNo": str(pg)},
                    label=f"낙찰자 {kind} {day:%m-%d} {pg}쪽")
        for it in got:
            no = str(pick(it, "bidNtceNo") or "").strip()
            if not no:
                continue
            row = {}
            nm = pick(it, "bidwinnrNm")
            if nm:
                row["win"] = str(nm).strip()[:40]
            for src, dst in (("bidwinnrBizno", "bno"), ("bidwinnrCeoNm", "ceo"),
                             ("bidwinnrAdrs", "adr"), ("bidwinnrTelNo", "tel")):
                v = pick(it, src)
                if v:
                    row[dst] = str(v).strip()[:60]
            np = to_int(pick(it, "prtcptCnum") or 0)
            if np:
                row["np"] = np
            sa = to_int(pick(it, "sucsfbidAmt") or 0)
            if sa:
                row["sAmt"] = sa
            sr = to_rate(pick(it, "sucsfbidRate"))
            if sr is not None:
                row["sRate"] = sr
            if row:
                out[no] = row
        if len(got) < 999:
            break
    return out


def backfill_bsis(key, first, live, days, sleep=0.4):
    """지난 N일치 기초금액·A값을 다시 훑어 빈 칸을 채웁니다.

    왜 필요한가:
      기초금액과 A값은 «공고» 오퍼레이션에서만 옵니다. 그런데 배치는 최근 2~3일만
      훑기 때문에, 그 시점에 아직 공개되지 않은 기초금액·A값은 영영 못 받습니다.
      실측: 개찰 500건 중 기초금액 167건(33%), A값 42건(8.4%) 뿐이었습니다.
      그러면 시뮬레이션이 «A값 가정» 으로 돌아갑니다.
    비용:
      bsis_by_day 는 하루당 1~2회 호출입니다. 45일이면 50~90회 — 하루 한 번이면 부담 없습니다.
    """
    got = 0
    for i in range(days, 0, -1):
        day = datetime.now(KST) - timedelta(days=i)
        for kind in KINDS:
            bm = bsis_by_day(key, day, kind)
            for store in (first, live):
                for no, b in bm.items():
                    row = store.get(kind, {}).get(no)
                    if row is None:
                        continue
                    for f, v in b.items():
                        if v not in (None, "", 0) and not row.get(f):
                            row[f] = v
                            if f == "aval":
                                got += 1
            time.sleep(sleep)
        if NET_DOWN:
            break
    print(f"  → 소급 보충 {days}일치 — A값 {got:,}건 새로 채움")


def bsis_by_day(key, day, kind):
    """하루치 기초금액을 통째로 받아 {공고번호: {...}} 로."""
    out = {}
    for it in fetch(BSIS[kind], key, day, None,
                    label=f"기초금액 {kind} {day:%m-%d}"):
        no = str(pick(it, "bidNtceNo") or "").strip()
        r = bsis_row(it)
        if no and r:
            r.update(extra_amounts(it))
            out[no] = r
    return out


LIC_PAGES = 12       # 면허제한은 물품·용역까지 섞여 와서 한 쪽(999건)으로는 턱없이 모자람


def lic_by_day(key, day, kind):
    """하루치 면허·업종 제한을 통째로 받아 {공고번호: [제한명, ...]} 로.

    한 공고에 여러 줄이 올 수 있습니다 (토목 + 건축 처럼).

    ⚠️ 쪽넘김을 반드시 해야 합니다.
       이 오퍼레이션은 공사·용역·물품을 다 섞어서 줍니다.
       한 쪽(999건)만 받으면 대부분 물품이라 공사는 거의 안 걸립니다.
       실제로 1,946건 중 3건만 채워졌습니다."""
    out = {}
    rows = []
    for pg in range(1, LIC_PAGES + 1):
        got = fetch(LIC[kind], key, day, {"pageNo": str(pg)},
                    label=f"면허제한 {kind} {day:%m-%d} {pg}쪽")
        rows.extend(got)
        if len(got) < 999:          # 마지막 쪽
            break
    for it in rows:
        no = str(pick(it, "bidNtceNo") or "").strip()
        # bsnsDivNm 이 «물품»·«용역» 인 줄이 섞여 옵니다. 공사만 씁니다.
        div = str(pick(it, "bsnsDivNm") or "").strip()
        if div and div != "공사":
            continue
        nm = pick(it, "lcnsLmtNm", "licenseLmtNm", "indstrytyNm",
                  "bidprcPsblIndstrytyNm", "lmtNm", "prmsnCorpNm")
        if not no or not nm:
            continue
        nm = str(nm).strip()
        cur = out.setdefault(no, [])
        if nm not in cur:
            cur.append(nm)
    return out


def bsis_one(key, no, kind):
    """공고번호로 한 건만. 목록조회에서 빠진 건을 메운다 (inqryDiv=2)."""
    for it in fetch(BSIS[kind], key, None,
                    {"inqryDiv": "2", "bidNtceNo": no}, label=f"기초금액 {no}"):
        r = bsis_row(it)
        if r:
            r.update(extra_amounts(it))
            return r
    return None


# 화면에 싣는 순위 개수. 전부 실으면 파일이 터집니다 —
# 실측(2026-09-02): 한 공고에 999곳까지 옵니다. 60건만 채웠는데 목록 파일이
# 310KB → 743KB 로 뛰었습니다. 전부 채우면 80MB 가 됩니다.
# 우리에게 필요한 것은 «가장 낮게 쓴 쪽»입니다 — 낙찰하한 근처가 승부처이기 때문입니다.
# 그래서 낮은 금액 순으로 30곳만 싣고, 전체 참가업체수는 np 로 따로 보여줍니다.
RANK_KEEP = 30


def openg_ranks(key, no, ord_="000"):
    """공고 하나의 투찰업체를 순위대로 전부 가져온다.

    돌려주는 모양은 parse_corps 와 **똑같이** 맞춥니다:
        [[업체명, 투찰금액, 투찰률, 사업자번호, 대표자, 추첨번호1, 추첨번호2], ...]
    화면 코드가 corps[i][0..2] 로 읽고 있어서, 자리를 바꾸면 화면이 깨집니다.

    ★ 2026-09-03 — 추첨번호(drwtNo1·drwtNo2)도 담습니다.
      소장님: 「클릭하는 번호에 따라 달라지는 거 아닌가?」 — 처음 순위를 받을 때
      이 두 칸을 버리고 있어서 «번호와 낙찰의 관계» 를 자료로 대볼 수 없었습니다.
      낮은 30곳의 번호는 corps 에, **전체 투찰자의 번호 분포**(1~15 번이 각각 몇 번
      찍혔나)는 15칸짜리 drw 로 돌려줍니다. 분포가 있어야 «가장 많이 찍힌 4개»
      (= 사정률에 들어간 예비가격)를 알 수 있습니다.

    돌려주는 것: (corps30, 전체건수, 사다리, drw15)
    """
    items = fetch(OPENG_RANK, key,
                  extra={"bidNtceNo": no, "bidNtceOrd": str(ord_ or "000") or "000"},
                  label=f"개찰순위 {no}")
    out = []
    drw = [0] * 15
    for it in items:
        nm = str(pick(it, "prcbdrNm") or "").strip()
        amt = to_int(pick(it, "bidprcAmt"))
        if not nm or amt <= 0:
            continue
        rank = to_int(pick(it, "opengRank")) or 9999
        bno = re.sub(r"[^0-9]", "", str(pick(it, "prcbdrBizno") or ""))
        ceo = str(pick(it, "prcbdrCeoNm") or "").strip()[:12]
        d1 = to_int(pick(it, "drwtNo1"))
        d2 = to_int(pick(it, "drwtNo2"))
        for d in (d1, d2):
            if 1 <= d <= 15:
                drw[d - 1] += 1
        out.append((rank, amt, [nm, amt, to_rate(pick(it, "bidprcrt")),
                                bno if len(bno) == 10 else "", ceo,
                                d1 if 1 <= d1 <= 15 else 0,
                                d2 if 1 <= d2 <= 15 else 0]))
    # 순위가 비어 오는 경우가 있어 «금액이 낮은 순»을 보조 기준으로 둡니다
    out.sort(key=lambda x: (x[0], x[1]))

    # ── 순위 사다리 ────────────────────────────────────────────
    #  화면에 «우리 금액이면 몇 위» 를 말하려면 낮은 30곳만으론 부족합니다.
    #  851곳이 붙은 공고에서 우리 금액은 30위권 밖이라 «최소 31위» 밖에 못 합니다.
    #  그렇다고 851개 금액을 다 실으면 파일이 터집니다.
    #  그래서 **몇 등이 얼마였는지 사다리만** 담습니다 — 10칸이면 전 구간을 덮습니다.
    #  화면은 우리 금액이 어느 칸 사이에 떨어지는지 보고 등수를 좁힙니다.
    ladder = []
    n = len(out)
    for pos in (1, 2, 3, 5, 10, 20, 50, 100, 200, 500, 1000):
        if pos <= n:
            ladder.append([pos, out[pos - 1][1]])
    if n and (not ladder or ladder[-1][0] != n):
        ladder.append([n, out[n - 1][1]])       # 꼴찌도 한 칸

    return [c for _, _, c in out[:RANK_KEEP]], n, ladder, drw


def parse_corps(raw, limit=6):
    """'업체명^사업자번호^대표^금액^투찰률|업체명^...'
       → [[이름, 금액, 투찰률, 사업자번호, 대표자], ...]

    앞의 세 자리는 예전과 같은 자리에 둡니다 (화면 코드가 [0][1][2] 로 읽습니다).
    번호·대표는 뒤에 덧붙여서, 옛 자료와 섞여도 깨지지 않게 합니다."""
    out = []
    for chunk in str(raw or "").split("|")[:limit]:
        p = chunk.split("^")
        if len(p) >= 5 and p[0].strip():
            bno = re.sub(r"[^0-9]", "", p[1])
            out.append([p[0].strip(), to_int(p[3]), to_rate(p[4]),
                        bno if len(bno) == 10 else "", p[2].strip()[:12]])
    return out


def row_first(item):
    no = str(item.get("bidNtceNo", "")).strip()
    if not no:
        return None
    corps = parse_corps(item.get("opengCorpInfo", ""), limit=10)
    if not corps:
        return None
    win, amt, rate = corps[0][0], corps[0][1], corps[0][2]
    return {
        "no": no,
        # 공고차수 — 나라장터 원문 주소를 정확히 만들려면 필요합니다
        "ord": str(item.get("bidNtceOrd", "") or "").strip(),
        # 참가업체수 — 그 공고에 몇 곳이 들어왔는지. 경쟁 강도 그 자체입니다.
        "np": to_int(pick(item, "prtcptCnum") or 0),
        "name": str(item.get("bidNtceNm", "")).strip(),
        "inst": str(item.get("ntceInsttNm", "")).strip(),
        "dt": str(item.get("opengDt", "")).strip(),
        "win": win, "amt": amt, "rate": rate,
        # 업체를 이름이 아니라 «사업자번호» 로 가리기 위한 값
        "bno": corps[0][3] if len(corps[0]) > 3 else "",
        "ceo": corps[0][4] if len(corps[0]) > 4 else "",
        # 낙찰금액·낙찰률은 응답에 있을 때만 채워진다 (없으면 1순위 투찰금액이 곧 낙찰가)
        "sAmt": to_int(pick(item, "sucsfbidAmt", "sucsfbidPrce")) or 0,
        "sRate": to_rate(pick(item, "sucsfbidRate")),
        "corps": corps,
    }


def row_live(item):
    no = str(item.get("bidNtceNo", "")).strip()
    if not no:
        return None
    url = str(item.get("bidNtceDtlUrl", "") or "").replace(":8081", "").replace(":8101", "")

    def txt(*names):
        v = pick(item, *names)
        return str(v).strip() if v is not None else ""

    # 공고서에 있는 내용을 되도록 사이트 안에서 볼 수 있게 담아 옵니다.
    # 항목 이름이 문서와 다를 때가 있어 후보를 여러 개 적어 pick() 으로 찾습니다.
    # 비어 오면 화면에서 그 줄만 빠집니다.
    return {
        "no": no,
        "ord": txt("bidNtceOrd"),
        "name": str(item.get("bidNtceNm", "")).strip(),
        "inst": str(item.get("ntceInsttNm", "")).strip(),
        "dmnd": txt("dminsttNm"),                       # 수요기관
        "dt": str(item.get("bidNtceDt", "")).strip(),
        "budget": to_int(item.get("bdgtAmt", 0) or item.get("presmptPrce", 0)),
        "est": to_int(pick(item, "presmptPrce", "presmptPrceAmt") or 0),   # 추정가격
        "close": str(item.get("bidClseDt", "") or "").strip(),
        "openg": txt("opengDt"),                        # 개찰 일시
        "kind": txt("ntceKindNm"),                      # 공고종류(일반/긴급/재공고)
        "mthd": txt("cntrctCnclsMthdNm", "bidMethdNm"), # 계약방법
        "swin": txt("sucsfbidMthdNm"),                  # 낙찰자 결정방법
        # ★ 낙찰하한율 — 공고가 직접 알려주면 우리가 추정할 필요가 없습니다
        "llr": to_rate(pick(item, "sucsfbidLwltRate")),
        "rgn": txt("prtcptPsblRgnNm"),                  # 참가가능 지역
        "ind": txt("bidprcPsblIndstrytyNm", "indstrytyNm"),   # 참가가능 업종
        "joint": txt("cmmnSpldmdMethdNm"),              # 공동수급 방식
        "rebid": txt("rbidPermsnYn"),                   # 재입찰 허용
        "ofcl": txt("ntceInsttOfclNm"),                 # 담당자
        "tel": txt("ntceInsttOfclTelNo"),               # 연락처
        # 복수예비가격 — 사정률이 어떻게 정해지는지의 핵심입니다
        "pmth": txt("prearngPrceDcsnMthdNm"),           # 예정가격 결정방법
        "ptot": to_int(pick(item, "totPrdprcNum") or 0),   # 예비가격 개수 (보통 15)
        "pdrw": to_int(pick(item, "drwtPrdprcNum") or 0),  # 추첨 개수 (보통 4)
        "site": txt("cnstrtsiteRgnNm"),                 # 공사 현장 지역
        "rgnb": txt("rgnLmtBidLocplcJdgmBssNm"),        # 지역제한 판단기준
        "main": txt("mainCnsttyNm"),                    # 주공종
        # 공고문 첨부파일 — 나라장터에 안 가고 여기서 바로 받게 합니다
        "docs": [[n, u] for n, u in (
            (txt(f"ntceSpecFileNm{i}"), txt(f"ntceSpecDocUrl{i}"))
            for i in range(1, 11)) if n and u][:6],
        "url": url or "https://www.g2b.go.kr/index.jsp",
        # 조달청이 주는 또 다른 주소. 공고상세(bidNtceDtlUrl)와 다른 것인지,
        # 혹시 투찰 입구인지 확인하려고 담아둡니다.
        # 나라장터는 화면이 바뀌어도 주소가 안 바뀌는 구조라 투찰 딥링크가
        # 없을 가능성이 큽니다 — 추측하지 않고 실제 값을 보고 정합니다.
        "url2": txt("bidNtceUrl"),
    }



# ─────────────────────────────────────────────
#  진단 — 투찰업체 «전체» 를 주는 오퍼레이션 찾기
#
#  지금 쓰는 개찰결과 오퍼레이션은 1순위(낙찰자) 한 곳만 돌려줍니다.
#  (3년치 118,847건 전수 확인 — 전부 1곳)
#  그래서 사이트에 «1위~10위» 를 못 띄웁니다.
#  조달청에 투찰업체 목록을 주는 다른 오퍼레이션이 있는지
#  하루 한 번만 두드려 보고 결과를 기록에 남깁니다. (자료는 건드리지 않습니다)
# ─────────────────────────────────────────────
#  ⚠️ 2026-09-02 실측으로 확인된 것 (추측이 아닙니다):
#     지금 쓰는 «개찰결과»(getOpengResultListInfoCnstwk) 는
#     opengCorpInfo 에 **업체 한 곳만** 담아 줍니다.
#       예: "주식회사 와이티건설^1528601041^홍주호^6383600^90.341"
#       같은 공고의 prtcptCnum(참가업체수) 은 23 이었습니다.
#     저장소에 쌓인 개찰 10,913건 **전부** corps 가 1곳이었습니다.
#     → 이 오퍼레이션으로는 2~10위를 절대 못 가져옵니다.
#
#  그래서 아래에 «투찰내역을 줄 만한» 오퍼레이션 후보를 더 넣어 두고,
#  하루 한 번 두드려 응답이 오는지만 봅니다 (자료는 건드리지 않습니다).
#  결과는 data/diag.json 에 남으므로 다음 날 확인할 수 있습니다.
PROBE_OPS = [
    # ── 지금 쓰는 것 (1순위만 옵니다 — 위에 증거) ──
    ("낙찰자 목록",       f"{BASE}/as/ScsbidInfoService/getScsbidListSttusCnstwk"),
    ("면허·업종 제한",    f"{BASE}/ad/BidPublicInfoService/getBidPblancListInfoLicenseLimit"),
    # ── «전체 투찰내역» 후보 ──
    ("개찰결과 투찰업체", f"{BASE}/as/ScsbidInfoService/getOpengResultListInfoCnstwkPPSSrch"),
    ("개찰 참가업체",     f"{BASE}/as/ScsbidInfoService/getBidPblancListInfoCnstwkBidPrceList"),
    ("투찰 목록",         f"{BASE}/as/ScsbidInfoService/getOpengResultListInfoBidPrceList"),
    ("예비가격 상세",     f"{BASE}/as/ScsbidInfoService/getOpengResultListInfoCnstwkPreparPcDetail"),
    ("개찰결과 상세",     f"{BASE}/as/ScsbidInfoService/getOpengResultListInfoCnstwkDetail"),
    ("투찰가 상세",       f"{BASE}/as/ScsbidInfoService/getOpengResultListInfoCnstwkBidPrceDetail"),
    ("개찰 순위",         f"{BASE}/as/ScsbidInfoService/getOpengResultListInfoOpengCompt"),
    ("낙찰자 현황(용역)", f"{BASE}/as/ScsbidInfoService/getScsbidListSttusServc"),
    ("개찰결과(물품)",    f"{BASE}/as/ScsbidInfoService/getOpengResultListInfoThng"),
    ("입찰가격산식A",     f"{BASE}/ad/BidPublicInfoService/getBidPblancListInfoBidPrceCalclA"),
]


def save_diag():
    """진단 결과를 사이트에 파일로 남깁니다.
    GitHub 기록 화면은 무거워서 읽기 어렵습니다. 파일이면 언제든 볼 수 있습니다."""
    if not DIAG:
        return
    try:
        os.makedirs(OUT, exist_ok=True)
        path = os.path.join(OUT, "diag.json")
        # 수집 단계와 진단 단계가 따로 돌기 때문에 서로 지우지 않도록 합칩니다
        old = {}
        if os.path.exists(path):
            try:
                with open(path, encoding="utf-8") as f:
                    old = (json.load(f) or {}).get("ops", {}) or {}
            except Exception:
                old = {}
        merged = dict(old)
        for k, v in DIAG.items():
            if k == "_probe" and isinstance(old.get(k), dict):
                m = dict(old[k])
                m.update(v)
                merged[k] = m
            else:
                merged[k] = v
        with open(path, "w", encoding="utf-8") as f:
            json.dump({"built": datetime.now(KST).strftime("%Y-%m-%d %H:%M"),
                       "ops": merged}, f, ensure_ascii=False, indent=1)
        print(f"  · 진단 결과를 data/diag.json 에 남겼습니다 ({len(merged)}개)")
    except Exception as e:
        print(f"  ! 진단 저장 실패 ({type(e).__name__})")


def probe_ops(key, day, sample_no="", sample_ord="000"):
    """«투찰업체 전체»를 주는 오퍼레이션이 있는지 확인합니다.

    ⚠️ 2026-09-02 — 예전 진단은 **날짜로만** 두드렸습니다.
       공고번호를 받아야 하는 오퍼레이션은 그때 «응답 없음» 으로 나왔고,
       저는 그걸 «그런 오퍼레이션이 없다» 로 잘못 읽었습니다.
       그래서 이제 두 가지 방식으로 다 두드리고, **실패 이유(resultMsg)** 까지 적습니다.
         ① 날짜 범위 (inqryBgnDt~inqryEndDt)
         ② 공고번호 (bidNtceNo + bidNtceOrd)
       «응답 없음» 과 «필수값 누락» 과 «없는 서비스» 는 완전히 다른 이야기입니다.
    """
    print("-" * 52)
    print(f"진단 — 투찰업체 전체를 주는 오퍼레이션 찾기 (표본 공고 {sample_no or '없음'})")
    for label, url in PROBE_OPS:
        if NET_DOWN:
            print("  · 통신이 막혀 진단을 건너뜁니다")
            return
        op = url.rsplit("/", 1)[-1]
        rec = {"op": op, "tries": {}}
        modes = [("날짜", dict(day=day, extra=None))]
        if sample_no:
            modes.append(("공고번호", dict(
                day=None,
                extra={"bidNtceNo": sample_no, "bidNtceOrd": sample_ord})))
        for mname, kw in modes:
            why = {}
            items = fetch(url, key, label=f"[진단]{label}/{mname}", why=why, **kw)
            if items:
                one = items[0] if isinstance(items, list) else items
                keys = sorted(one.keys()) if isinstance(one, dict) else []
                rec["tries"][mname] = {
                    "rows": len(items), "fields": keys,
                    "sample": {k: str(v)[:80] for k, v in list(one.items())[:60]}
                    if isinstance(one, dict) else {}}
                print(f"  \u2713 {label}/{mname}: {len(items)}건 · "
                      f"항목 {', '.join(keys)[:300]}")
            else:
                rec["tries"][mname] = {"rows": 0, **why}
                print(f"  · {label}/{mname}: 응답 없음 "
                      f"({why.get('code', '')} {why.get('msg', '')})".rstrip())
            time.sleep(0.4)
        DIAG.setdefault("_probe", {})[label] = rec
    print("-" * 52)


def merge_ranks(first):
    """data/ranks.csv (사람이 조달데이터허브에서 받아 inbox 에 넣은 전체 투찰내역)를
       개찰 자료의 corps 에 붙인다 — 화면의 «투찰 순위»가 1위부터 10위까지 나옵니다.

    ⚠️ 조달청 공개 API 는 1순위만 줍니다(실측 10,913건 전부 1곳).
       그래서 순위는 이 파일이 있을 때만 채워집니다. 없으면 조용히 넘어갑니다.
    """
    path = os.path.join(ARCHIVE_DIR, "ranks.csv")
    if not os.path.exists(path):
        return 0
    by_no = {}
    try:
        with io.open(path, encoding="utf-8-sig", newline="") as f:
            for row in csv.DictReader(f):
                no = (row.get("공고번호") or "").strip()
                if not no:
                    continue
                try:
                    rank = int(float(row.get("순위") or 0))
                    amt = int(float(re.sub(r"[^0-9.]", "", str(row.get("투찰금액") or "0")) or 0))
                except Exception:
                    continue
                if not (1 <= rank <= 10) or amt <= 0:
                    continue
                rt = to_rate(row.get("투찰률"))
                bno = re.sub(r"[^0-9]", "", str(row.get("사업자번호") or ""))
                by_no.setdefault(no, []).append(
                    (rank, [str(row.get("업체명") or "").strip(), amt, rt,
                            bno if len(bno) == 10 else "", ""]))
    except Exception as e:
        print(f"  ! ranks.csv 읽기 실패 ({type(e).__name__}: {e}) — 넘어갑니다")
        return 0

    n = 0
    for kind in KINDS:
        for no, row in first[kind].items():
            got = by_no.get(no)
            if not got or len(got) < 2:
                continue          # 1위 하나뿐이면 지금 것과 같습니다
            got.sort(key=lambda x: x[0])
            row["corps"] = [c for _, c in got][:10]
            row["nrank"] = len(row["corps"])
            n += 1
    if n:
        print(f"  \u2192 투찰 순위 붙임 {n:,}건 (ranks.csv {len(by_no):,}공고)")
    return n


# ══════════════════════════════════════════════════════════════
#  씨앗(seed) 저장소 — 2026-09-03 사고의 재발 방지
#
#  사이트의 1순위가 «35쪽(691건 · 이틀치)» 뿐이었다. 소장님 PC 에는 11,541건(70일)이 있는데.
#  GitHub Actions 로그를 직접 열어 보니 「Cache Size: ~0 MB (532 B)」 —
#  회차 사이에 넘겨주는 data/store 캐시가 **빈 저장소**였고, 그 뒤 회차마다 빈 것을
#  이어받아 «2일치»만 새로 쌓고 있었다. 한 번 비면 영영 비는 구조.
#
#  두 가지 방어:
#   ① 저장소가 비어 있거나 거의 비어 있으면(200건 미만) data/seed/{first,live}.json.gz 에서
#      먼저 복구한다. 씨앗은 소장님 PC 의 70일치를 압축해 저장소(git)에 넣어 둔 것이다.
#      (4.4MB — 한 번만 커밋. 워크플로 파일은 못 고치므로 collect.py 안에서 해결한다)
#   ② 이미 1,000건 넘게 있는 저장소를 그 절반도 안 되는 것으로 덮어쓰려 하면 거부한다.
#      (조달청 장애·파싱 실패로 빈 것이 만들어져도 좋은 저장소를 지우지 못하게)
# ══════════════════════════════════════════════════════════════
SEED = os.path.join(ROOT, "data", "seed")


def _store_rows(d):
    return sum(len(v) for v in d.values() if isinstance(v, dict))


def _load_seed(name):
    p = os.path.join(SEED, f"{name}.json.gz")
    if not os.path.exists(p):
        return None
    try:
        import gzip
        with gzip.open(p, "rt", encoding="utf-8") as f:
            d = json.load(f)
        return {"con": d.get("con", {}), "serv": d.get("serv", {})}
    except Exception as e:
        print(f"  ! 씨앗 {name} 읽기 실패 ({type(e).__name__}: {e})")
        return None


def load_store(name):
    p = os.path.join(STORE, f"{name}.json")
    d = {"con": {}, "serv": {}}
    if os.path.exists(p):
        try:
            with open(p, encoding="utf-8") as f:
                raw = json.load(f)
            d = {"con": raw.get("con", {}), "serv": raw.get("serv", {})}
        except Exception as e:
            print(f"  ! 저장소 {name} 읽기 실패 ({type(e).__name__}: {e}) — 빈 것으로 시작")
    #   «거의 비었다» 의 기준은 씨앗의 절반 미만. 실제 사고 때 저장소가 691건이었다 —
    #   200건 같은 고정 숫자로는 못 잡는다. 정상 저장소(70일치 ≈ 11,000건+)는 씨앗보다 크므로 안 건드린다.
    if name in ("first", "live") and os.path.exists(os.path.join(SEED, f"{name}.json.gz")):
        seed = _load_seed(name)
        if seed and _store_rows(d) < _store_rows(seed) * 0.5:
            for kind in ("con", "serv"):
                merged = dict(seed.get(kind, {}))
                merged.update(d.get(kind, {}))        # 지금 것이 씨앗보다 새것이면 그것이 이깁니다
                d[kind] = merged
            print(f"  ⚠ 저장소 {name} 이 거의 비어 있어 씨앗(data/seed)에서 복구: {_store_rows(d):,}건")
    return d


def save_store(name, data):
    os.makedirs(STORE, exist_ok=True)
    p = os.path.join(STORE, f"{name}.json")
    if name in ("first", "live") and os.path.exists(p):
        try:
            with open(p, encoding="utf-8") as f:
                old = json.load(f)
            n_old = _store_rows({k: v for k, v in old.items() if isinstance(v, dict)})
            n_new = _store_rows(data)
            if n_old >= 1000 and n_new < n_old * 0.5:
                print(f"  ⛔ 저장소 {name} 를 {n_old:,}건 → {n_new:,}건 으로 줄이려 해서 저장을 거부합니다 "
                      f"(수집 실패로 빈 것을 덮어쓰는 사고 방지)")
                return
        except Exception:
            pass
    with open(p, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))


def archive(first, live=None):
    """이번에 받은 개찰 결과를 달별 누적 CSV 에 덧붙이고, 빈칸을 소급해 채운다.

    live 를 함께 받는 이유: 예가범위가 «공고» 쪽에만 실려 오기 때문입니다.
    """
    live = live or {"con": {}, "serv": {}}
    import glob as _glob

    have = set()
    for p in _glob.glob(os.path.join(ARCHIVE_DIR, "extra_*.csv")):
        try:
            with io.open(p, encoding="utf-8-sig", newline="") as f:
                for row in csv.DictReader(f):
                    no = (row.get("공고번호") or "").strip()
                    if no:
                        have.add(no)
        except Exception as e:
            print(f"  ! 누적 CSV 읽기 실패 {os.path.basename(p)} ({type(e).__name__})")

    buckets = {}
    for kind in ("con", "serv"):
        for no, r in first[kind].items():
            if no in have:
                continue
            d = dt_digits(r.get("dt"))
            if len(d) < 8:
                continue
            have.add(no)
            ym = f"{d[0:4]}-{d[4:6]}"
            buckets.setdefault(ym, []).append({
                "공고번호": no,
                "날짜": f"{d[0:4]}-{d[4:6]}-{d[6:8]}",
                "발주기관": r.get("inst", ""),
                "공고명": r.get("name", ""),
                "1순위업체": r.get("win", ""),
                "사업자번호": r.get("bno", "") or "",
                "대표자": r.get("ceo", "") or "",
                "투찰금액": r.get("amt", 0),
                "투찰률": "" if r.get("rate") is None else r.get("rate"),
                "기초금액": r.get("base", "") or "",
                "A값": r.get("aval", "") or "",
                "A값적용": r.get("ayn", "") or "",
                "예가하한": "" if r.get("lo") is None else r.get("lo"),
                "예가상한": "" if r.get("hi") is None else r.get("hi"),
                "참가업체수": r.get("np", "") or "",
            })

    # ══════════════════════════════════════════════════════════════
    #  이미 적힌 줄의 빈칸을, «지금 가진 것»으로 뒤늦게 채웁니다.
    #
    #  ⚠️ 2026-09-03 소장님 지적: 「오늘부터 쌓이는 게 아니라,
    #     조달청에서 가져와 지금 사이트에 있는 건 다 채워야 맞지 않아?」 — 맞습니다.
    #
    #  채우는 것: A값 · A값적용 · 예가하한 · 예가상한 · 참가업체수
    #  채우는 재료: data/store/{first,live}.json — **이미 받아 둔 것**입니다.
    #     → 조달청 호출 0번. 돈도 시간도 안 듭니다.
    #
    #  왜 live 까지 보나: 예가범위는 «공고» 쪽에만 실려 옵니다.
    #     개찰 저장소에는 12.6% 뿐인데 공고 저장소에는 77.8% 가 있습니다.
    #     같은 공고번호로 이어 붙이면 그만큼이 살아납니다.
    #
    #  ⚠️ 옛 달(4~6월)은 공고 저장소가 45일치라 재료가 없습니다.
    #     그건 이 함수로 못 채웁니다 — 애초에 가진 적이 없는 값입니다.
    #     시뮬레이션이 보는 창은 최근 30일이라 실질 손해는 없습니다.
    # ══════════════════════════════════════════════════════════════
    try:
        known = {}
        # 뒤에 넣는 쪽이 이깁니다 — 개찰(first)이 공고(live)보다 확정값입니다.
        for st in (live, first):
            for kind in ("con", "serv"):
                for no, r in (st.get(kind) or {}).items():
                    if not isinstance(r, dict):
                        continue
                    cur = known.setdefault(no, {})
                    for src, dst in (("aval", "A값"), ("ayn", "A값적용"),
                                     ("lo", "예가하한"), ("hi", "예가상한"),
                                     ("np", "참가업체수")):
                        v = r.get(src)
                        if v is not None and v != "" and v != 0:
                            cur[dst] = v
        known = {k: v for k, v in known.items() if v}

        if known:
            filled = {c: 0 for c in ("A값", "예가하한", "참가업체수")}
            n_files = 0
            for path in sorted(_glob.glob(os.path.join(ARCHIVE_DIR, "extra_*.csv"))):
                with io.open(path, encoding="utf-8-sig", newline="") as f:
                    rows = list(csv.DictReader(f))
                if not rows:
                    continue
                ch = False
                for row in rows:
                    k = known.get((row.get("공고번호") or "").strip())
                    if not k:
                        continue
                    # A값 — 적용여부(N)도 값이므로 둘을 한 쌍으로 봅니다
                    if k.get("A값") and not str(row.get("A값") or "").strip():
                        row["A값"] = k["A값"]
                        row["A값적용"] = k.get("A값적용", "")
                        filled["A값"] += 1
                        ch = True
                    # 예가범위 — 하한만 있고 상한이 없는 일은 없습니다. 쌍으로 채웁니다.
                    if k.get("예가하한") is not None and not str(row.get("예가하한") or "").strip():
                        row["예가하한"] = k["예가하한"]
                        row["예가상한"] = k.get("예가상한", "")
                        filled["예가하한"] += 1
                        ch = True
                    if k.get("참가업체수") and not str(row.get("참가업체수") or "").strip():
                        row["참가업체수"] = k["참가업체수"]
                        filled["참가업체수"] += 1
                        ch = True
                # 바뀐 파일만 다시 씁니다.
                # (안 바뀐 파일까지 쓰면 하루 21번 커밋이 통째로 부풀어 오릅니다)
                if ch:
                    n_files += 1
                    with io.open(path, "w", encoding="utf-8-sig", newline="") as g:
                        w = csv.DictWriter(g, fieldnames=ARCH_COLS, extrasaction="ignore")
                        w.writeheader()
                        for row in rows:
                            w.writerow({c: row.get(c, "") for c in ARCH_COLS})
            if any(filled.values()):
                print(f"  → 누적 CSV 소급 기록 ({n_files}개 파일) — "
                      f"A값 {filled['A값']:,} · 예가범위 {filled['예가하한']:,} · "
                      f"참가업체수 {filled['참가업체수']:,}칸 (조달청 호출 0번)")
    except Exception as e:
        print(f"  ! 소급 기록 실패 ({type(e).__name__}: {e}) — 넘어갑니다")

    if not buckets:
        print("  · 누적 CSV — 새로 추가할 건 없음")
        return

    added = 0
    for ym, rows in sorted(buckets.items()):
        p = archive_path(ym)
        fresh = not os.path.exists(p)

        # 칸이 늘어났을 때(사업자번호·대표자 추가) 옛 파일을 한 번 갈아끼웁니다.
        # 그냥 덧붙이면 머리글은 8칸인데 줄은 10칸이 되어 자료가 어긋납니다.
        if not fresh:
            try:
                with io.open(p, encoding="utf-8-sig", newline="") as f:
                    rd = csv.DictReader(f)
                    head = rd.fieldnames or []
                    if head != ARCH_COLS:
                        keep = [dict(x) for x in rd]
                        with io.open(p, "w", encoding="utf-8-sig", newline="") as g:
                            w = csv.DictWriter(g, fieldnames=ARCH_COLS,
                                               extrasaction="ignore")
                            w.writeheader()
                            for x in keep:
                                w.writerow({c: x.get(c, "") for c in ARCH_COLS})
                        print(f"  · {os.path.basename(p)} 칸 맞춤 "
                              f"({len(head)} → {len(ARCH_COLS)}칸, {len(keep):,}줄)")
            except Exception as e:
                print(f"  ! 누적 CSV 칸 맞춤 실패 {os.path.basename(p)} ({type(e).__name__})")

        with io.open(p, "a", encoding="utf-8-sig", newline="") as f:
            w = csv.DictWriter(f, fieldnames=ARCH_COLS, extrasaction="ignore")
            if fresh:
                w.writeheader()
            w.writerows(rows)
        added += len(rows)
    months = ", ".join(sorted(buckets))
    print(f"  · 누적 CSV +{added:,}건 ({months}) — 전체 {len(have):,}건")


def dt_digits(v):
    return re.sub(r"[^0-9]", "", str(v or ""))[:12]


def trim(bucket, days, date_field):
    """오래된 건 버린다 — 안 버리면 파일이 해마다 조용히 무거워진다"""
    cut = (datetime.now(KST) - timedelta(days=days)).strftime("%Y%m%d%H%M")
    return {k: v for k, v in bucket.items() if (dt_digits(v.get(date_field)) or "999") >= cut}




# ══════════════════════════════════════════════════════════════
#  🎯 공고 고르기 — «예상 참가 · 이런 자리 1순위율» 재료 (2026-09-03)
#
#  소장님: 「실격이 더 되더라도 1건이라도…」 → 실측 8,406건: 사정률 분위를 어떻게 잡아도 1순위율은
#  3.5~4.4% 에서 안 움직였다. 움직이는 건 «참가업체수» 뿐이다 (2~9곳 18% · 100곳+ 1.6%).
#  그래서 금액이 아니라 «어느 공고에 넣느냐» 를 돕는다. 재료 둘, 전부 개찰 저장소(70일)에서 센다:
#   ① 기관별 참가업체수 중앙(enp) — 그 기관 개찰이 6건 이상일 때만. 실측: 중앙<10 이면 실제<10 이 72%.
#   ② 규모(기초금액)×참가 칸별 «권장 금액의 1순위율» 표 — 화면이 «이런 자리 1순위 x%» 로 쓴다.
#      15건 미만 칸은 null (없는 숫자를 만들지 않는다).
#  ⚠️ 표의 금액 계산은 build_json._baro_amount / bidmath.js shownBid 와 같은 길(K 0.674·여유 0.3%·올림 한 번).
# ══════════════════════════════════════════════════════════════
PICK_P50 = 99.896
PICK_SZ = [1e8, 3e8, 1e9]          # <1억 · 1~3억 · 3~10억 · 10억+
PICK_NB = [10, 30, 100]            # 2~9 · 10~29 · 30~99 · 100+
PICK_LLR = [(0, 1e9, 89.745), (1e9, 5e9, 88.745), (5e9, 1e10, 87.495)]
PICK_AUTO = {0: (0.8416, 1.0), 1: (0.2533, 1.0)}   # 참가묶음 → (z, 여유). bidmath.js AUTO_RULE 와 같이 고칠 것


def _pick_bucket(v, edges):
    for i, e in enumerate(edges):
        if v < e:
            return i
    return len(edges)


def _pick_llr(r):
    llr = r.get("llr")
    if llr and 60 <= float(llr) <= 100:
        return float(llr)
    est = r.get("est") or (float(r.get("base") or 0) / 1.1)
    for lo, hi, v in PICK_LLR:
        if lo <= est < hi:
            return v
    return None


def pick_stats(fstore):
    """(기관별 예상 참가, 규모×참가 1순위율 표) — 개찰 저장소만으로."""
    import math as _m
    from statistics import median as _med
    by_inst = {}
    cells = {}
    for r in (fstore.get("con") or {}).values():
        np_ = int(r.get("np") or 0)
        inst = str(r.get("inst") or "").strip()
        if np_ > 0 and inst:
            by_inst.setdefault(inst, []).append(np_)
        b = float(r.get("base") or 0); amt = float(r.get("amt") or 0); rate = r.get("rate")
        lo, hi = r.get("lo"), r.get("hi")
        if not b or not amt or not rate or lo is None or hi is None or float(hi) <= float(lo) or np_ < 2:
            continue
        if r.get("ayn") == "N":
            A = 0.0
        elif (r.get("aval") or 0) > 0:
            A = float(r["aval"])
        else:
            continue
        llr = _pick_llr(r)
        if not llr:
            continue
        yeje = amt / (float(rate) / 100)
        sj = yeje / b * 100
        if not (95 <= sj <= 105):
            continue
        w = float(hi) - float(lo)
        sd = _m.sqrt((w * w / 12) / 4 * 11 / 14)
        nb = _pick_bucket(np_, PICK_NB)
        # ★ 공고별 자동 분위(bidmath.js AUTO_RULE 와 같은 규칙): 참가 2~9곳 → 80분위, 10~29곳 → 60분위, 그 밖 → 권장(75+0.3%)
        #   표의 1순위율은 «화면이 실제로 내는 금액» 으로 재야 하므로 여기서도 같은 규칙을 씁니다.
        z, mg = PICK_AUTO.get(nb, (0.674, 1.003))
        sjq = round((PICK_P50 + z * sd) * 1000) / 1000
        M = _m.ceil(_m.ceil((b * sjq / 100 - A) * llr / 100 + A) * mg)
        rt = _m.ceil(M / (b * PICK_P50 / 100) * 100 * 1000) / 1000
        M = _m.ceil(b * PICK_P50 / 100 * rt / 100)
        L = _m.ceil((yeje - A) * llr / 100 + A)
        key = f"s{_pick_bucket(b, PICK_SZ)}n{nb}"
        c = cells.setdefault(key, [0, 0])
        c[0] += 1
        if M >= L and M < amt:
            c[1] += 1
    enp = {k: [int(_med(v)), len(v)] for k, v in by_inst.items() if len(v) >= 6}
    tbl = {k: ([n, round(w / n * 100, 1)] if n >= 15 else None) for k, (n, w) in cells.items()}
    return enp, {"tbl": tbl, "sz": PICK_SZ, "nb": PICK_NB, "n": sum(n for n, _ in cells.values())}


# ══════════════════════════════════════════════════════════════════
#  지역 판정 — 2026-09-05
#
#  ⚠️ 왜 새로 만들었나 (소장님: 「철콘·전남·고흥 하면 아무것도 안 나와」)
#     지역을 «기관명+공고명에 그 글자가 들어 있나» 로 판정하고 있었습니다.
#     그런데 전남과 광주가 통합되어 기관명이 「전남광주통합특별시 장흥군」 이 됐습니다.
#     → 「광주」 를 고르면 962건이 나오는데 그 중 833건(87%)이 전남 시·군이었습니다.
#     조달청은 cnstrtsiteRgnNm(공사 현장 지역, store 의 site)을 8,164건에 주고 있었습니다.
#     (CLAUDE.md 1번 「조달청이 주는 값이 있으면 그대로 쓴다」)
#
#  순서: ① 현장지역(site) 앞부분 → ② 기관명 앞부분 → ③ 기관명 안의 시도 낱말
#        → ④ 기관명·공고명 안의 시·군 이름 (표는 «자료에서» 만듭니다)
#  실측: 못 정하는 것 5.8%. 광주 962 → 137건, 전남 1,015 → 964건.
# ══════════════════════════════════════════════════════════════════
SIDO_FULL = [("서울특별시", "서울"), ("부산광역시", "부산"), ("대구광역시", "대구"),
             ("인천광역시", "인천"), ("광주광역시", "광주"), ("대전광역시", "대전"),
             ("울산광역시", "울산"), ("세종특별자치시", "세종"), ("경기도", "경기"),
             ("강원특별자치도", "강원"), ("강원도", "강원"), ("충청북도", "충북"),
             ("충청남도", "충남"), ("전북특별자치도", "전북"), ("전라북도", "전북"),
             ("전라남도", "전남"), ("경상북도", "경북"), ("경상남도", "경남"),
             ("제주특별자치도", "제주"), ("제주도", "제주")]
SIDO_SCAN = [("서울", "서울"), ("부산", "부산"), ("대구", "대구"), ("인천", "인천"),
             ("대전", "대전"), ("울산", "울산"), ("세종", "세종"), ("경기", "경기"),
             ("강원", "강원"), ("충청북도", "충북"), ("충북", "충북"),
             ("충청남도", "충남"), ("충남", "충남"), ("전라북도", "전북"), ("전북", "전북"),
             ("전라남도", "전남"), ("전남", "전남"), ("경상북도", "경북"), ("경북", "경북"),
             ("경상남도", "경남"), ("경남", "경남"), ("제주", "제주"), ("광주", "광주")]
MERGED = "전남광주통합특별시"          # 전남·광주 통합 (2026)
GWANGJU_GU = {"동구", "서구", "남구", "북구", "광산구", "광주청사"}


def _head_sido(s):
    """문자열 «앞부분» 으로만 판정합니다. 안에 든 글자로 짐작하지 않습니다."""
    s = (s or "").strip()
    if not s:
        return ""
    if s.startswith(MERGED):
        t = s[len(MERGED):].strip()
        head = t.split()[0] if t else ""
        if head in GWANGJU_GU:
            return "광주"
        if head and (head[-1] in "시군" or head == "무안청사"):
            return "전남"
        return "전남,광주"        # 교육청·소방본부처럼 광역 기관은 둘 다에 걸립니다
    for full, ab in SIDO_FULL:
        if s.startswith(full):
            return ab
    return ""


def region_book(rows):
    """시·군 → 시도 표를 «자료에서» 만듭니다. 손으로 226개를 적어 두면 언젠가 어긋납니다."""
    book = {}
    for r in rows:
        for src in (str(r.get("site") or ""), str(r.get("inst") or "")):
            ab = _head_sido(src)
            if not ab or "," in ab:
                continue
            parts = src.split()
            if len(parts) >= 2 and parts[1] and parts[1][-1] in "시군구" and len(parts[1]) >= 3:
                book.setdefault(parts[1], {})
                book[parts[1]][ab] = book[parts[1]].get(ab, 0) + 1
    out = {}
    for gun, c in book.items():
        ab, n = max(c.items(), key=lambda kv: kv[1])
        if n >= 2:                       # 한 번만 나온 것은 오타일 수 있어 뺍니다
            out[gun] = ab
    return out


def sido_of(r, book=None):
    """이 공고가 어느 시도인가. 여럿이면 «전남,광주» 처럼 쉼표로 잇습니다."""
    for src in (str(r.get("site") or ""), str(r.get("inst") or "")):
        ab = _head_sido(src)
        if ab:
            return ab
    inst = str(r.get("inst") or "")
    for tok, ab in SIDO_SCAN:
        if tok in inst:
            return ab
    if book:
        for gun, ab in book.items():
            if gun in inst:
                return ab
        nm = str(r.get("name") or "")
        for gun, ab in book.items():
            if gun in nm:
                return ab
    return ""



# ══════════════════════════════════════════════════════════════════
#  내역서 모음 — 2026-09-05
#
#  조달청은 공고마다 붙임 파일 이름과 «내려받기 주소»를 줍니다
#  (ntceSpecFileNm1~10 / ntceSpecDocUrl1~10). 우리는 이미 store 의 docs 에
#  받아 두고 있었는데 화면에서 한 번도 쓰지 않고 있었습니다.
#  실측: 공고 12,625건 중 7,070건(56%)에 붙임 22,624개 · 그 중 «내역» 4,235개.
#
#  ⚠️ 파일 «주소» 는 조달청이 준 것을 그대로 씁니다 (CLAUDE.md 1번).
#     2026-09-05 부터, 단가가 든 갈래에 한해 파일 자체도 받아 둡니다
#     (소장님: 「파일은 퍼 와도 돼, 사이트에서 사용자가 다운 받을수 있게 해줘」).
#     받아 둔 것은 화면에 «출처(발주기관·공고번호)» 와 «삭제 요청 안내» 를 함께 답니다.
# ══════════════════════════════════════════════════════════════════
NAEYEOK_KIND = [
    ("설계내역", "설계내역서"),        # ★ 발주처 설계 단가가 들어 있습니다
    ("단가산출", "단가산출서"),
    ("일위대가", "단가산출서"),
    ("공내역", "공내역서"),            # 단가가 비어 있습니다 (낙찰자가 채움)
    ("물량내역", "물량내역서"),
    ("수량산출", "수량산출서"),
    ("내역", "그 밖의 내역서"),
]
# 단가가 «들어 있는» 갈래 — 첫 화면에서 이것만 받습니다 (가벼움)
NAEYEOK_PRICED = ("설계내역서", "단가산출서")


def naeyeok_kind(name):
    """붙임 파일 이름으로 «어떤 내역서인가» 를 가릅니다.

    ⚠️ 이름에 「무단가」 가 붙은 것이 실제로 16개 있습니다
       (「(무단가)설계내역서.xlsx」). 조달청 파일 이름이 스스로 «단가 없음» 이라고
       말하고 있으므로 짐작이 아닙니다 — 공내역서로 가릅니다.
       이걸 설계내역서로 두면 화면이 «단가 있음» 이라고 거짓말을 합니다.
    """
    n = str(name or "").replace(" ", "")
    # ── «단가가 없다» 는 표시를 먼저 봅니다 ──────────────────────
    #    「설계내역서(공내역서).xlsx」 처럼 두 낱말이 다 든 이름이 실측 227개 중 18개.
    #    설계내역서를 먼저 맞히면 화면이 «단가 있음» 이라고 거짓말을 합니다.
    #    이름이 스스로 «비었다» 고 말하면 그 말을 믿습니다.
    if "무단가" in n or "단가없" in n:
        return "공내역서"
    if "공내역" in n:
        return "공내역서"
    if "물량내역" in n:
        return "물량내역서"
    for key, label in NAEYEOK_KIND:
        if key in n:
            return label
    return ""


# ── 내역서 파일 받기 설정 (2026-09-05) ────────────────────────────
#   실측: 설계내역서 5개를 실제로 받아 재보니 129KB · 96KB · 52KB · 5.4MB · 65KB.
#   4,290개를 다 받으면 1GB 가 넘어 Firebase 무료 10GB 에 부담이 됩니다.
#   그래서 «단가가 든 갈래» 만, 한 회차에 조금씩, 크기 상한을 두고 받습니다.
NAEYEOK_BOOK = os.path.join(STORE, "naeyeok_files.json")
# ── 내역서 목록을 «쌓아» 둡니다. (2026-09-06)
#   소장님: 「딱 3년 치만 저장」 → 비용을 재고 나서 「1년 치만 하자」 로 정정.
#   ⚠️ 기간은 NAEYEOK_KEEP_DAYS 하나로만 정합니다. 화면에 찍는 말도 여기서 가져옵니다 —
#      숫자를 글에 손으로 박아 두면 상한을 바꿨을 때 «말만 옛것» 이 됩니다(실제로 그랬습니다).
#   ⚠️ store(first/live)는 70일치라, 여기에만 기대면 목록도 70일치입니다.
#      그래서 목록만 따로 쌓습니다. 파일 자체가 아니라 «줄» 이라 가볍습니다.
#   ⚠️ 오래 쌓인 것을 화면에 통째로 내면 안 됩니다 — 3년치 실측 gzip 3.4MB 였습니다.
#      저장 기간과 별개로, 화면에 내는 것은 갈래별 상한(NAEYEOK_SHOW)까지입니다.
NAEYEOK_INDEX = os.path.join(STORE, "naeyeok_index.json")
# 소장님 결정(2026-09-06): 「비용이 많이 들면 내역서는 1년 치만 하자」
NAEYEOK_KEEP_DAYS = int(os.environ.get("NAEYEOK_KEEP_DAYS", "365"))    # 1년
# 화면 상한은 «언제 받는 파일인가» 로 갈립니다 — 같은 숫자를 쓰면 한쪽이 반드시 틀립니다.
#   naeyeok.json     : 화면을 열자마자 받습니다 → 작게
#   naeyeok-all.json : 그 갈래를 눌렀을 때만 받습니다 → 크게 잡아도 됩니다
NAEYEOK_SHOW_TOP = int(os.environ.get("NAEYEOK_SHOW_TOP", "800"))     # 단가 든 갈래(첫 화면)
NAEYEOK_SHOW = int(os.environ.get("NAEYEOK_SHOW", "2000"))            # 나머지 갈래(누를 때)
# 받은 파일이 사는 곳. ★ web/public 이 아니라 data/store 입니다 —
#   GitHub Actions 가 회차 사이에 넘겨주는 것이 data/store 하나뿐이기 때문입니다.
#   (자세한 까닭은 fetch_naeyeok_files 의 설명)
NAEYEOK_DIR = os.path.join(STORE, "naeyeok")
# 보관 상한. 넘으면 오래된 것부터 내립니다 (공고는 날마다 새로 나옵니다).
# ⚠️ 300 → 150 으로 내렸습니다 (2026-09-06 실측).
#    배포 1벌이 93MB 인 줄 알았는데 **183MB** 였습니다(공고 8,000장 + 카톡 카드 2,509장).
#    Firebase Hosting 은 «출시 보관 10개» 라, 1벌에 얹히는 것이 10배로 곱해집니다:
#      지금            183MB × 10 = 1.8GB  (무료 10GB 의 18%)
#      + 내역서 300MB              = 4.7GB  (47%)  ← 절반
#      + 내역서 150MB              = 3.3GB  (33%)  ← 이걸로 정함
#    150MB 면 설계내역서 약 750개(다섯 달치)입니다. 그보다 오래된 것은 링크로 남습니다.
NAEYEOK_KEEP_MB = int(os.environ.get("NAEYEOK_KEEP_MB", "150"))
NAEYEOK_FETCH = int(os.environ.get("NAEYEOK_FETCH", "100"))     # 한 회차에 새로 받는 개수
#   40 → 100 (2026-09-06). 실측 40개에 155.7초였으니 100개면 약 390초.
#   회차 전체가 14분이고 상한이 45분이라 여유가 있습니다.
NAEYEOK_MAXBYTES = int(os.environ.get("NAEYEOK_MAXBYTES", str(8 * 1024 * 1024)))
# ⚠️ 시간 상한. 40개 × 60초 타임아웃 = 40분이라 그것만으로 회차(45분)가 터집니다.
#    내역서는 «있으면 좋은 것» 이지 배포를 막아도 되는 것이 아닙니다.
NAEYEOK_BUDGET_S = int(os.environ.get("NAEYEOK_BUDGET_S", "420"))
NAEYEOK_TIMEOUT_S = int(os.environ.get("NAEYEOK_TIMEOUT_S", "25"))


def load_json(path, dflt):
    try:
        with io.open(path, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return dflt


def save_json(path, obj):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    tmp = path + ".tmp"
    with io.open(tmp, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, separators=(",", ":"))
    os.replace(tmp, path)


def xlsx_has_price(path):
    """엑셀을 열어 «단가에 값이 들어 있는지» 확인합니다.

    ⚠️ 이름으로 짐작하지 않습니다 — 「(무단가) 설계내역서」처럼
       이름은 설계내역서인데 단가가 빈 것이 실제로 16개 있었습니다.
    머리글에 «단가» 가 든 열을 찾아, 그 열에 숫자가 몇 개나 있는지 셉니다.
    """
    try:
        from openpyxl import load_workbook
    except Exception:
        return None
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return None
    try:
        for ws in wb.worksheets:
            cols, n = [], 0
            for row in ws.iter_rows(max_row=400, values_only=True):
                if not row:
                    continue
                if not cols:
                    for i, v in enumerate(row):
                        if isinstance(v, str) and "단가" in v:
                            cols.append(i)
                    continue
                for i in cols:
                    if i < len(row) and isinstance(row[i], (int, float)) and row[i] > 0:
                        n += 1
                if n >= 5:
                    return True
        return False
    except Exception:
        return None
    finally:
        try:
            wb.close()
        except Exception:
            pass


def doc_flag(r):
    """붙임 파일에 내역서가 있는가. 2 = 단가가 든 것(설계내역서·단가산출서) · 1 = 내역서 · 0 = 없음."""
    best = 0
    for y in (r.get("docs") or []):
        k = naeyeok_kind(str(y[0] if isinstance(y, (list, tuple)) else y))
        if not k:
            continue
        best = max(best, 2 if k in NAEYEOK_PRICED else 1)
    return best


def lic_pairs(r):
    """조달청 lic 값 «철근ㆍ콘크리트공사업/4994» 를 (코드, 이름) 으로 가릅니다.

    ⚠️ 손으로 이름을 적지 않습니다. 조달청이 준 문자열을 그대로 가릅니다
       (CLAUDE.md 1번). 코드가 없으면 이름 자체를 코드로 씁니다.
    """
    v = r.get("lic") or []
    if not isinstance(v, list):
        v = [v]
    out = []
    for x in v:
        t = str(x).strip()
        if not t:
            continue
        if "/" in t:
            label, code = t.rsplit("/", 1)
            label, code = label.strip(), code.strip()
        else:
            label, code = t, t
        if code:
            out.append((code, label or code))
    return out


def lic_codes(r):
    return [c for c, _ in lic_pairs(r)]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--days", type=int, default=3)
    ap.add_argument("--backfill", type=int, default=0)
    ap.add_argument("--sleep", type=float, default=0.6)
    ap.add_argument("--ranks", type=int, default=-1,
                    help="개찰 순위(1위~꼴찌)를 이번 회차에 몇 건까지 채울지. "
                         "안 주면 시각을 보고 알아서 정합니다(정밀 회차 250 / 그 외 60). 0이면 안 함")
    ap.add_argument("--reranks", type=int, default=0,
                    help="순위는 있는데 추첨번호 분포(drw)가 없는 개찰을 N건 다시 받습니다 (한 번만)")
    ap.add_argument("--fillbsis", type=int, default=0,
                    help="지난 N일치 기초금액·A값을 소급해서 채웁니다 (하루 한 번이면 충분)")
    ap.add_argument("--fillonly", action="store_true",
                    help="조달청을 부르지 않고, 이미 받아 둔 자료로 "
                         "누적 CSV 의 빈칸만 채웁니다 (호출 0번 · 몇 초)")
    ap.add_argument("--exportonly", action="store_true",
                    help="조달청을 부르지 않고 저장소만으로 화면 파일을 다시 굽습니다 "
                         "(색인·면허·지역을 고친 뒤 확인할 때)")
    ap.add_argument("--probe", action="store_true",
                    help="투찰업체 전체를 주는 오퍼레이션이 있는지 한 번 확인만 합니다")
    args = ap.parse_args()
    if args.exportonly:
        # 조달청을 아예 부르지 않습니다. fetch 가 차단기(NET_DOWN)를 보고 바로 []를 냅니다.
        global NET_DOWN, NO_NET
        NET_DOWN = True
        NO_NET = True
        print('  · --exportonly : 조달청을 부르지 않고 저장소로만 다시 굽습니다')

    # ── --fillonly : 조달청을 한 번도 부르지 않습니다 ────────────────
    #   이미 data/store/ 에 받아 둔 것으로 누적 CSV 의 빈칸만 채웁니다.
    #   API 키도 필요 없고, 몇 초면 끝납니다.
    #   («--days 30 이 중간에 끊겼다» 같은 상황에서 그때까지 받은 것만이라도
    #     바로 자료에 반영하려고 만들었습니다)
    if args.fillonly:
        print("=" * 52)
        print("  이미 받아 둔 자료로 누적 CSV 빈칸 채우기 (조달청 호출 0번)")
        print("=" * 52)
        _first = load_store("first")
        _live = load_store("live")
        archive(_first, _live)
        print("완료.")
        return

    load_env()

    key = api_key()

    days = args.backfill or args.days
    today = datetime.now(KST)

    if args.probe:
        # 이미 받아 둔 개찰 중 가장 최근 것을 표본 공고로 씁니다.
        # 공고번호를 받아야 하는 오퍼레이션을 «없다» 고 오판하지 않기 위해서입니다.
        _f = load_store("first")
        _rows = sorted((_f.get("con") or {}).values(),
                       key=lambda r: dt_digits(r.get("dt")), reverse=True)
        _no = _rows[0]["no"] if _rows else ""
        _ord = str((_rows[0].get("ord") if _rows else "") or "000") or "000"
        # 진단은 자료를 바꾸지 않습니다. 조달청이 느린 날 여기서 5분(20호출×15초)을 더 쓰면
        # 뒤 단계(빌드·배포)가 45분 상한에 밀립니다 → 3분만 씁니다.
        global NET_BUDGET_S
        NET_BUDGET_S = min(NET_BUDGET_S, 180)
        probe_ops(key, today - timedelta(days=1), _no, _ord)
        save_diag()
        return

    print("=" * 52)
    print(f"  조달청 수집 — 최근 {days}일")
    print("=" * 52)

    first = load_store("first")
    live = load_store("live")

    # ── 순위 조회 건수를 «지난 회차로부터 얼마나 지났나» 로 정합니다 ──
    #   워크플로 파일(.github/workflows/update.yml)은 보안상 원격에서 못 고칩니다.
    #   그래서 명령에 --ranks 를 안 붙여도 알아서 배분하도록 여기에 둡니다.
    #
    #   ⚠️ 2026-09-04 — 전에는 «한국시간 08시·13시면 250건» 이었습니다. **틀린 설계였습니다.**
    #      GitHub 예약 실행은 정시에 안 옵니다(실측: 예약 21회 중 하루 4회, 10:03·14:40·20:04·22:29).
    #      시각이 08 이나 13 에 딱 걸릴 확률이 낮아, 사실상 매번 60건만 돌고 있었습니다.
    #      → 시각(hour)을 조건으로 쓰지 않습니다. **지난 회차와의 간격**으로 정합니다.
    #        · 3시간 넘게 비었으면(= 그날 첫 회차이거나 오래 밀렸으면) 250건
    #        · 그보다 촘촘하면                                        60건
    #      회차가 몇 번 오든 하루 1,000~1,500건 언저리로 스스로 맞춰집니다.
    #   ⚠️ 손으로 --ranks 를 주면 그 값이 이깁니다 (0 을 주면 순위 조회를 건너뜁니다).
    if args.ranks < 0:
        _now = datetime.now(timezone.utc)
        _prev = first.get("_lastrun") if isinstance(first, dict) else None
        _gap = None
        if _prev:
            try:
                _gap = (_now - datetime.fromisoformat(_prev)).total_seconds() / 3600.0
            except Exception:
                _gap = None
        args.ranks = 60 if (_gap is not None and _gap < 3) else 250
        print(f"  · 개찰 순위 조회: 이번 회차 {args.ranks:,}건 "
              f"(지난 회차와 {'%.1f시간' % _gap if _gap is not None else '기록 없음'} 간격)")
        first["_lastrun"] = _now.isoformat(timespec="seconds")

    # 다루지 않기로 한 종류(용역)는 저장소에서도 비웁니다.
    # 안 그러면 안 쓰는 자료가 70일 동안 남아 파일만 무거워집니다.
    for _st in (first, live):
        for _k in list(_st.keys()):
            # ⚠️ 사전(dict)인 칸만 비웁니다. 그냥 «KINDS 에 없으면 비우기» 로 두면
            #    _lastrun 같은 기록용 값까지 {} 로 지워집니다(2026-09-04에 실제로 그랬습니다).
            if _k not in KINDS and isinstance(_st[_k], dict):
                _st[_k] = {}
    added = {"first": 0, "live": 0}

    n_base = 0
    n_lic = 0
    n_aval = 0
    n_win = 0
    scan = days_to_scan(today, days)
    _extra = len(scan) - days
    if _extra > 0:
        print("  · 기록에 빠져 있는 날을 다시 훑습니다: "
              + ", ".join(d.strftime("%m-%d") for d in scan[:_extra]))
    daylog = load_days()
    for day in scan:
        ds = day.strftime("%m-%d")
        day_ok = True
        for kind in KINDS:
            _rows, _why = fetch_paged(ENDPOINTS[("first", kind)], key, day,
                                      label=f"개찰 {kind} {ds}")
            if _why:
                day_ok = False
            for item in _rows:
                r = row_first(item)
                if r:
                    prev = first[kind].get(r["no"]) or {}
                    # 이미 받아둔 기초금액을 덮어쓰지 않는다
                    for f in ("base", "lo", "hi", "lic",
                              "aval", "aparts", "ayn", "gmtrl", "np",
                              "llr", "est", "ptot", "pdrw"):
                        if prev.get(f) is not None:
                            r[f] = prev[f]
                    # ⚠️ 2026-09-03 — 순위 조회로 채운 것을 개찰목록 수집이 덮어쓰면 안 됩니다.
                    #    개찰목록 API 의 corps 는 «1순위 한 곳»뿐이라, 그대로 두면 힘들게 받은
                    #    30곳 순위(corps)·사다리(rq)·전체건수(nrank)가 통째로 날아갑니다.
                    #    실제로 그래서 store 에 순위가 0건이었습니다 — bidresult 엔 46건 남아 있었는데도.
                    #    (조회 직후 export 된 건 남고, 다음 회차 목록 수집이 원본을 지웠던 것)
                    #    rask(순위 조회를 마친 표시)가 있으면 순위 관련 값을 지킵니다.
                    if prev.get("rask"):
                        for f in ("corps", "rq", "nrank", "rask", "drw",
                                  "win", "amt", "rate", "bno", "ceo"):
                            if prev.get(f) is not None:
                                r[f] = prev[f]
                    first[kind][r["no"]] = r
                    added["first"] += 1
            time.sleep(args.sleep)
            _rows, _why = fetch_paged(ENDPOINTS[("live", kind)], key, day,
                                      label=f"공고 {kind} {ds}")
            if _why:
                day_ok = False
            for item in _rows:
                r = row_live(item)
                if r:
                    prev = live[kind].get(r["no"]) or {}
                    for f in ("base", "lo", "hi"):
                        if prev.get(f) is not None:
                            r[f] = prev[f]
                    live[kind][r["no"]] = r
                    added["live"] += 1
            time.sleep(args.sleep)

            # ── 기초금액: 하루치를 통째로 받아 두 저장소에 같이 붙인다 ──
            bm = bsis_by_day(key, day, kind)
            for store in (first, live):
                for no, b in bm.items():
                    row = store[kind].get(no)
                    if row is None:
                        continue
                    if not row.get("base"):
                        row.update(b)
                        n_base += 1
                    else:
                        # ⚠️ 기초금액은 이미 있는데 «다른 칸»이 빈 경우.
                        #    2026-09-03 이전에는 A값만 채우고 예가범위는 안 채웠습니다.
                        #    그래서 base 를 먼저 받아 둔 공고는 lo/hi 를 영영 못 받았습니다.
                        #    응답에 이미 실려 온 값인데 버리고 있었던 것입니다.
                        #    («조달청이 주는 값은 그대로 쓴다» — 버리는 것도 안 쓰는 것입니다)
                        hit = False
                        for f in ("lo", "hi", "aval", "aparts", "ayn", "gmtrl", "lic"):
                            if row.get(f) is None and b.get(f) is not None:
                                row[f] = b[f]
                                hit = True
                        if hit:
                            n_aval += 1
            time.sleep(args.sleep)

            # ── 낙찰자 상세 (주소·전화·참가업체수) ──
            sm = scsbid_by_day(key, day, kind)
            for no, v in sm.items():
                row = first[kind].get(no)
                if row is not None:
                    for f, val in v.items():
                        if not row.get(f):
                            row[f] = val
                    n_win += 1
            time.sleep(args.sleep)

            # ── 면허·업종 제한 ──────────────────────
            # 면허제한은 «등록일» 기준으로 옵니다.
            # 그날 공고만 붙이면 대부분 비어 있게 됩니다 — 저장소 전체에 맞춰봅니다.
            lm = lic_by_day(key, day, kind)
            for no, names in lm.items():
                for store in (live, first):
                    row = store[kind].get(no)
                    if row is not None and not row.get("lic"):
                        row["lic"] = names[:6]
                        n_lic += 1
            time.sleep(args.sleep)

        print(f"  {ds}  1순위 {len(first['con']) + len(first['serv']):,}건 "
              f"/ 공고 {len(live['con']) + len(live['serv']):,}건 "
              f"/ 기초금액 {n_base:,}건 / 빈칸메움 {n_aval:,}건 "
              f"/ 면허 {n_lic:,}건 / 낙찰자상세 {n_win:,}건 누적")

        # ⚠️ 2026-09-03 — 하루가 끝날 때마다 저장합니다.
        #   전에는 «맨 끝에 한 번»만 저장했습니다. 그래서 --days 30 을 돌리다
        #   Ctrl+C 로 멈추면 16일치 조회가 통째로 날아갔습니다 (실제로 겪음).
        #   조달청을 다시 부르는 건 시간도 호출한도도 쓰는 일입니다. 버리면 안 됩니다.
        #   쓰는 비용은 7.8MB 두 번 — 0.3초쯤입니다. 날리는 쪽이 훨씬 비쌉니다.
        try:
            save_store("first", first)
            save_store("live", live)
        except Exception as e:
            print(f"    ! 중간 저장 실패 ({type(e).__name__}) — 계속합니다")

        # 이 날짜를 «받았다» 고 적어 둡니다. 실패한 날은 적지 않으므로
        # 다음 회차가 «빠진 날» 로 보고 다시 훑습니다.
        daylog[day.strftime("%Y-%m-%d")] = {
            "ok": bool(day_ok),
            "at": datetime.now(KST).strftime("%Y-%m-%d %H:%M"),
        }
        save_days(daylog)

    # ── 화면에 실릴 최근 건 중 기초금액이 빈 것만 공고번호로 개별 보충 ──
    todo = []
    for kind in KINDS:
        # 마감 전 공고를 마감 임박 순으로 — 사람들이 실제로 계산하는 순서입니다
        live_rows = [r for r in live[kind].values()
                     if dt_digits(r.get("close")).ljust(14, "0")
                     >= datetime.now(KST).strftime("%Y%m%d%H%M%S")]
        live_rows.sort(key=lambda r: dt_digits(r.get("close")))
        for r in live_rows:
            if not r.get("base") or not r.get("aval"):
                todo.append((kind, r))
        # 그다음 개찰 목록에서 기초금액이 빈 것
        rows = sorted(trim(first[kind], SHOW_DAYS, "dt").values(),
                      key=lambda r: dt_digits(r.get("dt")), reverse=True)
        for r in rows[:MAX_ROWS]:
            if not r.get("base"):
                todo.append((kind, r))
    seen = set()
    uniq = []
    for kind, r in todo:
        k = (kind, r["no"])
        if k not in seen:
            seen.add(k)
            uniq.append((kind, r))
    # ⚠️ 2026-09-02 — 여기가 조달청 호출의 85% 를 먹고 있었습니다.
    #   빈 건이 1,931건인데 회차마다 «목록 앞에서 250건» 만 물어봤습니다.
    #   목록 순서가 매번 같으니 **같은 250건을 하루 21번 다시 물었고**,
    #   뒤쪽 1,681건은 영영 차례가 오지 않았습니다.
    #   고침: «마지막으로 물어본 시각»을 적어 두고 **오래된 것부터** 돌아가며 묻습니다.
    #   250건씩 돌아가면 1,931건을 8회차(약 4시간)면 한 바퀴 다 돕니다.
    #   ⚠️ 횟수 제한은 두지 않습니다 — 기초금액은 마감 직전에 공개되는 공고가 있어서,
    #      한 번 없다고 끊으면 영영 못 받습니다.
    now_ts = datetime.now(KST).strftime("%Y%m%d%H%M%S")
    uniq.sort(key=lambda kr: str(kr[1].get("bask") or ""))
    if uniq:
        print(f"  · 기초금액 개별조회 {min(len(uniq), BSIS_ONE_CAP):,}건 "
              f"(빈 건 {len(uniq):,} — 오래 안 물어본 것부터)")
        for kind, r in uniq[:BSIS_ONE_CAP]:
            b = bsis_one(key, r["no"], kind)
            r["bask"] = now_ts          # 물어본 시각 (다음 회차에서 뒤로 밀립니다)
            if b:
                had = bool(r.get("base"))
                r.update(b)
                if had:
                    n_aval += 1
                else:
                    n_base += 1
            time.sleep(args.sleep / 2)
    print(f"  · 기초금액 확보 {n_base:,}건")

    # ⚠️ 순서 중요: 잘라내기 전에 누적 CSV 로 먼저 옮긴다
    # ── 소급 보충 ────────────────────────────────────────────────
    #  A값이 비어 있으면 시뮬레이션이 «가정» 으로 돌아갑니다.
    #  그래서 «시켜야 하는» 일로 두지 않고, **모자라면 스스로 채우게** 합니다.
    #    · 최근 45일 개찰 중 A값(또는 A값미적용 표시)을 아는 비율이 60% 미만이면 자동 실행
    #    · 하루 한 번만 (data/store/.fill 에 날짜를 적어 둡니다)
    #    · --fillbsis N 을 주면 조건 없이 N일치 실행
    auto_days = 0
    try:
        cut45 = (datetime.now(KST) - timedelta(days=45)).strftime("%Y%m%d%H%M")
        recent = [r for r in first.get("con", {}).values()
                  if (dt_digits(r.get("dt")) or "0") >= cut45]
        if recent:
            known = sum(1 for r in recent if r.get("aval") or r.get("ayn"))
            cov = known / len(recent)
            mark = os.path.join(STORE, ".fill")
            today = datetime.now(KST).strftime("%Y-%m-%d")
            done_today = os.path.exists(mark) and open(mark).read().strip() == today
            print(f"  · 최근 45일 개찰 {len(recent):,}건 중 A값 아는 것 "
                  f"{known:,}건 ({cov*100:.1f}%)")
            if cov < 0.60 and not done_today:
                auto_days = 45
                print("  · A값이 모자랍니다 — 소급 보충을 스스로 돌립니다")
            elif cov < 0.60:
                print("  · 오늘 이미 소급 보충을 했습니다 — 건너뜁니다")
    except Exception as e:
        print(f"  ! 커버리지 확인 실패 ({type(e).__name__}) — 넘어갑니다")

    fill_days = args.fillbsis or auto_days
    if fill_days > 0 and not NET_DOWN:
        print("-" * 52)
        try:
            backfill_bsis(key, first, live, fill_days, args.sleep)
            try:
                with open(os.path.join(STORE, ".fill"), "w") as f:
                    f.write(datetime.now(KST).strftime("%Y-%m-%d"))
            except Exception:
                pass
            save_store("first", first)
            save_store("live", live)
        except Exception as e:
            print(f"  ! 소급 보충 실패 ({type(e).__name__}: {e}) — 넘어갑니다")

    # ── ★ 개찰 순위 (1위~꼴찌) — 공고번호로 하나씩 받아 채웁니다 ──
    #   날짜로 부르면 안 옵니다. 공고번호로만 옵니다.
    #   한 회차에 args.ranks 건만 채웁니다 (하루 21회차 × 40건 = 840건,
    #   하루 개찰이 약 570건이므로 하루면 다 따라잡습니다).
    #   최근 개찰부터, 아직 1곳뿐인 것만 채웁니다.
    if (args.ranks > 0 or args.reranks > 0) and not NET_DOWN:
        todo_rank = []
        for kind in KINDS:
            rows = sorted(trim(first[kind], SHOW_DAYS, "dt").values(),
                          key=lambda r: dt_digits(r.get("dt")), reverse=True)
            for r in rows:
                if len(r.get("corps") or []) > 1:
                    continue            # 이미 순위가 붙은 공고는 건너뜁니다
                if r.get("nrank") == 1:
                    continue            # 참가업체가 정말 1곳인 공고 (다시 안 물어봅니다)
                todo_rank.append(r)
        # ── --reranks N : 순위는 받았지만 «추첨번호 분포(drw)» 가 없는 개찰을 다시 묻습니다.
        #   추첨번호를 담기 전(2026-09-03 이전)에 받은 순위가 여기에 해당합니다.
        #   한 번만 쓰는 옵션입니다 — 새로 받는 순위에는 처음부터 drw 가 들어갑니다.
        if args.reranks > 0:
            redo = []
            for kind in KINDS:
                for r in trim(first[kind], SHOW_DAYS, "dt").values():
                    if len(r.get("corps") or []) > 1 and not r.get("drw"):
                        redo.append(r)
            redo.sort(key=lambda r: dt_digits(r.get("dt")), reverse=True)
            print(f"  · 추첨번호 다시 받기 대상 {len(redo):,}건 중 {min(len(redo), args.reranks):,}건")
        else:
            redo = []
        # ⚠️ 기초금액 개별조회에서 겪은 것과 같은 함정을 피합니다.
        #   «앞에서 N건» 만 집으면 응답이 안 오는 공고를 매 회차 다시 묻고,
        #   뒤쪽은 영영 차례가 안 옵니다.
        #   ① 아직 한 번도 안 물어본 것 — 최근 개찰부터
        #   ② 물어봤는데 못 받은 것 — 오래된 것부터 돌아가며
        def _rank_key(r):
            asked = str(r.get("rask") or "")
            dt = (dt_digits(r.get("dt")) or "0")[:14].ljust(14, "0")
            #   ① 안 물어본 것(0) 먼저, ② 물어본 것은 오래된 순,
            #   ③ 같은 조건이면 최근 개찰부터 (문자열을 뒤집어 내림차순)
            return (1 if asked else 0, asked, "".join(chr(9 - int(c)) for c in dt))
        todo_rank.sort(key=_rank_key)
        # 다시 받을 것(추첨번호 없음)을 맨 앞에 — 정렬에 섞이면 rask 가 있어 뒤로 밀립니다
        todo_rank = redo[:args.reranks] + todo_rank[:args.ranks]
        got = ranked = 0
        for r in todo_rank:
            if NET_DOWN:
                break
            cs, total, ladder, drw = openg_ranks(key, r["no"], r.get("ord"))
            r["rask"] = datetime.now(KST).strftime("%Y%m%d%H%M%S")
            time.sleep(args.sleep)
            if not cs:
                continue
            got += 1
            r["corps"] = cs          # 낮은 금액 순 30곳까지
            r["nrank"] = total       # 실제로 받은 전체 투찰 건수
            r["rq"] = ladder         # [[등수, 금액], ...] — 전 구간 사다리
            if any(drw):
                r["drw"] = drw       # 1~15번이 각각 몇 번 찍혔나 (전체 투찰자 기준)
            # 100건마다 저장 — 한 번에 1,500건을 받다가 끊겨도 그때까지는 남습니다
            if got % 100 == 0:
                try:
                    save_store("first", first)
                    print(f"    · 순위 {got:,}건 받음 (중간 저장)")
                except Exception:
                    pass
            if len(cs) > 1:
                ranked += 1
                # 1순위 정보도 조달청 순위 자료로 맞춰 둡니다 (더 정확합니다)
                r["win"], r["amt"] = cs[0][0], cs[0][1]
                if cs[0][2]:
                    r["rate"] = cs[0][2]
                if len(cs[0]) > 3 and cs[0][3]:
                    r["bno"] = cs[0][3]
                if len(cs[0]) > 4 and cs[0][4]:
                    r["ceo"] = cs[0][4]
        if todo_rank:
            print(f"  → 개찰 순위 조회 {len(todo_rank):,}건 시도 · "
                  f"응답 {got:,}건 · 2곳 이상 {ranked:,}건")

    # 사람이 넣어 둔 전체 투찰내역이 있으면 여기서도 붙입니다 (파일로 받은 경우)
    merge_ranks(first)

    # ── 공고(live)에만 실려 오는 값을 개찰(first)에 이어 붙입니다 ──────────
    #   ★ 2026-09-03 소장님: 「바로투찰하고 1순위 채점에서 권장투찰가 금액이 달라.」
    #   원인 중 하나가 여기였습니다. 낙찰하한율(llr)·추정가격(est)은 «공고» 쪽에만 오고
    #   개찰 저장소에는 0건이었습니다(실측 11,542건 중 0). 그래서 바로투찰은 공고서의
    #   하한율(예: 87.675%)로 금액을 내고, 채점은 규모로 추정한 하한율(87.745%)로 다시
    #   계산했습니다 — 같은 공고인데 두 화면의 «우리 금액»이 달랐습니다.
    #   실측: 공고서 하한율이 규모 추정과 다른 공고가 6,212건 중 128건(2.1%), 최대 3.7% 차이.
    #   조달청 호출 0번 — 이미 받아 둔 공고 저장소에서 같은 공고번호로 옮겨 적습니다.
    joined = 0
    for kind in KINDS:
        lv = live.get(kind) or {}
        for no, r in (first.get(kind) or {}).items():
            src = lv.get(no)
            if not isinstance(src, dict):
                continue
            #   기초금액·예가범위·A값도 «공고엔 있는데 개찰엔 없는» 줄이 있다(실측 9,703건 중 77건).
            #   같은 공고번호면 같은 값이다. 채점이 «값 부족» 으로 빠질 이유가 없다.
            #   site(공사 현장 지역)도 개찰 쪽에는 0건입니다 — 공고에서 가져오면
            #   개찰 화면의 «지역 못 정함» 이 13.0% → 7.0% 로 줄어듭니다 (실측 11,638건).
            for f in ("llr", "est", "ptot", "pdrw",
                      "base", "lo", "hi", "aval", "ayn", "aparts", "gmtrl", "lic",
                      "site"):
                if (r.get(f) in (None, "", 0, [])) and src.get(f) not in (None, "", 0, []):
                    r[f] = src[f]
                    joined += 1
    if joined:
        print(f"  · 공고→개찰 이어붙임: 낙찰하한율·추정가격 등 {joined:,}칸")

    archive(first, live)

    for kind in ("con", "serv"):
        first[kind] = trim(first[kind], KEEP_DAYS, "dt")
        live[kind] = trim(live[kind], KEEP_DAYS, "dt")

    save_store("first", first)
    save_store("live", live)

    # 사이트용으로 최근분만 잘라서 내보낸다
    os.makedirs(OUT, exist_ok=True)
    built = datetime.now(KST).strftime("%Y-%m-%d %H:%M")

    def export(name, store, date_field):
        out = {"built": built}
        for kind in ("con", "serv"):   # serv 는 수집하지 않으므로 항상 [] — 형식 유지용
            rows = list(trim(store[kind], SHOW_DAYS, date_field).values())
            rows.sort(key=lambda r: dt_digits(r.get(date_field)), reverse=True)
            # 첫 화면에서 내려받는 파일이라 무겁게 두면 그대로 전송량이 된다
            out[kind] = rows[:MAX_ROWS]
        p = os.path.join(OUT, f"{name}.json")
        with open(p, "w", encoding="utf-8") as f:
            json.dump(out, f, ensure_ascii=False, separators=(",", ":"))
        print(f"  → {name}.json  공사 {len(out['con']):,} / 용역 {len(out['serv']):,}"
              f"  ({os.path.getsize(p)/1024:.0f}KB)")

    def export_board(name, store, date_field, enp_map=None):
        """한 달치를 500건씩 나눠 담는다.

        하루에 1순위 570건·공고 600건이 나오므로 300건만 실으면 반나절치도 안 된다.
        그렇다고 한 달치(약 18,000건, 6MB)를 한 파일에 넣으면 휴대폰에서 무겁다.
        그래서 묶음으로 쪼갠다 — 첫 화면은 0번만 받고,
        검색하거나 지역을 고를 때 나머지를 뒤에서 받아온다.
        """
        out_dir = os.path.join(OUT, "board")
        os.makedirs(out_dir, exist_ok=True)
        # rankChunk 는 화면이 «몇 번째 순위 파일인지» 를 셀 때 씁니다.
        # 여기서만 정합니다 — 화면에 같은 숫자를 두 번 적지 않습니다.
        meta = {"built": built, "chunk": BOARD_CHUNK,
                "rankChunk": BOARD_RANK_CHUNK}
        total = 0
        for kind in ("con", "serv"):   # serv 는 비어 있음 (형식 유지용)
            rows = list(trim(store[kind], SHOW_DAYS, date_field).values())
            rows.sort(key=lambda r: dt_digits(r.get(date_field)), reverse=True)
            # ★ 공고 묶음에도 예상 참가(enp)를 붙입니다 — 원클릭 카드가 «자리 찾기» 를 안 켜도 같은 금액을 내도록.
            #   bidindex·bidresult 와 같은 enp_map 이라 세 화면이 같은 분위로 같은 금액을 냅니다 (2026-09-03).
            if enp_map:
                for r in rows:
                    e = enp_map.get(str(r.get("inst") or "").strip())
                    if e:
                        r["enp"], r["enpn"] = e[0], e[1]
            # ══════════════════════════════════════════════════════
            #  ★ 순위 30곳(corps)은 «펼칠 때만» 받습니다 — 2026-09-06
            #
            #  실측: first-con-0.json 이 gzip 376KB 인데 그 중 **303KB(80%)** 가
            #  corps·rq·drw 였습니다. 1순위 탭을 열기만 해도 이게 다 옵니다.
            #  그런데 화면이 순위를 쓰는 곳은 **카드를 펼쳤을 때 한 곳뿐**이고,
            #  rq·drw 는 목록 화면 어디서도 안 읽습니다(통째로 뺍니다).
            #
            #      묶음 376KB → 73KB · 순위는 50건씩 따로 (한 개 약 5KB)
            #
            #  ⚠️ rows 는 저장소의 줄을 **그대로 가리킵니다.** 여기서 corps 를 지우면
            #     저장소에서도 지워집니다 — 힘들게 받은 30곳이 날아갑니다
            #     (2026-09-03 에 row_first 보존 목록에서 실제로 겪은 사고입니다).
            #     그래서 «지운 사본» 을 만들어 씁니다. rows 는 안 건드립니다.
            #  ⚠️ 순위 파일의 자리는 **전체에서 몇 번째 줄인가(pos)** 입니다.
            #     묶음 번호가 아닙니다. 화면은 useBoard 가 줄마다 붙여 주는 _rk 로 찾습니다.
            # ══════════════════════════════════════════════════════
            ranks = [r.get("corps") or 0 for r in rows]
            n_rank = 0
            if any(ranks):
                for k in range(0, len(ranks), BOARD_RANK_CHUNK):
                    with open(os.path.join(
                            out_dir,
                            f"{name}-{kind}-rank-{k // BOARD_RANK_CHUNK}.json"),
                            "w", encoding="utf-8") as f:
                        json.dump(ranks[k:k + BOARD_RANK_CHUNK], f,
                                  ensure_ascii=False, separators=(",", ":"))
                    n_rank += 1

            parts = [rows[i:i + BOARD_CHUNK]
                     for i in range(0, len(rows), BOARD_CHUNK)] or [[]]
            for i, part in enumerate(parts):
                slim = [{k: v for k, v in r.items() if k not in BOARD_RANK_KEYS}
                        for r in part]
                with open(os.path.join(out_dir, f"{name}-{kind}-{i}.json"),
                          "w", encoding="utf-8") as f:
                    json.dump(slim, f, ensure_ascii=False, separators=(",", ":"))
            # 지난번보다 묶음 수가 줄었을 때 남는 옛 파일을 정리한다.
            # (마운트된 폴더는 삭제가 막힐 수 있어, 안 되면 빈 파일로 덮어쓴다)
            i = len(parts)
            while True:
                stale = os.path.join(out_dir, f"{name}-{kind}-{i}.json")
                if not os.path.exists(stale):
                    break
                try:
                    os.remove(stale)
                except Exception:
                    with open(stale, "w", encoding="utf-8") as f:
                        f.write("[]")
                i += 1

            j = n_rank
            while True:
                stale = os.path.join(out_dir, f"{name}-{kind}-rank-{j}.json")
                if not os.path.exists(stale):
                    break
                try:
                    os.remove(stale)
                except Exception:
                    with open(stale, "w", encoding="utf-8") as f:
                        f.write("[]")
                j += 1

            # ══════════════════════════════════════════════════════════
            #  ★ 검색 색인 — 2026-09-03
            #
            #  검색하면 7주치를 다 뒤져야 하니, 전에는 묶음을 **전부** 받았습니다.
            #  실측 1순위 1,528KB · 공고 1,767KB. 한 번 검색이 안 하는 방문 16명분입니다.
            #
            #  그런데 사람이 실제로 보는 건 «첫 쪽 20건» 입니다. 실측:
            #      「도로」 597건 · 전체를 그리려면 23묶음 · **첫 20건은 1묶음**
            #      「경주시」 65건 · 19묶음      · **첫 20건은 5묶음**
            #  그래서 «걸러내기»에 필요한 최소한만 담은 색인을 따로 만듭니다.
            #  화면은 이 색인으로 몇 건인지·몇 쪽인지를 정확히 세고,
            #  **보고 있는 쪽에 필요한 묶음만** 받아옵니다.
            #
            #  색인 크기: 1순위 358KB · 공고 352KB (전부 받기의 1/4~1/5)
            #  ⚠️ 순서는 묶음을 이어붙인 것과 **정확히 같아야** 합니다.
            #     n번째 항목이 (n÷500)묶음의 (n%500)번째 줄입니다. 이게 어긋나면
            #     검색 결과가 엉뚱한 공고를 가리킵니다. 아래 IDX_F 를 바꿀 땐
            #     useBoard.js 의 읽는 순서도 같이 고치세요.
            # ══════════════════════════════════════════════════════════
            if kind == "con":
                # ★ 면허(lic)를 색인에 넣습니다 — 2026-09-05
                #   전에는 공고명 낱말(«배수»·«기초»…)로 면허를 «추측» 했습니다.
                #   실측: 철근·콘크리트를 고르면 580건이 나오는데 진짜는 91건(15.7%)이고,
                #   반대로 진짜 514건 중 423건(82%)을 놓쳤습니다.
                #   조달청이 lic 로 정확히 주고 있었는데 안 쓰고 있었던 것입니다
                #   (CLAUDE.md 1번 「조달청이 주는 값이 있으면 그대로 쓴다」).
                #   색인에는 «코드만» 넣습니다(4994 등) — 이름까지 넣으면 168KB, 코드만이면 72KB.
                # 지역(rgn)도 색인에 넣습니다 — 화면이 «기관명에 그 글자가 있나» 로
                # 짐작하면 「전남광주통합특별시 장흥군」 이 광주로 잡힙니다(실측 833건).
                rbook = region_book(rows)
                if name == "first":
                    # 1순위: 검색은 공고명·기관·낙찰업체
                    idx = [[r.get("name") or "", r.get("inst") or "",
                            r.get("win") or "", lic_codes(r), sido_of(r, rbook)]
                           for r in rows]
                    fields = ["name", "inst", "win", "lic", "sido"]
                else:
                    # 공고: 검색은 공고명·기관. base/lo/hi 는 「해볼 만한 공고만」 등급 계산용
                    # dsn — 붙임에 «설계내역서»(발주처 설계 단가가 든 것)가 있는가.
                    #       0 없음 · 1 내역서 있음 · 2 설계내역서/단가산출서 있음
                    #       공고 목록에서 「단가 든 내역서 있는 공고만」 을 거르는 데 씁니다.
                    idx = [[r.get("name") or "", r.get("inst") or "",
                            int(r.get("base") or 0),
                            r.get("lo"), r.get("hi"), lic_codes(r), sido_of(r, rbook),
                            doc_flag(r)]
                           for r in rows]
                    fields = ["name", "inst", "base", "lo", "hi", "lic", "sido", "dsn"]
                with open(os.path.join(out_dir, f"{name}-{kind}-idx.json"),
                          "w", encoding="utf-8") as f:
                    json.dump({"f": fields, "chunk": BOARD_CHUNK, "r": idx},
                              f, ensure_ascii=False, separators=(",", ":"))

            days = sorted({dt_digits(r.get(date_field))[:8] for r in rows if r.get(date_field)})
            meta[kind] = {
                "n": len(rows),
                "parts": len(parts),
                "ranks": n_rank,
                "from": days[0] if days else "",
                "to": days[-1] if days else "",
            }
            if kind == "con":
                # 면허 칩 목록을 «자료에서» 만듭니다. 손으로 적어 두면 조달청이 이름을
                # 바꿨을 때 조용히 안 맞습니다. [코드, 이름, 건수] · 건수 순 상위 40가지.
                cnt, nm = {}, {}
                for r in rows:
                    for code, label in lic_pairs(r):
                        cnt[code] = cnt.get(code, 0) + 1
                        nm[code] = label
                top = sorted(cnt.items(), key=lambda kv: -kv[1])[:40]
                meta[kind]["lics"] = [[c, nm[c], n] for c, n in top]
                meta[kind]["nolic"] = sum(1 for r in rows if not lic_codes(r))
                rc = {}
                for r in rows:
                    for g in (sido_of(r, rbook) or "").split(","):
                        if g:
                            rc[g] = rc.get(g, 0) + 1
                meta[kind]["rgns"] = rc
                meta[kind]["norgn"] = sum(1 for r in rows if not sido_of(r, rbook))
            total += len(rows)
        with open(os.path.join(out_dir, f"{name}.json"), "w", encoding="utf-8") as f:
            json.dump(meta, f, ensure_ascii=False, separators=(",", ":"))
        size = sum(os.path.getsize(os.path.join(out_dir, x))
                   for x in os.listdir(out_dir) if x.startswith(name + "-"))
        _ix = os.path.join(out_dir, f"{name}-con-idx.json")
        _ixs = os.path.getsize(_ix) / 1024 if os.path.exists(_ix) else 0
        print(f"  → board/{name}  {total:,}건 "
              f"(공사 {meta['con']['parts']}묶음 / 용역 {meta['serv']['parts']}묶음, "
              f"{size/1024/1024:.1f}MB · 검색색인 {_ixs:.0f}KB)")


    def fetch_naeyeok_files(store):
        """설계내역서를 실제로 «받아» 사이트에 올립니다. (2026-09-05)

        소장님: 「파일은 퍼 와도 돼, 사이트에서 사용자가 다운 받을수 있게 해줘.
                 그래야 홍보 문구를 넣지」

        ■ 왜 설계내역서만 받나 — 실측하고 정했습니다
          크롬으로 5개를 실제로 받아 재봤습니다: 129KB · 96KB · 52KB · **5.4MB** · 65KB.
          4,290개를 다 받으면 1GB 가 넘습니다(Firebase 무료 10GB 에 부담).
          단가가 든 «설계내역서»만 받으면 209개 · 약 30~60MB 입니다.
          나머지(공내역서·물량내역서)는 단가가 없어 받아 둘 값어치가 적습니다 — 링크로 둡니다.

        ■ 받아서 «열어 봅니다» — 이게 링크만 걸 때와 결정적으로 다른 점입니다
          이름이 「설계내역서」인데 실제로는 단가가 빈 것이 섞여 있습니다
          (「(무단가) 설계내역서」·「설계내역서(공내역)」 실측 18/227개).
          받아서 열면 «단가 열에 값이 있는지» 를 진짜로 확인할 수 있습니다.

        ■ ★ 받은 파일은 data/store 안에 둡니다 — 여기서 설계가 갈렸습니다
          web/public/data 는 .gitignore 되어 있고, GitHub Actions 가 회차 사이에
          넘겨주는 것은 **`data/store` 하나뿐**입니다(워크플로의 cache 경로).
          web/public/naeyeok 에 바로 받으면 다음 회차에는 파일이 사라지는데
          기록(book)만 «이미 받음» 으로 남아 **화면의 «바로 받기» 가 404** 가 됩니다.
          → 받는 곳은 data/store/naeyeok, 배포 전에 web/public/naeyeok 로 복사합니다.
          (워크플로는 보호 파일이라 못 고칩니다 — CLAUDE.md. 여기서 해결합니다.)

        ■ 무한히 쌓이지 않게
          공고는 날마다 새로 나옵니다. 상한(NAEYEOK_KEEP_MB)을 넘으면
          **오래된 것부터** 지우고 기록에서도 뺍니다. 지운 것은 다시 안 받습니다
          (기록에 skip 을 남깁니다 — 안 그러면 지우고 받기를 되풀이합니다).

        ■ 출처를 반드시 남깁니다
          파일마다 발주기관·공고번호·공고주소를 목록에 함께 싣습니다.
        """
        os.makedirs(NAEYEOK_DIR, exist_ok=True)
        if NO_NET:
            # --exportonly 처럼 «바깥을 아예 안 부른다» 고 정한 때만 건너뜁니다.
            # ⚠️ NET_DOWN(조달청 OpenAPI 차단기)으로는 건너뛰지 않습니다 —
            #    첨부파일은 www.g2b.go.kr 이라 다른 서버입니다. 여기서 막았더니
            #    수집 막바지에 차단기가 내려간 회차마다 0건이 되었습니다.
            DIAG["naeyeok_fetch"] = {"건너뜀": "--exportonly (바깥을 부르지 않는 회차)"}
            print("  → naeyeok 파일  받기 건너뜀 (--exportonly)")
            return load_json(NAEYEOK_BOOK, {})
        # ★ 2026-09-06 — 갈래를 가리지 않고 받습니다.
        #   소장님: 「직접 다운 받을 수 있게 해줘」
        #   전에는 «단가 든 갈래» 만 받아서, 4,294개 중 바로 받을 수 있는 것이 35개뿐이었습니다.
        #   나머지는 「나라장터에서 받기」로 가는데 거긴 로그인을 요구합니다 — 사용자에겐 «안 되는» 것입니다.
        #   ⚠️ 그래도 **순서**는 값어치 순입니다: 단가 든 갈래(설계내역서·단가산출서) 먼저,
        #      그다음이 나머지. 보관 상한(NAEYEOK_KEEP_MB)에 걸리면 뒤엣것부터 못 받으니,
        #      «먼저 받는 것» 이 곧 «남는 것» 입니다.
        want = []
        for r in store["con"].values():
            for i, y in enumerate(r.get("docs") or []):
                fname = str(y[0] if isinstance(y, (list, tuple)) else y)
                furl = str(y[1]) if isinstance(y, (list, tuple)) and len(y) > 1 else ""
                k = naeyeok_kind(fname)
                if not furl or not k:
                    continue
                want.append((0 if k in NAEYEOK_PRICED else 1,
                             str(r.get("dt") or "")[:10], r, fname, furl, i))
        # 단가 든 갈래 먼저, 그 안에서 최신부터
        want.sort(key=lambda x: (x[0], [-ord(c) for c in x[1]]))
        # ⚠️ 그다음(단가 없는 갈래)은 **갈래를 돌아가며** 집습니다 (2026-09-06).
        #    한 줄로 늘어놓으면 공내역서 2,000개가 앞을 다 막아, 27개뿐인 수량산출서는
        #    영영 «바로 받기» 가 0개입니다. 실제로 그랬습니다 — 소장님이 바로 알아채셨습니다.
        #    갈래마다 돌아가며 집으면 어느 갈래든 금방 «받아지는 예» 가 생깁니다.
        head = [x for x in want if x[0] == 0]
        tail = [x for x in want if x[0] != 0]
        buckets = {}
        for x in tail:
            buckets.setdefault(naeyeok_kind(x[3]), []).append(x)
        mixed = []
        while any(buckets.values()):
            for k in list(buckets):
                if buckets[k]:
                    mixed.append(buckets[k].pop(0))
        want = [(dt, r, fn, fu, i) for _p, dt, r, fn, fu, i in head + mixed]

        book = load_json(NAEYEOK_BOOK, {})                # 이미 받은 것 기록
        got = new = 0
        errs, firsts = {}, []                             # 왜 못 받았는지 — diag 로 나갑니다
        t0, ran_out = time.time(), False
        # 나라장터가 아예 안 열리는 회차가 있습니다(2026-09-05 실측: ConnectTimeout 10건 251초).
        # 그럴 때 40건을 끝까지 두드리면 4분을 버리고도 0건입니다.
        # → 처음부터 연달아 실패하고 한 건도 못 받았으면 일찍 접습니다. 다음 회차에 다시 합니다.
        gave_up = False
        nonlocal_streak = [0]                              # _fail 이 안쪽 함수라 리스트로 셉니다

        def _fail(key, b0, why, url="", head=""):
            """실패를 «다시 해 볼 수 있게» 적습니다.

            ⚠️ 전에는 실패를 skip 으로 못박아 다시는 안 물어봤습니다.
               조달청이 한 시간 먹통이면 그 40건은 영영 못 받습니다
               (CLAUDE.md 4번 — 한 번의 실패로 «안 된다» 고 단정하지 않는다).
            """
            errs[why] = errs.get(why, 0) + 1
            nonlocal_streak[0] += 1
            if len(firsts) < 3:
                firsts.append({"why": why, "url": url[:120], "head": head[:200]})
            book[key] = {"why": why, "try": int(b0.get("try") or 0) + 1, "at": built}
            return book[key]

        for dt, r, fname, furl, i in want:
            key = "%s_%d" % (str(r.get("no") or ""), i)
            ext = os.path.splitext(fname)[1].lower() or ".xlsx"
            if ext not in (".xlsx", ".xls", ".xlsm"):
                continue                                   # zip·pdf 는 열어 볼 수 없어 건너뜁니다
            local = key + ext
            path = os.path.join(NAEYEOK_DIR, local)
            b0 = book.get(key) or {}
            if b0.get("file") and os.path.exists(path):
                got += 1
                continue
            if b0.get("perm") or int(b0.get("try") or 0) >= 5:
                continue                                   # 다섯 번 해 보고 그만둡니다
            if new >= NAEYEOK_FETCH:
                break
            if time.time() - t0 > NAEYEOK_BUDGET_S:
                ran_out = True                             # 시간을 다 썼습니다 — 다음 회차에 이어서
                break
            if got == 0 and nonlocal_streak[0] >= 5:
                gave_up = True                             # 나라장터가 안 열리는 회차
                break
            new += 1
            try:
                resp = requests.get(furl, timeout=NAEYEOK_TIMEOUT_S, verify=False, headers={
                    "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                                   "AppleWebKit/537.36 (KHTML, like Gecko) "
                                   "Chrome/126.0 Safari/537.36"),
                    "Referer": "https://www.g2b.go.kr/",
                    "Accept": "*/*",
                })
            except Exception as e:
                _fail(key, b0, "연결 실패: %s" % type(e).__name__, furl)
                continue
            body = resp.content
            if resp.status_code != 200:
                _fail(key, b0, "HTTP %d" % resp.status_code, furl,
                      body[:200].decode("utf-8", "replace"))
                continue
            if len(body) > NAEYEOK_MAXBYTES:
                _fail(key, b0, "너무 큼 %.1fMB" % (len(body) / 1024 / 1024), furl)["perm"] = True
                continue
            # ⚠️ 200 이라고 엑셀이 아닙니다 — 로그인 안내 HTML 이 200 으로 올 수 있습니다.
            #    확장자만 믿고 저장하면 «열리지 않는 xlsx» 를 사용자에게 내밀게 됩니다.
            if not (body[:4] == b"PK\x03\x04" or body[:4] == b"\xd0\xcf\x11\xe0"):
                _fail(key, b0, "엑셀이 아님(%d바이트)" % len(body), furl,
                      body[:200].decode("utf-8", "replace"))
                continue
            with open(path, "wb") as f:
                f.write(body)
            book[key] = {"file": local, "n": len(body), "dt": dt,
                         "priced": xlsx_has_price(path), "at": built}
            got += 1
            nonlocal_streak[0] = 0
            time.sleep(0.15)

        # ── 보관 기간이 지난 파일은 버립니다 (목록과 같은 상한) ──────────
        old_cut = (datetime.now(KST) - timedelta(days=NAEYEOK_KEEP_DAYS)).strftime("%Y-%m-%d")
        aged = 0
        for k, v in list(book.items()):
            if v.get("file") and (v.get("dt") or "") and v["dt"] < old_cut:
                try:
                    p_ = os.path.join(NAEYEOK_DIR, v["file"])
                    if os.path.exists(p_):
                        os.remove(p_)
                except Exception:
                    continue
                book[k] = {"skip": "%d일 지남" % NAEYEOK_KEEP_DAYS, "perm": True}
                aged += 1

        # ── 크기 상한을 넘으면 오래된 것부터 버립니다 ────────────────
        #   ⚠️ 3년치를 다 두면 실측 추정 810MB 입니다. 배포 1벌에 얹혀 ×10 보관이면
        #      Firebase 무료 10GB 에 닿습니다. 그래서 날짜 상한과 «별도로» 크기 상한을 둡니다.
        keep = [(v.get("dt") or "", k, v) for k, v in book.items() if v.get("file")]
        keep.sort(reverse=True)                            # 최신이 앞
        budget, used, dropped = NAEYEOK_KEEP_MB * 1024 * 1024, 0, 0
        for _dt, k, v in keep:
            p = os.path.join(NAEYEOK_DIR, v["file"])
            n = os.path.getsize(p) if os.path.exists(p) else 0
            if used + n <= budget:
                used += n
                continue
            try:
                if os.path.exists(p):
                    os.remove(p)
            except Exception:
                continue
            book[k] = {"skip": "보관 상한(%dMB) 초과로 내림" % NAEYEOK_KEEP_MB}
            dropped += 1
        save_json(NAEYEOK_BOOK, book)

        have = sum(1 for v in book.values() if v.get("file"))
        pr = sum(1 for v in book.values() if v.get("priced"))
        # 「0건」 을 이유 없이 남기지 않습니다 — data/diag.json 을 열면 까닭이 있습니다.
        DIAG["naeyeok_fetch"] = {
            "대상(단가 든 갈래)": len(want), "보관": have, "단가 확인됨": pr,
            "이번에 두드린 것": new, "실패 이유": errs, "첫 실패 3건": firsts,
            "걸린 초": round(time.time() - t0, 1),
            "시간 상한에 걸림": ran_out,
            "일찍 접음(연달아 5번 실패)": gave_up,
            "상한MB": NAEYEOK_KEEP_MB, "쓴MB": round(used / 1024 / 1024, 1),
            "보관일수": NAEYEOK_KEEP_DAYS, "3년 지나 내린 것": aged,
        }
        if aged:
            print("    보관 기간(%d일)이 지난 파일 %d개를 내렸습니다"
                  % (NAEYEOK_KEEP_DAYS, aged))
        if errs:
            print("    실패 이유: " + " · ".join("%s %d건" % (k, v) for k, v in errs.items()))
            for x in firsts:
                print("      %s ← %s%s" % (x["why"], x["url"],
                                           ("  |  " + x["head"].replace("\n", " ")) if x["head"] else ""))
        print("  → naeyeok 파일  보관 %s개 (이번 %d개 두드림 · "
              "단가 확인됨 %s개 · %.0fMB / 상한 %dMB"
              "%s)" % (f"{have:,}", new, f"{pr:,}", used / 1024 / 1024, NAEYEOK_KEEP_MB,
                       (" · 오래된 것 %d개 내림" % dropped) if dropped else ""))
        return book

    def publish_naeyeok_files():
        """data/store/naeyeok → web/public/naeyeok 로 옮겨 담습니다. (2026-09-05)

        받는 곳과 내보내는 곳을 나눈 이유는 fetch_naeyeok_files 의 설명대로입니다.
        빌드(npm run build)가 web/public 을 dist 로 복사하므로 여기 있으면 배포됩니다.
        반환값은 «실제로 사이트에 올라간 파일 이름» 의 집합입니다 —
        export_naeyeok 이 이 집합에 있는 것만 «바로 받기» 로 내보냅니다.
        (없는 파일을 «바로 받기» 로 내면 사용자는 404 를 봅니다 —
         CLAUDE.md 「없는 자료를 없는 채로 그리지 않는다」)
        """
        pub = os.path.join(ROOT, "web", "public", "naeyeok")
        os.makedirs(pub, exist_ok=True)
        live = set()
        if os.path.isdir(NAEYEOK_DIR):
            for fn_ in os.listdir(NAEYEOK_DIR):
                src = os.path.join(NAEYEOK_DIR, fn_)
                dst = os.path.join(pub, fn_)
                try:
                    if not os.path.exists(dst) or os.path.getsize(dst) != os.path.getsize(src):
                        shutil.copyfile(src, dst)
                    live.add(fn_)
                except Exception:
                    pass
        # 저장소에서 내려간 것은 사이트에서도 내립니다
        for fn_ in os.listdir(pub):
            if fn_ not in live:
                try:
                    os.remove(os.path.join(pub, fn_))
                except Exception:
                    pass
        return live

    def export_naeyeok(store, live=None):
        """내역서 모음 — 조달청 붙임 파일을 갈래별로 모읍니다. (2026-09-06 3년 보관)

        ■ 저장은 3년, 화면은 상한 — 여기가 이 함수의 핵심입니다
          소장님: 「공사내역서 누적하지 말고. 딱 3년 치만 저장」
          store(first/live)는 70일치라 목록도 70일치였습니다. 그래서 목록만 따로
          data/store/naeyeok_index.json 에 쌓고, 3년이 지난 줄은 버립니다.
          ⚠️ 그런데 3년치를 화면에 통째로 내면 **gzip 3.4MB** 입니다(실측).
             목록 한 번 여는 데 방문자 10명분 전송량입니다.
             → 화면에는 갈래별 최신 NAEYEOK_SHOW 개까지만 냅니다.
             화면에도 「저장은 3년치 · 목록은 최신 N개」 라고 적습니다.

        ■ 두 파일로 나눕니다
          naeyeok.json      단가가 «들어 있는» 갈래(설계내역서·단가산출서) — 첫 화면에서 받습니다
          naeyeok-all.json  나머지(공내역서·물량내역서 등) — 그 갈래를 눌렀을 때만 받습니다
        """
        book = load_json(NAEYEOK_BOOK, {})     # 실제로 받아 «열어 본» 결과
        # live = 실제로 web/public/naeyeok 에 올라간 파일 이름들.
        # 기록에만 있고 파일이 없으면 «바로 받기» 를 내지 않습니다 (404 방지).
        if live is None:
            live = publish_naeyeok_files()

        # ── 1) 이번 회차에 보이는 것을 누적 색인에 합칩니다 ──────────
        idx = load_json(NAEYEOK_INDEX, {})
        before = len(idx)
        for r in store["con"].values():
            docs = r.get("docs") or []
            if not docs:
                continue
            nm = str(r.get("name") or "")
            inst = str(r.get("inst") or "")
            dt = str(r.get("dt") or "")[:10]
            no = str(r.get("no") or "")
            purl = str(r.get("url") or "")      # 조달청이 준 공고 주소 (손으로 만들지 않습니다)
            for i, y in enumerate(docs):
                fname = str(y[0] if isinstance(y, (list, tuple)) else y)
                furl = str(y[1]) if isinstance(y, (list, tuple)) and len(y) > 1 else ""
                kind = naeyeok_kind(fname)
                if not kind or not furl:
                    continue
                # 갈래는 «지금 규칙» 으로 다시 매깁니다 — 규칙을 고치면 옛 줄도 따라옵니다.
                idx["%s_%d" % (no, i)] = [kind, fname, furl, nm, inst, dt, no, purl]

        # ── 2) 3년이 지난 줄은 버립니다 (누적하지 않습니다) ──────────
        cut = (datetime.now(KST) - timedelta(days=NAEYEOK_KEEP_DAYS)).strftime("%Y-%m-%d")
        dropped = 0
        for k in [k for k, v in idx.items() if (v[5] or "") and v[5] < cut]:
            del idx[k]
            dropped += 1
        save_json(NAEYEOK_INDEX, idx)

        # ── 3) 화면에 낼 것만 골라 냅니다 ────────────────────────────
        rows_p, rows_a = [], []
        seen = set()
        # 최신 공고가 위로 (dt 내림차순)
        for key, v in sorted(idx.items(), key=lambda kv: kv[1][5] or "", reverse=True):
            kind, fname, furl, nm, inst, dt, no, purl = v
            if (fname, furl) in seen:
                continue
            seen.add((fname, furl))
            # 우리가 받아 둔 파일이 있으면 그 주소와 «열어 본 결과» 를 함께 싣습니다.
            #   priced  1 단가 확인됨 · 0 열어 보니 단가 없음 · -1 아직 안 열어 봄
            b = book.get(key) or {}
            # ⚠️ 주소 뒤에 «크기 도장» 을 붙입니다 (2026-09-06).
            #    firebase.json 이 /naeyeok/** 에 7일 immutable 을 겁니다. 그런데 그 경로는
            #    **없는 파일에도 200 + index.html** 을 돌려줍니다(catch-all rewrite).
            #    파일이 아직 없던 때 그 주소를 한 번이라도 두드린 브라우저는
            #    **HTML 을 xlsx 로 알고 7일 동안 붙잡습니다** — 눌러도 «손상된 파일» 이 됩니다.
            #    크기가 주소에 들어가면 그때의 주소와 지금의 주소가 달라, 낡은 것을 못 씁니다.
            local = ("/naeyeok/%s?v=%d" % (b["file"], int(b.get("n") or 0))
                     if (b.get("file") in live) else "")
            priced = 1 if b.get("priced") is True else (0 if b.get("priced") is False else -1)
            (rows_p if kind in NAEYEOK_PRICED else rows_a).append(
                [kind, fname, furl, nm, inst, dt, no, purl, local, priced])

        # 갈래마다 상한 — 전송량을 지키는 자리입니다
        kept, cap = {}, []
        for v, lim in ((rows_p, NAEYEOK_SHOW_TOP), (rows_a, NAEYEOK_SHOW)):
            out = []
            for row in v:                       # 이미 최신 순입니다
                k = row[0]
                if kept.get(k, 0) >= lim:
                    continue
                kept[k] = kept.get(k, 0) + 1
                out.append(row)
            cap.append(out)
        rows_p, rows_a = cap

        # 「전부」와 「보여주는 것」 을 갈래마다 함께 냅니다 — 화면이 정직하게 적을 수 있도록
        allcnt = {}
        for v in idx.values():
            allcnt[v[0]] = allcnt.get(v[0], 0) + 1
        cnt = {}
        for x in rows_p + rows_a:
            cnt[x[0]] = cnt.get(x[0], 0) + 1

        fields = ["kind", "file", "url", "name", "inst", "dt", "no", "purl", "local", "priced"]
        meta = {"built": built, "f": fields, "n": len(rows_p) + len(rows_a),
                "kinds": cnt, "all": allcnt,
                "show": NAEYEOK_SHOW, "showTop": NAEYEOK_SHOW_TOP,
                "days": NAEYEOK_KEEP_DAYS, "total": len(idx)}
        for fn_, rows in (("naeyeok.json", rows_p), ("naeyeok-all.json", rows_a)):
            path = os.path.join(OUT, fn_)
            with open(path, "w", encoding="utf-8") as f:
                json.dump(dict(meta, r=rows), f, ensure_ascii=False, separators=(",", ":"))
        a = os.path.getsize(os.path.join(OUT, "naeyeok.json")) / 1024
        b2 = os.path.getsize(os.path.join(OUT, "naeyeok-all.json")) / 1024
        print("  → naeyeok  쌓아 둔 것 %s개(%d일치, 이번에 +%s%s) · "
              "화면에 내는 것 %s개 (단가 있는 것 %s개 %.0fKB / 나머지 %s개 %.0fKB)"
              % (f"{len(idx):,}", NAEYEOK_KEEP_DAYS, f"{len(idx) - before + dropped:,}",
                 (" · 기간 지나 버림 %d" % dropped) if dropped else "",
                 f"{len(rows_p) + len(rows_a):,}", f"{len(rows_p):,}", a, f"{len(rows_a):,}", b2))

    def export_bidindex(store, fstore):
        """«바로투찰» 전용 — 아직 마감되지 않은 공고만 담은 가벼운 목록.

        투찰가를 계산하는 사람은 «앞으로 넣을 공고» 만 찾습니다.
        그런데 board/live-* 를 전부 받으면 4MB 가까이 됩니다.
        그래서 마감 전 공고만, 계산에 꼭 필요한 값만 골라
        배열 형태로 담습니다. (이름표를 빼면 크기가 절반쯤 됩니다)
        """
        # GitHub 서버는 세계표준시로 돕니다. 마감시각은 한국시간이라 KST 로 비교해야 합니다.
        now = datetime.now(KST).strftime("%Y%m%d%H%M%S")
        enp_map, pick = pick_stats(fstore)
        _rbook = region_book(list(store["con"].values()))
        rows = []
        for r in store["con"].values():
            c = re.sub(r"[^0-9]", "", str(r.get("close") or ""))
            if not c:
                continue
            if c.ljust(14, "0") < now:      # 이미 마감된 공고는 뺀다
                continue
            rows.append([
                r.get("no") or "",
                r.get("name") or "",
                r.get("inst") or "",
                int(r.get("base") or 0),
                int(r.get("budget") or 0),
                r.get("close") or "",
                r.get("lo") if r.get("lo") is not None else -3,
                r.get("hi") if r.get("hi") is not None else 3,
                r.get("llr"),                      # 공고가 알려준 낙찰하한율
                int(r.get("est") or 0),            # 공고가 알려준 추정가격
                r.get("lic") or [],                # 면허·업종 제한
                int(r.get("aval") or 0),           # A값 합계 (법정경비)
                int(r.get("gmtrl") or 0),          # 관급자재금액
                r.get("ayn") or "",                # A값 적용 공고인지 (Y/N)
                # ⚠️ aparts(A값 내역)는 여기 없습니다 — aparts.json 으로 뺐습니다.
                #    이유는 아래 export_aparts 주석 참고. 첫 화면 전송량 −29%.
                r.get("ptot") or 0,                # 예비가격 개수
                r.get("pdrw") or 0,                # 추첨 개수
                # 나라장터 공고 주소 — 조달청이 준 것을 그대로 씁니다.
                # 손으로 만들면 차수(000/001/002)를 틀립니다. 실제로 틀렸습니다.
                r.get("url") or "",
                r.get("site") or "",               # 공사 현장 지역 (500/500 채워짐)
                r.get("rgnb") or "",               # 지역제한 판단기준 — 있으면 지역제한 공고
                r.get("joint") or "",              # 공동수급 방식
                r.get("mthd") or "",               # 계약방법 (제한경쟁/일반경쟁…)
                r.get("swin") or "",               # 낙찰방법 상세 (적격심사 기준까지 들어옵니다)
                r.get("rebid") or "",              # 재입찰 여부
                # ★ 공고 고르기 (2026-09-03) — 이 기관 개찰의 참가업체수 중앙과 그 근거 건수. 없으면 0.
                (enp_map.get(str(r.get("inst") or "").strip()) or [0, 0])[0],
                (enp_map.get(str(r.get("inst") or "").strip()) or [0, 0])[1],
                r.get("dt") or "",                 # 공고일 (목록 카드가 보여줍니다)
                # ⚠️ 이름은 반드시 «sido» 입니다. 조달청 rgn(참가가능지역)이 이미 있어서
                #    rgn 으로 두면 공고 카드가 「참가지역: 전남」 이라고 엉뚱하게 적습니다.
                sido_of(r, _rbook),                # 시도 (지역 거르기 — 짐작하지 않습니다)
                doc_flag(r),                       # 붙임 내역서: 2 단가 있음 · 1 있음 · 0 없음
            ])
        rows.sort(key=lambda x: re.sub(r"[^0-9]", "", str(x[5])))
        out = {"built": built,
               "f": ["no", "name", "inst", "base", "budget", "close", "lo", "hi",
                     "llr", "est", "lic", "aval", "gmtrl",
                     "ayn", "ptot", "pdrw", "url",
                     "site", "rgnb", "joint", "mthd", "swin", "rebid",
                     "enp", "enpn", "dt", "sido", "dsn"],
               "pick": pick,
               "r": rows}
        path = os.path.join(OUT, "bidindex.json")
        with open(path, "w", encoding="utf-8") as f:
            json.dump(out, f, ensure_ascii=False, separators=(",", ":"))
        have = sum(1 for x in rows if x[3] > 0)
        print(f"  \u2192 bidindex  \ub9c8\uac10\uc804 {len(rows):,}\uac74 "
              f"(\uae30\ucd08\uae08\uc561 \uc788\ub294 \uac83 {have:,}\uac74, "
              f"{os.path.getsize(path)/1024:.0f}KB)")

    def export_aparts(store):
        """A값 내역(법정경비 항목별)만 따로 담습니다 — 2026-09-03.

        ══════════════════════════════════════════════════════════════
        왜 뺐나 — 실측하고 나서 답이 바뀐 사례입니다.

        bidindex.json 927KB 중 이 항목이 244KB 였습니다. 두 가지를 만들어 재봤습니다:

            ① 이름표(「산업안전보건관리비」…)를 번호로 + 범례
               → raw 927 → 766KB 인데 **gzip 153.7 → 148.1KB. 겨우 5.6KB.**
                 gzip 이 이미 반복 문자열을 지우고 있었습니다. 헛수고였습니다.
            ② 이 파일로 분리
               → **gzip 153.7 → 109.2KB (−29%)**. 이 파일은 37KB.

        ②가 이긴 이유는 «작아서»가 아니라 **첫 화면에서 아예 안 받아서** 입니다.
        A값 내역은 공고 하나를 «고른 뒤»에만 보여주는 것이라, 목록에 실을 이유가 없습니다.

        ⚠️ 교훈: 전송량을 논할 때 raw 크기를 보면 틀립니다. 반드시 gzip 후로 재세요.
        ══════════════════════════════════════════════════════════════
        """
        now = datetime.now(KST).strftime("%Y%m%d%H%M%S")
        ap = {}
        for r in store["con"].values():
            c = re.sub(r"[^0-9]", "", str(r.get("close") or ""))
            if not c or c.ljust(14, "0") < now:
                continue
            v = r.get("aparts")
            if v:
                ap[r.get("no") or ""] = v
        out = {"built": built, "a": ap}
        path = os.path.join(OUT, "aparts.json")
        with open(path, "w", encoding="utf-8") as f:
            json.dump(out, f, ensure_ascii=False, separators=(",", ":"))
        print(f"  \u2192 aparts    A값 내역 {len(ap):,}건 "
              f"({os.path.getsize(path)/1024:.0f}KB · 바로투찰 열 때만 받음)")

    def export_bidresult(fstore):
        """«내 계산이 맞았나»를 확인하기 위한 최근 개찰 결과 색인.

        바로투찰 첫 화면에 얹으면 전송량이 두 배가 되므로 별도 파일로 두고
        필요할 때만 받아갑니다. 공고번호 하나로 바로 찾을 수 있게 «지도» 모양입니다.
        """
        cut = (datetime.now(KST) - timedelta(days=7)).strftime("%Y%m%d%H%M")
        enp_map, _ = pick_stats(fstore)      # 기관별 예상 참가 — 채점이 «그날 자동 분위였나» 를 알기 위해
        out = {}
        for r in (fstore.get("con") or {}).values():
            if (dt_digits(r.get("dt")) or "0") < cut:
                continue
            no = r.get("no")
            if not no:
                continue
            # 전화·주소·대표자는 «있는 공고만» 옵니다(조달청 낙찰자 상세에 실릴 때만).
            # 없으면 빈 문자열로 두고, 화면에서는 아예 감춥니다.
            out[no] = [r.get("win") or "", int(r.get("amt") or 0),
                       r.get("rate"), int(r.get("np") or 0),
                       int(r.get("base") or 0), r.get("dt") or "",
                       r.get("tel") or "", r.get("ceo") or "",
                       r.get("bno") or "", r.get("adr") or "",
                       1 if r.get("tsrc") else 0,
                       # 어떤 공고를 채점하는지 화면에 밝혀야 합니다.
                       # 이게 없어서 «무슨 공고인지 모르겠다» 는 화면이 나왔습니다.
                       str(r.get("name") or "")[:60], str(r.get("inst") or "")[:30],
                       # 채점의 실격 판정에 필요합니다. A값을 모르면 하한을 낮게 잡아
                       # «가져갔을 자리»가 남발됩니다.
                       int(r.get("aval") or 0), r.get("ayn") or "",
                       # ★ 투찰금액 목록(낮은 순, 최대 12개) — «우리 금액이면 몇 위였나» 를
                       #   채점 화면에서 바로 셀 수 있게 합니다. 이름은 넣지 않습니다(용량).
                       [int(c[1]) for c in (r.get("corps") or [])[:12] if c and c[1]],
                       # 순위 사다리 [[등수, 금액], ...] — 우리 금액이 몇 등쯤인지 좁힙니다
                       r.get("rq") or [],
                       int(r.get("nrank") or 0),
                       # ⚠️ 예가범위 — 채점이 그 공고의 사정률 분포를 재현하려면 꼭 필요합니다.
                       #    없어서 ±2% 공고를 ±3% 로 채점했고, 5억 공고 기준 115만원이
                       #    어긋났습니다(실측 106건 중앙값).
                       r.get("lo"), r.get("hi"),
                       # ★ 면허·업종 제한 — 2026-09-03 추가.
                       #   소장님: 「투찰 업체가 1곳이라는 게 말이 돼?」
                       #   말이 됩니다. 그 공고는 면허를 「산림조합(지역조합)」으로 묶어
                       #   자격 되는 곳이 사실상 하나였습니다. 같은 «가곡지구» 사업의
                       #   일반 토목 공고는 같은 홍성군에서 323곳·377곳이 들어왔습니다.
                       #   **자료는 맞는데 화면이 «왜»를 안 알려줘서 의심을 샀습니다.**
                       #   조달청이 lcnsLmtNm 으로 주는 값이라 만들 필요도 없습니다.
                       (r.get("lic") or [])[:3],
                       # ★ 낙찰하한율·추정가격 — 2026-09-03 추가. 공고서에 적힌 하한율이 있으면
                       #   채점도 그걸로 계산해야 바로투찰과 같은 금액이 나옵니다.
                       r.get("llr"), int(r.get("est") or 0),
                       # ★ 예상 참가 — 그날 바로투찰이 «자동 분위» 였는지 채점이 알아야 같은 금액이 나옵니다 (2026-09-03)
                       (enp_map.get(str(r.get("inst") or "").strip()) or [0, 0])[0],
                       (enp_map.get(str(r.get("inst") or "").strip()) or [0, 0])[1]]
        path = os.path.join(OUT, "bidresult.json")
        with open(path, "w", encoding="utf-8") as f:
            json.dump({"built": built, "f": ["win", "amt", "rate", "np", "base", "dt",
                             "tel", "ceo", "bno", "adr", "tsrc", "name", "inst",
                             "aval", "ayn", "amts", "rq", "nrank", "lo", "hi", "lic",
                             "llr", "est", "enp", "enpn"],
                       "r": out}, f, ensure_ascii=False, separators=(",", ":"))
        print(f"  → bidresult 최근 7일 개찰 {len(out):,}건 "
              f"({os.path.getsize(path)/1024:.0f}KB)")

    def export_bandstat(fstore, lstore):
        """규모(추정가격)별 «경쟁 강도»와 «A값 비율».

        - 참가업체수(np): 개찰결과에 조달청이 실어 줍니다. 몇 개사와 붙는지 알면
          투찰률을 얼마나 공격적으로 잡을지 감이 잡힙니다.
        - A값 비율: A값이 아직 공개되지 않은 공고에서 «대략 이만큼»을 보여주기 위한 값입니다.
          추정치라고 화면에 분명히 적습니다.
        """
        BND = [("s", 0, 10e8), ("m", 10e8, 50e8), ("l", 50e8, 100e8), ("xl", 100e8, None)]

        def band_of(est):
            for k, lo, hi in BND:
                if est >= lo and (hi is None or est < hi):
                    return k
            return None

        def med(v):
            if not v:
                return None
            v = sorted(v)
            n = len(v)
            return v[n // 2] if n % 2 else (v[n // 2 - 1] + v[n // 2]) / 2

        cut = (datetime.now(KST) - timedelta(days=60)).strftime("%Y%m%d%H%M")
        nps, ars = {k: [] for k, _, _ in BND}, {k: [] for k, _, _ in BND}

        for r in (fstore.get("con") or {}).values():
            if (dt_digits(r.get("dt")) or "0") < cut:
                continue
            base = int(r.get("base") or 0)
            np_ = int(r.get("np") or 0)
            if base <= 0 or np_ <= 0:
                continue
            b = band_of(base / 1.1)
            if b:
                nps[b].append(np_)

        for r in (lstore.get("con") or {}).values():
            base = int(r.get("base") or 0)
            av = int(r.get("aval") or 0)
            if base <= 0 or av <= 0 or r.get("ayn") == "N":
                continue
            b = band_of(int(r.get("est") or 0) or base / 1.1)
            if b:
                ars[b].append(av / base)

        out = {"built": built, "bands": {}}
        for k, _, _ in BND:
            v = sorted(nps[k])
            out["bands"][k] = {
                "n": len(v),
                "npMed": med(v),
                "npLo": v[int(len(v) * 0.25)] if v else None,
                "npHi": v[int(len(v) * 0.75)] if v else None,
                "arN": len(ars[k]),
                "ar": round(med(ars[k]), 5) if ars[k] else None,
            }
        path = os.path.join(OUT, "bandstat.json")
        with open(path, "w", encoding="utf-8") as f:
            json.dump(out, f, ensure_ascii=False, separators=(",", ":"))
        line = " / ".join(
            f"{k}:{out['bands'][k]['n']}건·중앙{out['bands'][k]['npMed']}개사"
            for k, _, _ in BND)
        print(f"  → bandstat  {line}")

    def fill_contacts(fstore):
        """주소·전화 메우기.

        조달청은 낙찰자 상세를 «주는 공고»에만 실어 줍니다 — 실측 31%.
        그런데 같은 업체(사업자번호)가 다른 공고에서는 주소·전화와 함께 나옵니다.
        그래서 사업자번호로 이어 붙입니다. 가장 최근에 확인된 값을 씁니다.
        빌려온 값은 tsrc 로 표시해, 화면에서 «다른 공고에서 확인» 이라고 밝힙니다.
        (없는 걸 지어내지 않습니다 — 조달청이 준 값을 옮겨 담을 뿐입니다)"""
        book = {}
        for r in (fstore.get("con") or {}).values():
            bno = (r.get("bno") or "").strip()
            if not bno or not (r.get("tel") or r.get("adr")):
                continue
            d = dt_digits(r.get("dt")) or ""
            cur = book.get(bno)
            if cur is None or d > cur[0]:
                book[bno] = (d, r.get("tel") or "", r.get("adr") or "")
        n = 0
        for r in (fstore.get("con") or {}).values():
            if r.get("tel") or r.get("adr"):
                continue
            got = book.get((r.get("bno") or "").strip())
            if not got:
                continue
            _, tel, adr = got
            if tel:
                r["tel"] = tel
            if adr:
                r["adr"] = adr
            r["tsrc"] = 1
            n += 1
        have = sum(1 for r in (fstore.get("con") or {}).values() if r.get("tel") or r.get("adr"))
        tot = len(fstore.get("con") or {})
        print(f"  → 연락처 {n:,}건을 다른 공고에서 이어붙임 "
              f"(전체 {have:,}/{tot:,} = {have/max(tot,1)*100:.0f}%)")

    try:
        fill_contacts(first)
    except Exception as e:
        print(f"  ! 연락처 잇기 실패 ({type(e).__name__}: {e}) — 넘어갑니다")


    print("-" * 52)
    export("first", first, "dt")
    export("live", live, "dt")
    export_board("first", first, "dt")
    export_board("live", live, "dt", enp_map=pick_stats(first)[0])
    export_bidindex(live, first)
    export_aparts(live)
    try:
        fetch_naeyeok_files(live)
    except Exception as e:
        print(f"  ! \ub0b4\uc5ed\uc11c \ud30c\uc77c \ubc1b\uae30 \uc2e4\ud328 ({type(e).__name__}) \u2014 \ub118\uc5b4\uac11\ub2c8\ub2e4")
    export_naeyeok(live)
    # 새로 붙인 통계라 혹시 터져도 배치 전체를 멈추지 않게 감쌉니다.
    try:
        export_bidresult(first)
        export_bandstat(first, live)
    except Exception as e:
        print(f"  ! bandstat 실패 ({type(e).__name__}: {e}) — 넘어갑니다")
    save_diag()
    print("✅ 수집 완료")


if __name__ == "__main__":
    main()

# -*- coding: utf-8 -*-
"""
formsgen.py — 건설 서식(엑셀)을 굽습니다. (2026-09-04)

왜 파이썬으로 굽나
  화면에서 자바스크립트로 엑셀을 만들면 라이브러리를 브라우저로 내려보내야 합니다.
  서식은 «변하지 않는 것» 이라 미리 구워 두고 링크만 걸면 전송량이 0 입니다.

⚠️ 서식의 «내용»은 web/src/data/forms.json **한 곳**에만 적습니다.
   화면(Forms.jsx)·엑셀(여기)·미리굽기(prerender.py)가 모두 그 파일을 읽습니다.
   두 벌로 적으면 «화면과 엑셀이 다른 서식» 이 조용히 생깁니다.

만드는 곳: web/public/forms/{slug}.xlsx  (배포되면 /forms/{slug}.xlsx 로 열립니다)
"""

import io
import json
import math
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(ROOT, "web", "src", "data", "forms.json")
OUT = os.path.join(ROOT, "web", "public", "forms")
FONT_PATH = os.path.join(ROOT, "assets", "KcmKR-Bold.otf")

COLS = 24                      # 가로를 24칸으로 쪼개 놓고 병합해서 씁니다
COLW = 3.6                     # 24 × 3.6 ≈ A4 세로 한 장 폭
KFONT = "맑은 고딕"            # 윈도우 기본. 없으면 시스템이 알아서 대체합니다

THIN = Side(style="thin", color="BFC7D5")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEAD_FILL = PatternFill("solid", fgColor="EEF3FB")
LABEL_FILL = PatternFill("solid", fgColor="F6F8FC")


def _logo_png():
    """K-건설맵 로고를 그려 메모리에 담아 돌려줍니다 (파일로 안 남깁니다)."""
    from PIL import Image, ImageDraw, ImageFont
    W, H = 360, 84
    im = Image.new("RGBA", (W, H), (255, 255, 255, 0))
    d = ImageDraw.Draw(im)
    d.rounded_rectangle([0, 10, 64, 74], radius=14, fill=(15, 32, 63, 255))
    try:
        f1 = ImageFont.truetype(FONT_PATH, 38)
        f2 = ImageFont.truetype(FONT_PATH, 30)
        f3 = ImageFont.truetype(FONT_PATH, 17)
    except Exception:
        f1 = f2 = f3 = ImageFont.load_default()
    d.text((32, 42), "K", font=f1, fill=(255, 255, 255, 255), anchor="mm")
    d.text((78, 30), "K-건설맵", font=f2, fill=(15, 32, 63, 255), anchor="lm")
    d.text((80, 60), "k-conmap.com", font=f3, fill=(116, 142, 178, 255), anchor="lm")
    buf = io.BytesIO()
    im.save(buf, "PNG")
    buf.seek(0)
    return buf


def _spans(weights, total=COLS, least=2):
    """가중치(대략 100 합계)를 24칸으로 나눕니다. 마지막 칸이 남는 것을 흡수합니다."""
    s = sum(weights) or 1
    out = [max(least, round(total * w / s)) for w in weights]
    diff = total - sum(out)
    i = out.index(max(out))
    out[i] += diff
    if out[i] < least:                      # 그래도 모자라면 균등분할
        out = [total // len(weights)] * len(weights)
        out[-1] += total - sum(out)
    return out


class Sheet:
    def __init__(self, ws):
        self.ws = ws
        self.r = 1

    def merge(self, c0, span, text="", *, bold=False, fill=None, size=10,
              align="left", h=None, wrap=False, border=True):
        ws = self.ws
        c1 = c0 + span - 1
        ws.merge_cells(start_row=self.r, start_column=c0, end_row=self.r, end_column=c1)
        cell = ws.cell(row=self.r, column=c0, value=text)
        cell.font = Font(name=KFONT, size=size, bold=bold,
                         color="0F203F" if bold else "222222")
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        if fill:
            cell.fill = fill
        if border:
            for c in range(c0, c1 + 1):
                ws.cell(row=self.r, column=c).border = BOX
        if h:
            ws.row_dimensions[self.r].height = h
        return cell

    def blank(self, h=6):
        self.ws.row_dimensions[self.r].height = h
        self.r += 1


def build(form):
    wb = Workbook()
    ws = wb.active
    ws.title = form["title"][:28]
    for c in range(1, COLS + 1):
        ws.column_dimensions[get_column_letter(c)].width = COLW

    # ⚠️ 기준을 6 → 7 로 올렸습니다. 계약서의 «당사자» 표가 6칸이라
    #    계약서가 전부 가로로 누웠습니다(실제로 그렇게 나왔습니다).
    #    조문이 있는 문서는 무조건 세로입니다 — 가로로 누우면 읽기가 나빠집니다.
    has_clause = any(b["t"] == "cl" for b in form["sheet"]["blocks"])
    wide = (not has_clause) and any(
        b["t"] == "table" and len(b["cols"]) >= 7 for b in form["sheet"]["blocks"])
    ws.page_setup.orientation = "landscape" if wide else "portrait"
    ws.page_setup.paperSize = 9                     # A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = ws.page_margins.right = 0.3
    ws.page_margins.top = ws.page_margins.bottom = 0.4

    s = Sheet(ws)

    # ── 1행: 우리 표시 (지우기 쉽게 «한 줄»로) ─────────
    #  2026-09-05 소장님 결정: 「로고가 보이게 하되 지울 수 있게 해주면 돼.
    #   사용자가 다운받아 쓸 때 지울 수 있으면 그만이지. 그래도 건설맵은 보게 되잖아」
    #
    #  ⚠️ **그림(이미지)으로 넣으면 안 됩니다.** 엑셀에서 그림은 셀 위에 «떠» 있어서
    #     행을 지워도 그림만 남습니다. 그러면 지우는 일이 더 성가셔집니다.
    #     그래서 **글자**로, 그리고 **1행 하나에만** 넣습니다 —
    #     마우스 우클릭 → 「행 삭제」 한 번이면 흔적 없이 사라집니다.
    #  ⚠️ 지울 수 있다는 것을 «그 줄에» 적습니다. 모르면 못 지웁니다.
    s.r = 1
    s.merge(1, 12, "  K-건설맵   |   k-conmap.com   무료 건설 서식",
            bold=True, size=10.5, align="left", h=22, border=False)
    s.merge(13, COLS - 12, "← 이 1행을 지우고 쓰셔도 됩니다   ",
            size=8.5, align="right", border=False)
    s.r = 2
    s.blank(6)

    # ── 제목 ──────────────────────────────────────────
    s.merge(1, COLS, form["sheet"]["heading"], bold=True, size=19,
            align="center", h=38, border=False)
    s.r += 1
    s.blank(10)

    for b in form["sheet"]["blocks"]:
        t = b["t"]
        if t == "kv":
            for label, val in b["rows"]:
                s.merge(1, 6, "  " + label, bold=True, fill=LABEL_FILL, h=24)
                s.merge(7, COLS - 6, val, wrap=True)
                s.r += 1
            s.blank()
        elif t == "text":
            s.merge(1, COLS, b["text"], size=10, align="center", h=26,
                    wrap=True, border=False)
            s.r += 1
            s.blank()
        elif t == "table":
            spans = _spans(b.get("w") or [100 / len(b["cols"])] * len(b["cols"]))
            c = 1
            for name, sp in zip(b["cols"], spans):
                s.merge(c, sp, name, bold=True, fill=HEAD_FILL, align="center", h=24)
                c += sp
            s.r += 1
            for _ in range(b["n"]):
                c = 1
                for sp in spans:
                    s.merge(c, sp, "", h=22)
                    c += sp
                s.r += 1
            s.blank()
        elif t == "cl":
            # 조문 — 계약서용. 「제1조(목적) …」 을 한 줄씩 폅니다.
            # ⚠️ openpyxl 은 자동 행높이가 없습니다. 글자 수로 줄 수를 계산해 높이를 줍니다.
            for head, body in b["items"]:
                s.merge(1, COLS, head, bold=True, size=10.5, align="left",
                        h=20, border=False)
                s.r += 1
                if body:
                    # ⚠️ 열 너비 단위는 «숫자 한 글자» 기준입니다. 한글은 두 칸을 먹습니다.
                    #    처음에 1.35 를 곱했다가 줄 수를 3배로 적게 잡아 **글자가 잘렸습니다.**
                    #    (LibreOffice 로 PDF 를 뽑아 보고서야 보였습니다 — 눈으로 확인하는 이유)
                    per = max(20, int(COLS * COLW / 2.2))
                    lines = math.ceil(len(body) / per) + 1      # 한 줄 여유
                    s.merge(1, COLS, body, size=10, align="left",
                            h=14.5 * lines, wrap=True, border=False)
                    s.r += 1
                s.blank(4)
            s.blank()
        elif t == "sign":
            s.blank(10)
            s.merge(1, COLS, "년        월        일", align="center",
                    h=26, border=False)
            s.r += 1
            for who in b["who"]:
                s.merge(COLS - 13, 8, who, bold=True, align="right", border=False)
                s.merge(COLS - 5, 6, "(서명 또는 인)", align="center", border=False)
                s.r += 1
            if b.get("note"):
                s.merge(1, COLS, b["note"], size=9, align="right", border=False)
                s.r += 1
            s.blank()

    # ── 여기까지가 «인쇄되는 서식» 입니다 ─────────────
    #   print_area 를 여기서 끊습니다. 아래에 적는 것은 인쇄·PDF 에 나오지 않습니다.
    last = s.r - 1
    ws.print_area = f"A1:{get_column_letter(COLS)}{last}"

    # ── 인쇄 영역 «밖» — 파일을 열어 스크롤해야 보이는 안내 한 줄 ──
    #   우리 표시는 1행에 있습니다. 여기는 «주의사항»만 둡니다(인쇄 안 됨).
    s.r = last + 3
    s.merge(1, COLS,
            "발주기관이 정한 서식이 있으면 그 서식을 사용하세요. "
            "이 줄과 1행은 인쇄 전에 지우셔도 됩니다.",
            size=8.5, align="left", border=False)
    return wb


def main():
    with open(SRC, encoding="utf-8") as f:
        data = json.load(f)
    os.makedirs(OUT, exist_ok=True)
    made = []
    for form in data["forms"]:
        wb = build(form)
        p = os.path.join(OUT, f"{form['slug']}.xlsx")
        wb.save(p)
        made.append((form["slug"], os.path.getsize(p)))
    # 설계변경 탭이 쓰는 «작은 목록» — forms.json(217KB)을 통째로 보내지 않기 위해서입니다.
    # 사본이 아니라 여기서 굽는 산출물입니다. 손으로 고치지 마세요.
    mini = {f["slug"]: [f["title"], f.get("icon", ""), f.get("short", ""), f.get("group", "")]
            for f in data["forms"]}
    mp = os.path.join(ROOT, "web", "src", "data", "forms-min.json")
    with open(mp, "w", encoding="utf-8", newline="\r\n") as f:
        json.dump(mini, f, ensure_ascii=False, indent=0)
    print(f"  ✅ 작은 목록 {len(mini)}가지 → {mp} ({os.path.getsize(mp)/1024:.0f} KB)")

    total = sum(n for _, n in made)
    print(f"  ✅ 서식 {len(made)}개 · 합계 {total/1024:.0f} KB → {OUT}")
    for slug, n in made:
        print(f"     {slug:<18} {n/1024:6.1f} KB")


if __name__ == "__main__":
    main()

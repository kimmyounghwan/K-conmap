# -*- coding: utf-8 -*-
"""
chgbook.py — 설계변경 자동계산 통합 엑셀(11장)을 굽습니다. (2026-09-05)

소장님: 「수식이 다 연결돼 있어서 단가 하나 바꾸면 끝까지 다시 계산되고,
         당초·변경 내역을 넣으면 증감이 자동으로 나오는 그 파일」

■ 다른 서식과 무엇이 다른가
   formsgen.py 가 굽는 105가지는 «빈 표» 입니다 — 사람이 계산해서 채웁니다.
   이 파일은 **계산기**입니다. 단가마스터의 단가 한 칸을 고치면
   일위대가 → 내역 → 증감대비표 → 원가계산서 → 제출서식까지 다시 계산됩니다.

■ 시트 11장과 자료가 흐르는 방향
   설정 ─────────────┐ (낙찰률·요율은 여기 한 곳에만 적습니다)
   단가마스터 ─┬→ 일위대가 ─┐
               │             ├→ 당초내역 ─┐
   수량산출서 ─┴─────────────┤            ├→ 증감대비표 → 공종별집계
                             └→ 변경내역 ─┘        │
                                    └──────→ 원가계산서 → 제출서식
                                                   검증시트가 전부를 훑습니다

■ 계약금액 조정 단가 기준 (국가계약법 시행령 제65조 — 확인하고 넣었습니다)
   감소분        → 계약단가
   증가분        → 계약단가 (계약단가 > 예정가격단가 이면 예정가격단가)
   신규 비목     → 설계변경 «당시» 단가 × 낙찰률
   발주기관 요구 → 협의, 불성립 시 (당시단가 + 당시단가×낙찰률) ÷ 2
   ⚠️ «증가분에 낙찰률을 곱하는» 흔한 오해를 수식이 따라가지 않게 할 것.

■ 확인 방법 (눈으로 보지 않습니다)
   tools/checkbook.py 가 LibreOffice 로 실제 재계산시킨 뒤 값을 읽어
   파이썬으로 따로 계산한 정답과 대조합니다.

쓰는 법:  python chgbook.py
결과:     web/public/forms/chg-tonghap.xlsx
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(ROOT, "web", "public", "forms")
NAME = "chg-tonghap.xlsx"

# 몇 줄까지 수식을 깔아 둘지
N_MASTER = 300      # 단가마스터
N_ILWI = 300        # 일위대가 상세
N_ILWI_SUM = 60     # 일위대가 요약(일위 번호 개수)
N_SURYANG = 250     # 수량산출서
N_ITEM = 200        # 당초/변경 내역 · 증감대비표
N_GROUP = 25        # 공종별집계

HEAD = 4            # 1행 표시 · 2행 제목 · 3행 머리 · 4행부터 자료

BLUE = "1F4E79"
LITE = "DDEBF7"
GRAY = "F2F2F2"
YELLOW = "FFF2CC"
GREEN = "E2EFDA"
RED = "FCE4E4"

thin = Side(style="thin", color="BFBFBF")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)


def brand(ws, cols):
    """1행 — 지우고 쓰셔도 된다고 스스로 밝힙니다. 그림이 아니라 «글자» 입니다
       (엑셀에서 그림은 행을 지워도 남습니다)."""
    ws.cell(1, 1, "K-건설맵  |  k-conmap.com   ← 이 1행을 지우고 쓰셔도 됩니다")
    ws.cell(1, 1).font = Font(size=9, color="808080")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(cols, 2))


def title(ws, text, cols, sub=""):
    c = ws.cell(2, 1, text)
    c.font = Font(size=14, bold=True, color=BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(cols, 2))
    ws.row_dimensions[2].height = 24
    if sub:
        ws.cell(2, max(cols, 2))  # 자리만
    ws.freeze_panes = ws.cell(HEAD, 1)


def header(ws, cols, widths, row=3):
    for i, (name, w) in enumerate(zip(cols, widths), start=1):
        c = ws.cell(row, i, name)
        c.font = Font(size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BOX
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 30


def grid(ws, r0, r1, ncol, fills=None):
    """자료 칸에 테두리와 «입력칸/자동칸» 색을 깔아 둡니다.
       노랑 = 사람이 넣는 칸 · 회색 = 수식이 채우는 칸. 한눈에 갈립니다."""
    for r in range(r0, r1 + 1):
        for c in range(1, ncol + 1):
            cell = ws.cell(r, c)
            cell.border = BOX
            cell.font = Font(size=10)
            if fills and c in fills:
                cell.fill = PatternFill("solid", fgColor=fills[c])


def money(ws, r0, r1, cols, fmt="#,##0"):
    for r in range(r0, r1 + 1):
        for c in cols:
            ws.cell(r, c).number_format = fmt


def note(ws, row, text, cols):
    c = ws.cell(row, 1, text)
    c.font = Font(size=9, color="606060")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(cols, 2))


# ═══════════════════════════════════════════════════════════════
# 1. 설정 — 낙찰률·요율은 이 시트 «한 곳» 에만 적습니다
# ═══════════════════════════════════════════════════════════════
def sheet_config(wb):
    ws = wb.create_sheet("설정")
    brand(ws, 4)
    title(ws, "설  정", 4)
    ws.freeze_panes = None
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 52

    rows = [
        ("■ 공사 정보", "", ""),
        ("공 사 명", "", "예) ○○지구 배수개선사업"),
        ("계약번호", "", ""),
        ("발주기관", "", ""),
        ("계약일자", "", ""),
        ("변경 차수", "1차", ""),
        ("작성일자", "", ""),
        ("", "", ""),
        ("■ 금액 (여기 숫자가 모든 계산의 뿌리입니다)", "", ""),
        ("계약금액 (원)", 0, "부가세 포함 계약금액"),
        ("예정가격 (원)", 0, "개찰 때 확정된 예정가격"),
        ("낙찰률 (%)", 0, "=계약금액÷예정가격×100. 아래 «자동» 값을 참고해 적으세요"),
        ("  (자동 계산)", "", "예정가격이 0이면 비어 있습니다"),
        ("", "", ""),
        ("■ 원가 요율 (발주기관 기준을 그대로 넣으세요)", "", ""),
        ("일반관리비율 (%)", 6.0, "순공사원가 × 요율"),
        ("이윤율 (%)", 15.0, "(노무비+경비+일반관리비) × 요율 — 재료비는 뺍니다"),
        ("부가가치세율 (%)", 10.0, ""),
    ]
    r = HEAD
    for label, val, hint in rows:
        if label.startswith("■"):
            c = ws.cell(r, 2, label)
            c.font = Font(size=11, bold=True, color=BLUE)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
            c.fill = PatternFill("solid", fgColor=LITE)
        elif label:
            a = ws.cell(r, 2, label)
            a.font = Font(size=10, bold=True)
            a.fill = PatternFill("solid", fgColor=GRAY)
            a.border = BOX
            b = ws.cell(r, 3, val)
            b.border = BOX
            b.font = Font(size=10)
            b.fill = PatternFill("solid", fgColor=YELLOW)
            d = ws.cell(r, 4, hint)
            d.font = Font(size=9, color="808080")
        r += 1

    # 자리 (1-based). rows 목록 순서에 맞춰 계산합니다.
    C = {}
    r = HEAD
    for label, _, _ in rows:
        C[label.strip()] = r
        r += 1
    ws["C%d" % C["계약금액 (원)"]].number_format = "#,##0"
    ws["C%d" % C["예정가격 (원)"]].number_format = "#,##0"
    ws["C%d" % C["낙찰률 (%)"]].number_format = "0.000"
    auto = C["(자동 계산)"]
    ws["C%d" % auto] = ("=IF(C{p}>0,ROUND(C{k}/C{p}*100,5),\"\")"
                        .format(p=C["예정가격 (원)"], k=C["계약금액 (원)"]))
    ws["C%d" % auto].number_format = "0.00000"
    ws["C%d" % auto].fill = PatternFill("solid", fgColor=GRAY)

    note(ws, r + 1,
         "노랑 칸이 사람이 넣는 자리입니다. 회색 칸은 수식이 채웁니다. "
         "낙찰률은 «자동 계산» 값을 보고 적으세요 — 공고서에 적힌 값이 있으면 그 값이 우선입니다. "
         "신규 비목 단가에 이 낙찰률이 곱해지므로, 여기가 비어 있으면 증감대비표가 0을 냅니다.", 4)
    return ws, C


# ═══════════════════════════════════════════════════════════════
# 2. 단가마스터 — 여기 단가 한 칸이 끝까지 흘러갑니다
# ═══════════════════════════════════════════════════════════════
def sheet_master(wb):
    ws = wb.create_sheet("단가마스터")
    cols = ["코드", "품 명", "규 격", "단위", "재료비단가", "노무비단가",
            "경비단가", "합계단가", "적용시점", "산출 근거", "비고"]
    w = [12, 26, 16, 7, 13, 13, 13, 13, 11, 24, 14]
    brand(ws, len(cols))
    title(ws, "단 가 마 스 터", len(cols))
    header(ws, cols, w)
    r1 = HEAD + N_MASTER - 1
    grid(ws, HEAD, r1, len(cols),
         fills={1: YELLOW, 2: YELLOW, 3: YELLOW, 4: YELLOW, 5: YELLOW,
                6: YELLOW, 7: YELLOW, 8: GRAY, 9: YELLOW, 10: YELLOW, 11: YELLOW})
    for r in range(HEAD, r1 + 1):
        ws.cell(r, 8, "=IF(A{r}=\"\",\"\",N(E{r})+N(F{r})+N(G{r}))".format(r=r))
    money(ws, HEAD, r1, [5, 6, 7, 8])
    dv = DataValidation(type="list", formula1='"당초,변경당시"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("I%d:I%d" % (HEAD, r1))
    ws.cell(HEAD, 1, "예) 토공-001")
    ws.cell(HEAD, 1).font = Font(size=10, color="A6A6A6", italic=True)
    return ws


# ═══════════════════════════════════════════════════════════════
# 3. 일위대가 — 왼쪽은 쌓기, 오른쪽(N~R)은 일위별 단가 요약
# ═══════════════════════════════════════════════════════════════
def sheet_ilwi(wb):
    ws = wb.create_sheet("일위대가")
    cols = ["일위번호", "구성코드", "품 명", "규 격", "단위", "수량",
            "재료비단가", "노무비단가", "경비단가", "재료비금액", "노무비금액", "경비금액"]
    w = [12, 12, 24, 14, 7, 9, 12, 12, 12, 13, 13, 13]
    brand(ws, 20)
    title(ws, "일 위 대 가 표", 12)
    header(ws, cols, w)
    r1 = HEAD + N_ILWI - 1
    grid(ws, HEAD, r1, len(cols),
         fills={1: YELLOW, 2: YELLOW, 3: GRAY, 4: GRAY, 5: GRAY, 6: YELLOW,
                7: GRAY, 8: GRAY, 9: GRAY, 10: GRAY, 11: GRAY, 12: GRAY})
    M = "단가마스터!$A$%d:$H$%d" % (HEAD, HEAD + N_MASTER - 1)
    for r in range(HEAD, r1 + 1):
        ws.cell(r, 3, '=IF($B{r}="","",IFERROR(VLOOKUP($B{r},{M},2,FALSE),"⚠️단가마스터에 없음"))'.format(r=r, M=M))
        ws.cell(r, 4, '=IF($B{r}="","",IFERROR(VLOOKUP($B{r},{M},3,FALSE),""))'.format(r=r, M=M))
        ws.cell(r, 5, '=IF($B{r}="","",IFERROR(VLOOKUP($B{r},{M},4,FALSE),""))'.format(r=r, M=M))
        for k, idx in ((7, 5), (8, 6), (9, 7)):
            ws.cell(r, k, '=IF($B{r}="","",IFERROR(VLOOKUP($B{r},{M},{i},FALSE),0))'
                    .format(r=r, M=M, i=idx))
        for k, src in ((10, "G"), (11, "H"), (12, "I")):
            ws.cell(r, k, '=IF($B{r}="","",N($F{r})*N(${s}{r}))'.format(r=r, s=src))
    money(ws, HEAD, r1, [7, 8, 9, 10, 11, 12])

    # ── 오른쪽 요약: 일위번호별 단가 ──
    sw = ["일위번호", "일위 품명", "단위", "재료비계", "노무비계", "경비계", "합계단가"]
    for i, (nm, wd) in enumerate(zip(sw, [12, 22, 7, 13, 13, 13, 13]), start=14):
        c = ws.cell(3, i, nm)
        c.font = Font(size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="548235")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BOX
        ws.column_dimensions[get_column_letter(i)].width = wd
    s1 = HEAD + N_ILWI_SUM - 1
    grid(ws, HEAD, s1, 20, fills=None)
    for r in range(HEAD, s1 + 1):
        for c in range(14, 21):
            ws.cell(r, c).border = BOX
            ws.cell(r, c).fill = PatternFill("solid", fgColor=(YELLOW if c in (14, 15, 16) else GRAY))
    A = "$A${0}:$A${1}".format(HEAD, r1)
    for r in range(HEAD, s1 + 1):
        for c, src in ((17, "J"), (18, "K"), (19, "L")):
            ws.cell(r, c, '=IF($N{r}="","",SUMIF({A},$N{r},${s}${a}:${s}${b}))'
                    .format(r=r, A=A, s=src, a=HEAD, b=r1))
        ws.cell(r, 20, '=IF($N{r}="","",N(Q{r})+N(R{r})+N(S{r}))'.format(r=r))
    money(ws, HEAD, s1, [17, 18, 19, 20])
    ws.cell(2, 14, "▶ 오른쪽 = 일위번호별 단가 (내역서가 여기를 봅니다)")
    ws.cell(2, 14).font = Font(size=10, bold=True, color="548235")
    return ws


# ═══════════════════════════════════════════════════════════════
# 4. 수량산출서 — 산출식을 남기고 수량은 자동
# ═══════════════════════════════════════════════════════════════
def sheet_suryang(wb):
    ws = wb.create_sheet("수량산출서")
    cols = ["구분", "공종코드", "부위·위치", "가로", "세로", "높이", "개소",
            "수량", "단위", "도면번호", "비고"]
    w = [9, 13, 26, 9, 9, 9, 8, 11, 7, 13, 18]
    brand(ws, len(cols))
    title(ws, "수 량 산 출 서", len(cols))
    header(ws, cols, w)
    r1 = HEAD + N_SURYANG - 1
    grid(ws, HEAD, r1, len(cols),
         fills={1: YELLOW, 2: YELLOW, 3: YELLOW, 4: YELLOW, 5: YELLOW,
                6: YELLOW, 7: YELLOW, 8: GRAY, 9: YELLOW, 10: YELLOW, 11: YELLOW})
    for r in range(HEAD, r1 + 1):
        # 빈 칸은 1 로 봅니다 — «가로×개소» 처럼 두 칸만 쓰는 경우가 많습니다
        ws.cell(r, 8, '=IF($B{r}="","",IF(COUNT(D{r}:G{r})=0,"",'
                      'ROUND(IF(D{r}="",1,D{r})*IF(E{r}="",1,E{r})*'
                      'IF(F{r}="",1,F{r})*IF(G{r}="",1,G{r}),3)))'.format(r=r))
        ws.cell(r, 8).number_format = "#,##0.###"
    dv = DataValidation(type="list", formula1='"당초,변경"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("A%d:A%d" % (HEAD, r1))
    note(ws, r1 + 2,
         "구분에 «당초» 또는 «변경» 을 고르면 내역서의 «수량(산출서)» 칸이 자동으로 합쳐집니다. "
         "쓰지 않는 칸(세로·높이 등)은 비워 두세요 — 1 로 봅니다.", len(cols))
    return ws


# ═══════════════════════════════════════════════════════════════
# 5·6. 당초내역 / 변경내역 — 구조가 같아야 증감이 줄끼리 맞습니다
# ═══════════════════════════════════════════════════════════════
def sheet_naeyeok(wb, name, gubun):
    ws = wb.create_sheet(name)
    cols = ["번호", "공종코드", "품 명", "규 격", "단위",
            "수량(산출서)", "수량(직접)", "적용수량",
            "재료비단가", "노무비단가", "경비단가", "합계단가", "금 액", "비고"]
    w = [6, 13, 26, 15, 7, 12, 11, 11, 12, 12, 12, 12, 15, 16]
    brand(ws, len(cols))
    title(ws, "%s 산 출 내 역 서" % ("당 초" if gubun == "당초" else "변 경"), len(cols))
    header(ws, cols, w)
    r1 = HEAD + N_ITEM - 1
    grid(ws, HEAD, r1, len(cols),
         fills={1: YELLOW, 2: YELLOW, 3: GRAY, 4: GRAY, 5: GRAY, 6: GRAY,
                7: YELLOW, 8: GRAY, 9: GRAY, 10: GRAY, 11: GRAY, 12: GRAY,
                13: GRAY, 14: YELLOW})
    M = "단가마스터!$A$%d:$H$%d" % (HEAD, HEAD + N_MASTER - 1)
    I = "일위대가!$N$%d:$T$%d" % (HEAD, HEAD + N_ILWI_SUM - 1)
    SG = "수량산출서!$A$%d:$A$%d" % (HEAD, HEAD + N_SURYANG - 1)
    SC = "수량산출서!$B$%d:$B$%d" % (HEAD, HEAD + N_SURYANG - 1)
    SQ = "수량산출서!$H$%d:$H$%d" % (HEAD, HEAD + N_SURYANG - 1)
    for r in range(HEAD, r1 + 1):
        # 이름·규격·단위 — 단가마스터에 없으면 일위대가 요약에서 찾습니다
        ws.cell(r, 3, '=IF($B{r}="","",IFERROR(VLOOKUP($B{r},{M},2,FALSE),'
                      'IFERROR(VLOOKUP($B{r},{I},2,FALSE),"⚠️단가 없음")))'.format(r=r, M=M, I=I))
        ws.cell(r, 4, '=IF($B{r}="","",IFERROR(VLOOKUP($B{r},{M},3,FALSE),""))'.format(r=r, M=M))
        ws.cell(r, 5, '=IF($B{r}="","",IFERROR(VLOOKUP($B{r},{M},4,FALSE),'
                      'IFERROR(VLOOKUP($B{r},{I},3,FALSE),"")))'.format(r=r, M=M, I=I))
        ws.cell(r, 6, '=IF($B{r}="","",SUMIFS({SQ},{SG},"{g}",{SC},$B{r}))'
                .format(r=r, SQ=SQ, SG=SG, SC=SC, g=gubun))
        ws.cell(r, 8, '=IF($B{r}="","",IF($G{r}<>"",$G{r},N($F{r})))'.format(r=r))
        for c, mi, ii in ((9, 5, 4), (10, 6, 5), (11, 7, 6)):
            ws.cell(r, c, '=IF($B{r}="","",IFERROR(VLOOKUP($B{r},{M},{mi},FALSE),'
                          'IFERROR(VLOOKUP($B{r},{I},{ii},FALSE),0)))'
                    .format(r=r, M=M, I=I, mi=mi, ii=ii))
        ws.cell(r, 12, '=IF($B{r}="","",N($I{r})+N($J{r})+N($K{r}))'.format(r=r))
        ws.cell(r, 13, '=IF($B{r}="","",ROUND(N($H{r})*N($L{r}),0))'.format(r=r))
    money(ws, HEAD, r1, [9, 10, 11, 12, 13])
    for r in range(HEAD, r1 + 1):
        ws.cell(r, 6).number_format = "#,##0.###"
        ws.cell(r, 7).number_format = "#,##0.###"
        ws.cell(r, 8).number_format = "#,##0.###"
    # 합계
    t = r1 + 1
    ws.cell(t, 2, "합  계").font = Font(size=11, bold=True)
    ws.cell(t, 13, "=SUM(M{a}:M{b})".format(a=HEAD, b=r1))
    for c in (2, 13):
        ws.cell(t, c).fill = PatternFill("solid", fgColor=LITE)
        ws.cell(t, c).font = Font(size=11, bold=True)
        ws.cell(t, c).border = BOX
    ws.cell(t, 13).number_format = "#,##0"
    note(ws, t + 2,
         "공종코드만 넣으면 품명·규격·단위·단가가 단가마스터(없으면 일위대가)에서 따라옵니다. "
         "수량은 수량산출서에서 자동으로 합쳐지고, 직접 넣고 싶으면 «수량(직접)» 칸에 적으세요 — 그 값이 이깁니다. "
         "⚠️ 변경내역은 당초내역과 «같은 줄에 같은 코드» 를 두세요. 증감대비표가 줄끼리 맞춥니다 "
         "(신규 비목은 당초내역의 그 줄을 비워 두면 됩니다).", len(cols))
    return ws


# ═══════════════════════════════════════════════════════════════
# 7. 증감대비표 — 규정 단가가 수식으로 들어 있는 «심장»
# ═══════════════════════════════════════════════════════════════
def sheet_jeunggam(wb, C):
    ws = wb.create_sheet("증감대비표")
    cols = ["번호", "공종코드", "품 명", "규 격", "단위",
            "당초수량", "계약단가", "당초금액",
            "변경수량", "당시단가", "증감수량",
            "구분", "협의(Y)", "예정가격단가", "적용단가", "증감금액", "단가 근거"]
    w = [6, 13, 24, 14, 7, 11, 12, 14, 11, 12, 11, 9, 9, 13, 13, 15, 34]
    brand(ws, len(cols))
    title(ws, "증 감 대 비 표", len(cols))
    header(ws, cols, w)
    r1 = HEAD + N_ITEM - 1
    grid(ws, HEAD, r1, len(cols),
         fills={1: GRAY, 2: GRAY, 3: GRAY, 4: GRAY, 5: GRAY, 6: GRAY, 7: GRAY,
                8: GRAY, 9: GRAY, 10: GRAY, 11: GRAY, 12: GRAY, 13: YELLOW,
                14: YELLOW, 15: GREEN, 16: GREEN, 17: GRAY})
    RATE = "('설정'!$C${r}/100)".format(r=C["낙찰률 (%)"])
    for i, r in enumerate(range(HEAD, r1 + 1)):
        s = HEAD + i                      # 당초·변경 내역의 같은 줄
        ws.cell(r, 1, '=IF($B{r}="","",ROW()-{h}+1)'.format(r=r, h=HEAD))
        ws.cell(r, 2, '=IF(변경내역!$B{s}<>"",변경내역!$B{s},당초내역!$B{s})'.format(s=s))
        ws.cell(r, 3, '=IF($B{r}="","",IF(변경내역!$C{s}<>"",변경내역!$C{s},당초내역!$C{s}))'.format(r=r, s=s))
        ws.cell(r, 4, '=IF($B{r}="","",IF(변경내역!$D{s}<>"",변경내역!$D{s},당초내역!$D{s}))'.format(r=r, s=s))
        ws.cell(r, 5, '=IF($B{r}="","",IF(변경내역!$E{s}<>"",변경내역!$E{s},당초내역!$E{s}))'.format(r=r, s=s))
        ws.cell(r, 6, '=IF($B{r}="","",N(당초내역!$H{s}))'.format(r=r, s=s))
        ws.cell(r, 7, '=IF($B{r}="","",N(당초내역!$L{s}))'.format(r=r, s=s))
        ws.cell(r, 8, '=IF($B{r}="","",N(당초내역!$M{s}))'.format(r=r, s=s))
        ws.cell(r, 9, '=IF($B{r}="","",N(변경내역!$H{s}))'.format(r=r, s=s))
        ws.cell(r, 10, '=IF($B{r}="","",N(변경내역!$L{s}))'.format(r=r, s=s))
        ws.cell(r, 11, '=IF($B{r}="","",N($I{r})-N($F{r}))'.format(r=r))
        # 구분 — 당초에 없던 것이면 신규
        ws.cell(r, 12, '=IF($B{r}="","",IF(당초내역!$B{s}="","신규",'
                       'IF($K{r}<0,"감소",IF($K{r}>0,"증가","-"))))'.format(r=r, s=s))
        # 적용단가 — 국가계약법 시행령 제65조
        ws.cell(r, 15,
                '=IF($B{r}="","",'
                'IF($L{r}="감소",$G{r},'
                'IF($L{r}="증가",IF($M{r}="Y",ROUND(($J{r}+$J{r}*{RT})/2,0),'
                'IF(AND($N{r}<>"",$G{r}>$N{r}),$N{r},$G{r})),'
                'IF($L{r}="신규",IF($M{r}="Y",ROUND(($J{r}+$J{r}*{RT})/2,0),'
                'ROUND($J{r}*{RT},0)),0))))'.format(r=r, RT=RATE))
        ws.cell(r, 16, '=IF($B{r}="","",ROUND(N($K{r})*N($O{r}),0))'.format(r=r))
        ws.cell(r, 17,
                '=IF($B{r}="","",'
                'IF($L{r}="감소","감소분 → 계약단가",'
                'IF(AND($M{r}="Y",OR($L{r}="증가",$L{r}="신규")),'
                '"발주기관 요구 → 협의 (불성립 시 (당시단가+당시단가×낙찰률)÷2)",'
                'IF($L{r}="증가",IF(AND($N{r}<>"",$G{r}>$N{r}),'
                '"증가분 → 계약단가>예정가격단가 이므로 예정가격단가",'
                '"증가분 → 계약단가 (낙찰률을 곱하지 않습니다)"),'
                'IF($L{r}="신규","신규비목 → 당시단가 × 낙찰률","")))))'.format(r=r))
    money(ws, HEAD, r1, [7, 8, 10, 14, 15, 16])
    for r in range(HEAD, r1 + 1):
        for c in (6, 9, 11):
            ws.cell(r, c).number_format = "#,##0.###"
        ws.cell(r, 12).alignment = Alignment(horizontal="center")
        ws.cell(r, 13).alignment = Alignment(horizontal="center")
        ws.cell(r, 17).font = Font(size=9, color="606060")
    t = r1 + 1
    for lab, col, f in (("증가 소계", 16, '=SUMIF($L${a}:$L${b},"증가",$P${a}:$P${b})'),
                        ("감소 소계", 16, '=SUMIF($L${a}:$L${b},"감소",$P${a}:$P${b})'),
                        ("신규 소계", 16, '=SUMIF($L${a}:$L${b},"신규",$P${a}:$P${b})'),
                        ("증감 합계", 16, '=SUM($P${a}:$P${b})')):
        ws.cell(t, 14, lab).font = Font(size=11, bold=True)
        ws.cell(t, 14).fill = PatternFill("solid", fgColor=LITE)
        ws.cell(t, 14).border = BOX
        c = ws.cell(t, col, f.format(a=HEAD, b=r1))
        c.font = Font(size=11, bold=True)
        c.number_format = "#,##0"
        c.fill = PatternFill("solid", fgColor=LITE)
        c.border = BOX
        t += 1
    ws.cell(t, 14, "변경 후 계약금액").font = Font(size=11, bold=True)
    ws.cell(t, 14).fill = PatternFill("solid", fgColor=GREEN)
    ws.cell(t, 14).border = BOX
    c = ws.cell(t, 16, "='설정'!$C${k}+P{s}".format(k=C["계약금액 (원)"], s=t - 1))
    c.font = Font(size=12, bold=True, color=BLUE)
    c.number_format = "#,##0"
    c.fill = PatternFill("solid", fgColor=GREEN)
    c.border = BOX
    note(ws, t + 2,
         "이 표는 손으로 채우는 곳이 «협의(Y)» 와 «예정가격단가» 둘뿐입니다. 나머지는 전부 자동입니다.\n"
         "· 발주기관이 요구해서 늘어난 분이면 협의 칸에 Y 를 넣으세요 — 단가가 협의식으로 바뀝니다.\n"
         "· 예정가격단가는 «아는 경우에만» 적으세요. 계약단가가 그보다 높으면 예정가격단가가 적용됩니다.\n"
         "· 증가분에는 낙찰률을 곱하지 않습니다. 낙찰률은 «신규 비목» 에만 곱합니다.", len(cols))
    return ws, t


# ═══════════════════════════════════════════════════════════════
# 8. 공종별집계 — 코드 앞 2~3글자로 묶습니다
# ═══════════════════════════════════════════════════════════════
def sheet_group(wb):
    ws = wb.create_sheet("공종별집계")
    cols = ["대공종 코드", "대공종 명", "당초금액", "증감금액", "변경 후 금액", "증감률(%)"]
    w = [14, 26, 18, 18, 18, 12]
    brand(ws, len(cols))
    title(ws, "공 종 별 집 계", len(cols))
    header(ws, cols, w)
    r1 = HEAD + N_GROUP - 1
    grid(ws, HEAD, r1, len(cols),
         fills={1: YELLOW, 2: YELLOW, 3: GRAY, 4: GRAY, 5: GRAY, 6: GRAY})
    a, b = HEAD, HEAD + N_ITEM - 1
    for r in range(HEAD, r1 + 1):
        ws.cell(r, 3, '=IF($A{r}="","",SUMPRODUCT((LEFT(당초내역!$B${a}:$B${b},LEN($A{r}))=$A{r})'
                      '*N(당초내역!$M${a}:$M${b})))'.format(r=r, a=a, b=b))
        ws.cell(r, 4, '=IF($A{r}="","",SUMPRODUCT((LEFT(증감대비표!$B${a}:$B${b},LEN($A{r}))=$A{r})'
                      '*N(증감대비표!$P${a}:$P${b})))'.format(r=r, a=a, b=b))
        ws.cell(r, 5, '=IF($A{r}="","",N($C{r})+N($D{r}))'.format(r=r))
        ws.cell(r, 6, '=IF(OR($A{r}="",N($C{r})=0),"",ROUND($D{r}/$C{r}*100,2))'.format(r=r))
    money(ws, HEAD, r1, [3, 4, 5])
    for r in range(HEAD, r1 + 1):
        ws.cell(r, 6).number_format = "0.00"
    t = r1 + 1
    ws.cell(t, 2, "합  계").font = Font(size=11, bold=True)
    for c, col in ((3, "C"), (4, "D"), (5, "E")):
        x = ws.cell(t, c, "=SUM({0}{1}:{0}{2})".format(col, HEAD, r1))
        x.font = Font(size=11, bold=True)
        x.number_format = "#,##0"
        x.fill = PatternFill("solid", fgColor=LITE)
        x.border = BOX
    ws.cell(t, 2).fill = PatternFill("solid", fgColor=LITE)
    ws.cell(t, 2).border = BOX
    note(ws, t + 2,
         "대공종 코드에 «토공» 처럼 코드 앞부분을 적으면 그 글자로 시작하는 모든 줄을 묶습니다. "
         "예: 코드가 «토공-001, 토공-002» 이면 «토공» 이라고 적으세요.", len(cols))
    return ws


# ═══════════════════════════════════════════════════════════════
# 9. 원가계산서 — 당초 / 변경 후 두 열
# ═══════════════════════════════════════════════════════════════
def sheet_wonga(wb, C):
    ws = wb.create_sheet("원가계산서")
    cols = ["구 분", "비  목", "산출 근거", "당 초", "변경 후", "증 감"]
    w = [12, 24, 34, 18, 18, 18]
    brand(ws, len(cols))
    title(ws, "공 사 원 가 계 산 서", len(cols))
    header(ws, cols, w)
    a, b = HEAD, HEAD + N_ITEM - 1
    GEN = "'설정'!$C$%d" % C["일반관리비율 (%)"]
    PRF = "'설정'!$C$%d" % C["이윤율 (%)"]
    VAT = "'설정'!$C$%d" % C["부가가치세율 (%)"]

    def sp(sheet, qty, price):
        return "SUMPRODUCT(N({s}!${q}${a}:${q}${b}),N({s}!${p}${a}:${p}${b}))".format(
            s=sheet, q=qty, p=price, a=a, b=b)

    rows = [
        ("순공사원가", "재료비", "Σ(적용수량 × 재료비단가)", sp("당초내역", "H", "I"), sp("변경내역", "H", "I")),
        ("", "노무비", "Σ(적용수량 × 노무비단가)", sp("당초내역", "H", "J"), sp("변경내역", "H", "J")),
        ("", "경  비", "Σ(적용수량 × 경비단가)", sp("당초내역", "H", "K"), sp("변경내역", "H", "K")),
        ("", "순공사원가 계", "재료비+노무비+경비", "SUM(D{0}:D{1})", "SUM(E{0}:E{1})"),
        ("간접", "일반관리비", "순공사원가 × 요율", None, None),
        ("", "이  윤", "(노무비+경비+일반관리비) × 요율 — 재료비 제외", None, None),
        ("합계", "공급가액", "순공사원가+일반관리비+이윤", None, None),
        ("", "부가가치세", "공급가액 × 세율", None, None),
        ("", "총 공사비", "공급가액+부가세", None, None),
    ]
    r = HEAD
    idx = {}
    for g, item, basis, d, e in rows:
        ws.cell(r, 1, g).font = Font(size=10, bold=True)
        ws.cell(r, 2, item).font = Font(size=10, bold=True)
        ws.cell(r, 3, basis).font = Font(size=9, color="606060")
        idx[item] = r
        for c in range(1, 7):
            ws.cell(r, c).border = BOX
            if c in (4, 5, 6):
                ws.cell(r, c).fill = PatternFill("solid", fgColor=GRAY)
                ws.cell(r, c).number_format = "#,##0"
        r += 1
    R = idx
    ws.cell(R["재료비"], 4, "=" + rows[0][3]); ws.cell(R["재료비"], 5, "=" + rows[0][4])
    ws.cell(R["노무비"], 4, "=" + rows[1][3]); ws.cell(R["노무비"], 5, "=" + rows[1][4])
    ws.cell(R["경  비"], 4, "=" + rows[2][3]); ws.cell(R["경  비"], 5, "=" + rows[2][4])
    for col in ("D", "E"):
        ws.cell(R["순공사원가 계"], 4 if col == "D" else 5,
                "=SUM({c}{a}:{c}{b})".format(c=col, a=R["재료비"], b=R["경  비"]))
        ws.cell(R["일반관리비"], 4 if col == "D" else 5,
                "=ROUND({c}{s}*{g}/100,0)".format(c=col, s=R["순공사원가 계"], g=GEN))
        ws.cell(R["이  윤"], 4 if col == "D" else 5,
                "=ROUND(({c}{n}+{c}{e}+{c}{g})*{p}/100,0)".format(
                    c=col, n=R["노무비"], e=R["경  비"], g=R["일반관리비"], p=PRF))
        ws.cell(R["공급가액"], 4 if col == "D" else 5,
                "={c}{s}+{c}{g}+{c}{p}".format(c=col, s=R["순공사원가 계"],
                                               g=R["일반관리비"], p=R["이  윤"]))
        ws.cell(R["부가가치세"], 4 if col == "D" else 5,
                "=ROUND({c}{s}*{v}/100,0)".format(c=col, s=R["공급가액"], v=VAT))
        ws.cell(R["총 공사비"], 4 if col == "D" else 5,
                "={c}{s}+{c}{v}".format(c=col, s=R["공급가액"], v=R["부가가치세"]))
    for item, rr in R.items():
        ws.cell(rr, 6, "=E{r}-D{r}".format(r=rr))
    for c in range(1, 7):
        ws.cell(R["총 공사비"], c).fill = PatternFill("solid", fgColor=GREEN)
        ws.cell(R["총 공사비"], c).font = Font(size=11, bold=True, color=BLUE)
    note(ws, r + 1,
         "요율은 «설정» 시트에서 한 번만 고치면 여기가 따라옵니다. "
         "이윤은 노무비+경비+일반관리비에만 곱합니다 — 재료비는 넣지 않습니다. "
         "«변경 후» 열은 변경내역의 단가로 잡은 값입니다. 발주기관에 내는 조정금액은 "
         "증감대비표의 «변경 후 계약금액» 을 쓰세요(규정 단가가 적용된 값입니다).", 6)
    return ws, R


# ═══════════════════════════════════════════════════════════════
# 10. 검증시트 — 사람이 놓치는 것을 수식이 잡습니다
# ═══════════════════════════════════════════════════════════════
def sheet_check(wb, C, jg_total_row):
    ws = wb.create_sheet("검증시트")
    cols = ["번호", "무엇을 보는가", "값", "판정", "틀렸을 때 어떻게 되나"]
    w = [6, 40, 20, 12, 56]
    brand(ws, len(cols))
    title(ws, "검 증 시 트", len(cols))
    header(ws, cols, w)
    a, b = HEAD, HEAD + N_ITEM - 1
    ma, mb = HEAD, HEAD + N_MASTER - 1
    RATE = "'설정'!$C$%d" % C["낙찰률 (%)"]
    KUM = "'설정'!$C$%d" % C["계약금액 (원)"]
    SUM_JG = "증감대비표!$P$%d" % (jg_total_row - 1)

    checks = [
        ("낙찰률이 들어 있는가",
         "={R}".format(R=RATE), "=IF(N(C{r})>0,\"✅\",\"⚠️\")",
         "비어 있으면 신규 비목 단가가 0 이 됩니다. 조정금액이 통째로 틀립니다.", "0.000"),
        ("당초내역에서 단가를 못 찾은 줄",
         '=COUNTIF(당초내역!$C${a}:$C${b},"⚠️*")'.format(a=a, b=b),
         '=IF(N(C{r})=0,"✅","⚠️")',
         "단가마스터·일위대가 어디에도 없는 코드입니다. 그 줄 금액이 0 으로 잡힙니다.", "#,##0"),
        ("변경내역에서 단가를 못 찾은 줄",
         '=COUNTIF(변경내역!$C${a}:$C${b},"⚠️*")'.format(a=a, b=b),
         '=IF(N(C{r})=0,"✅","⚠️")',
         "위와 같습니다. 증감금액이 실제보다 적게 나옵니다.", "#,##0"),
        ("당초·변경의 줄이 서로 어긋난 개수",
         '=SUMPRODUCT((당초내역!$B${a}:$B${b}<>"")*(변경내역!$B${a}:$B${b}<>"")'
         '*(당초내역!$B${a}:$B${b}<>변경내역!$B${a}:$B${b}))'.format(a=a, b=b),
         '=IF(N(C{r})=0,"✅","⚠️")',
         "같은 줄에 다른 공종이 놓였습니다. 증감이 엉뚱한 항목끼리 계산됩니다.", "#,##0"),
        ("단가마스터에 코드가 겹치는 개수",
         '=SUMPRODUCT((단가마스터!$A${ma}:$A${mb}<>"")*'
         '(COUNTIF(단가마스터!$A${ma}:$A${mb},단가마스터!$A${ma}:$A${mb}&"")>1))'.format(ma=ma, mb=mb),
         '=IF(N(C{r})=0,"✅","⚠️")',
         "VLOOKUP 은 «맨 위 하나» 만 씁니다. 뒤에 적은 단가는 조용히 무시됩니다.", "#,##0"),
        ("적용수량이 음수인 줄 (당초+변경)",
         '=SUMPRODUCT((당초내역!$H${a}:$H${b}<0)*1)+SUMPRODUCT((변경내역!$H${a}:$H${b}<0)*1)'
         .format(a=a, b=b),
         '=IF(N(C{r})=0,"✅","⚠️")',
         "수량을 음수로 넣어 감소를 표현하면 안 됩니다. 변경수량을 줄여서 적으세요.", "#,##0"),
        ("증가분에 낙찰률이 잘못 곱해진 줄",
         '=SUMPRODUCT((증감대비표!$L${a}:$L${b}="증가")*(증감대비표!$M${a}:$M${b}<>"Y")'
         '*(증감대비표!$O${a}:$O${b}<>증감대비표!$G${a}:$G${b})'
         '*(증감대비표!$N${a}:$N${b}=""))'.format(a=a, b=b),
         '=IF(N(C{r})=0,"✅","⚠️")',
         "증가분은 계약단가 그대로여야 합니다. 낙찰률을 곱하면 조정액이 깎입니다.", "#,##0"),
        ("신규 비목에 낙찰률이 빠진 줄",
         '=SUMPRODUCT((증감대비표!$L${a}:$L${b}="신규")*(증감대비표!$M${a}:$M${b}<>"Y")'
         '*(증감대비표!$J${a}:$J${b}>0)'
         '*(증감대비표!$O${a}:$O${b}=증감대비표!$J${a}:$J${b}))'.format(a=a, b=b),
         '=IF(N(C{r})=0,"✅","⚠️")',
         "신규 비목은 «당시단가 × 낙찰률» 입니다. 안 곱하면 발주기관이 깎습니다.", "#,##0"),
        ("당초내역 합계와 설정의 계약금액 차이",
         '=IF(N({K})=0,"",당초내역!$M${t}-{K})'.format(K=KUM, t=HEAD + N_ITEM),
         '=IF(OR(C{r}="",ABS(N(C{r}))<=1),"✅","⚠️")',
         "당초내역이 계약금액과 안 맞습니다. 빠진 항목이나 부가세 처리를 확인하세요.", "#,##0"),
        ("증감 합계 (규정 단가 적용)",
         "={S}".format(S=SUM_JG), '="—"',
         "이 값이 발주기관에 내는 조정금액입니다.", "#,##0"),
        ("변경내역 − 당초내역 (단순 차)",
         '=당초내역!$M${t}*-1+변경내역!$M${t}'.format(t=HEAD + N_ITEM), '="—"',
         "위 «증감 합계» 와 다른 것이 정상입니다 — 규정 단가(증가는 계약단가, "
         "신규는 낙찰률 적용)를 쓰기 때문입니다.", "#,##0"),
        ("증감률 (%)",
         '=IF(N({K})=0,"",ROUND({S}/{K}*100,2))'.format(K=KUM, S=SUM_JG),
         '=IF(C{r}="","—",IF(ABS(N(C{r}))<30,"✅","⚠️"))',
         "30% 를 넘으면 발주기관 내부 절차가 길어집니다. 차수를 나누는 편이 빠를 수 있습니다.", "0.00"),
    ]
    r = HEAD
    for i, (what, val, verdict, why, fmt) in enumerate(checks, start=1):
        ws.cell(r, 1, i)
        ws.cell(r, 2, what)
        ws.cell(r, 3, val)
        ws.cell(r, 3).number_format = fmt
        ws.cell(r, 4, verdict.format(r=r))
        ws.cell(r, 5, why)
        for c in range(1, 6):
            ws.cell(r, c).border = BOX
            ws.cell(r, c).font = Font(size=10)
            ws.cell(r, c).alignment = Alignment(vertical="center", wrap_text=(c == 5))
        ws.cell(r, 3).fill = PatternFill("solid", fgColor=GRAY)
        ws.cell(r, 4).fill = PatternFill("solid", fgColor=LITE)
        ws.cell(r, 4).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(r, 4).font = Font(size=13, bold=True)
        ws.cell(r, 5).font = Font(size=9, color="606060")
        ws.row_dimensions[r].height = 30
        r += 1
    ws.cell(r + 1, 2, "⚠️ 가 하나라도 있으면 그 줄의 «틀렸을 때 어떻게 되나» 를 읽고 고친 뒤 제출하세요.")
    ws.cell(r + 1, 2).font = Font(size=10, bold=True, color="C00000")
    return ws


# ═══════════════════════════════════════════════════════════════
# 11. 제출서식 — 발주기관에 내는 1장
# ═══════════════════════════════════════════════════════════════
def sheet_submit(wb, C, jg_total_row):
    ws = wb.create_sheet("제출서식")
    brand(ws, 6)
    title(ws, "설 계 변 경  (계약금액 조정) 요 청 서", 6)
    for col, wd in zip("ABCDEF", [4, 18, 26, 20, 20, 16]):
        ws.column_dimensions[col].width = wd
    SUM_JG = "증감대비표!$P$%d" % (jg_total_row - 1)
    AFTER = "증감대비표!$P$%d" % jg_total_row
    KUM = "'설정'!$C$%d" % C["계약금액 (원)"]

    rows = [
        ("공 사 명", "='설정'!$C$%d" % C["공 사 명"], ""),
        ("계약번호", "='설정'!$C$%d" % C["계약번호"], ""),
        ("발주기관", "='설정'!$C$%d" % C["발주기관"], ""),
        ("변경 차수", "='설정'!$C$%d" % C["변경 차수"], ""),
        ("", "", ""),
        ("당초 계약금액", "=" + KUM, "#,##0"),
        ("증  가", '=SUMIF(증감대비표!$L$%d:$L$%d,"증가",증감대비표!$P$%d:$P$%d)'
         % (HEAD, HEAD + N_ITEM - 1, HEAD, HEAD + N_ITEM - 1), "#,##0"),
        ("감  소", '=SUMIF(증감대비표!$L$%d:$L$%d,"감소",증감대비표!$P$%d:$P$%d)'
         % (HEAD, HEAD + N_ITEM - 1, HEAD, HEAD + N_ITEM - 1), "#,##0"),
        ("신 규 비 목", '=SUMIF(증감대비표!$L$%d:$L$%d,"신규",증감대비표!$P$%d:$P$%d)'
         % (HEAD, HEAD + N_ITEM - 1, HEAD, HEAD + N_ITEM - 1), "#,##0"),
        ("증감 합계", "=" + SUM_JG, "#,##0"),
        ("증감률", '=IF(N(%s)=0,"",ROUND(%s/%s*100,2)&" %%")' % (KUM, SUM_JG, KUM), ""),
        ("변경 후 계약금액", "=" + AFTER, "#,##0"),
        ("", "", ""),
        ("변경 사유", "", ""),
        ("근거 서류", "", ""),
    ]
    r = HEAD
    for label, val, fmt in rows:
        if label:
            a = ws.cell(r, 2, label)
            a.font = Font(size=11, bold=True)
            a.fill = PatternFill("solid", fgColor=GRAY)
            a.border = BOX
            a.alignment = Alignment(horizontal="center", vertical="center")
            c = ws.cell(r, 3, val)
            c.border = BOX
            c.font = Font(size=11)
            if fmt:
                c.number_format = fmt
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
            for cc in (4, 5):
                ws.cell(r, cc).border = BOX
            ws.row_dimensions[r].height = 22
            if label in ("변경 사유", "근거 서류"):
                c.fill = PatternFill("solid", fgColor=YELLOW)
                c.alignment = Alignment(vertical="top", wrap_text=True)
                ws.row_dimensions[r].height = 54
            elif label == "변경 후 계약금액":
                for cc in (2, 3, 4, 5):
                    ws.cell(r, cc).fill = PatternFill("solid", fgColor=GREEN)
                c.font = Font(size=13, bold=True, color=BLUE)
        r += 1
    r += 1
    ws.cell(r, 2, "위와 같이 설계변경에 따른 계약금액 조정을 요청합니다.").font = Font(size=11)
    r += 3
    for i, who in enumerate(("현장대리인", "대표자")):
        ws.cell(r, 2 + i * 2, who).font = Font(size=11, bold=True)
        ws.cell(r, 3 + i * 2, "(서명 또는 인)").font = Font(size=10, color="808080")
    ws.print_area = "A1:F%d" % (r + 1)
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    return ws


# ═══════════════════════════════════════════════════════════════
# 안내 시트 + 조립
# ═══════════════════════════════════════════════════════════════
def sheet_guide(wb):
    ws = wb.create_sheet("사용법", 0)
    brand(ws, 3)
    title(ws, "이 파일을 쓰는 법", 3)
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 86
    steps = [
        ("① 설정", "공사명·계약금액·예정가격·낙찰률·요율을 넣습니다. 낙찰률은 «자동 계산» 값을 참고하세요."),
        ("② 단가마스터", "쓸 단가를 코드와 함께 넣습니다. 여기 단가 한 칸을 고치면 끝까지 다시 계산됩니다."),
        ("③ 일위대가", "복합 공종은 여기서 쌓습니다. 오른쪽(N열~)의 일위별 단가를 내역서가 가져다 씁니다."),
        ("④ 수량산출서", "구분(당초/변경)과 공종코드를 넣고 가로·세로·높이·개소를 적으면 수량이 자동입니다."),
        ("⑤ 당초내역", "공종코드만 넣으면 품명·규격·단가가 따라오고, 수량은 수량산출서에서 합쳐집니다."),
        ("⑥ 변경내역", "당초내역과 «같은 줄에 같은 코드». 신규 비목은 당초 쪽 줄을 비워 두면 «신규» 로 잡힙니다."),
        ("⑦ 증감대비표", "전부 자동입니다. 손댈 곳은 «협의(Y)» 와 «예정가격단가» 둘뿐입니다."),
        ("⑧ 공종별집계", "대공종 코드(예: 토공)를 적으면 그 글자로 시작하는 줄을 묶어 보여줍니다."),
        ("⑨ 원가계산서", "당초 / 변경 후를 나란히 냅니다. 요율은 설정에서 한 번만 고칩니다."),
        ("⑩ 검증시트", "제출 전에 여기부터 보세요. ⚠️ 가 있으면 그 줄을 고치고 냅니다."),
        ("⑪ 제출서식", "발주기관에 내는 1장. 인쇄 영역이 잡혀 있습니다."),
    ]
    r = HEAD
    for a, b in steps:
        x = ws.cell(r, 2, a)
        x.font = Font(size=11, bold=True, color=BLUE)
        x.fill = PatternFill("solid", fgColor=LITE)
        x.border = BOX
        y = ws.cell(r, 3, b)
        y.font = Font(size=10)
        y.border = BOX
        y.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 26
        r += 1
    r += 1
    ws.cell(r, 2, "칸 색").font = Font(size=11, bold=True)
    for lab, color in (("노랑 = 사람이 넣는 칸", YELLOW), ("회색 = 수식이 채우는 칸(건드리지 마세요)", GRAY),
                       ("초록 = 결과", GREEN)):
        r += 1
        c = ws.cell(r, 3, lab)
        c.fill = PatternFill("solid", fgColor=color)
        c.border = BOX
        c.font = Font(size=10)
    r += 2
    ws.cell(r, 2, "계약금액 조정 단가 기준").font = Font(size=11, bold=True, color="C00000")
    for lab in ("감소분 → 계약단가",
                "증가분 → 계약단가 (계약단가가 예정가격단가보다 높으면 예정가격단가)",
                "신규 비목 → 설계변경 «당시» 단가 × 낙찰률",
                "발주기관 요구 → 협의, 불성립 시 (당시단가 + 당시단가×낙찰률) ÷ 2",
                "※ 증가분에는 낙찰률을 곱하지 않습니다 — 가장 흔한 실수입니다."):
        r += 1
        c = ws.cell(r, 3, lab)
        c.font = Font(size=10)
        c.border = BOX
    r += 2
    ws.cell(r, 2, "⚠️ 참고").font = Font(size=10, bold=True)
    ws.cell(r, 3, "국가계약법 시행령 등을 바탕으로 만들었지만 계약서 특수조건과 발주기관 판단이 "
                  "우선합니다. 금액이 큰 건은 전문가 검토를 받으세요. 1행의 K-건설맵 표시는 지우고 쓰셔도 됩니다.")
    ws.cell(r, 3).font = Font(size=9, color="606060")
    ws.cell(r, 3).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 40
    return ws


def build():
    wb = Workbook()
    wb.remove(wb.active)
    _, C = sheet_config(wb)
    sheet_master(wb)
    sheet_ilwi(wb)
    sheet_suryang(wb)
    sheet_naeyeok(wb, "당초내역", "당초")
    sheet_naeyeok(wb, "변경내역", "변경")
    _, jg_total = sheet_jeunggam(wb, C)
    sheet_group(wb)
    sheet_wonga(wb, C)
    sheet_check(wb, C, jg_total)
    sheet_submit(wb, C, jg_total)
    sheet_guide(wb)          # 맨 앞으로
    order = ["사용법", "설정", "단가마스터", "일위대가", "수량산출서", "당초내역",
             "변경내역", "증감대비표", "공종별집계", "원가계산서", "검증시트", "제출서식"]
    wb._sheets = [wb[n] for n in order]
    wb.active = 0
    if not os.path.isdir(OUT):
        os.makedirs(OUT)
    p = os.path.join(OUT, NAME)
    wb.save(p)
    return p


if __name__ == "__main__":
    p = build()
    print("  ✅ 설계변경 통합 엑셀 → %s (%.0f KB · 시트 12장)"
          % (p, os.path.getsize(p) / 1024))
